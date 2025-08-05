from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.db import models
import time
import re
import json
from .forms import LoginForm, RegistrationForm, UserPasswordResetForm, UserSetPasswordForm, UserPasswordChangeForm, CursoCapacitacionForm, PostForm, CommentForm, BlogPostForm, EvaluacionForm, CalificacionForm, EntregaForm, TicketSoporteForm, ComentarioTicketForm, TicketSoporteAdminForm
from django.contrib.auth import logout
from django.contrib.auth import views as auth_views
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from .models import RegistrationLink, Post, Comment, Curso, BlogPost, InscripcionCurso, Evaluacion, Calificacion, Entrega, TicketSoporte, ClasificacionTicket, SubclasificacionTicket, Rubrica, CriterioRubrica, Esperable, ResultadoRubrica, PuntajeCriterio, CustomUser, ComentarioTicket
from django.http import JsonResponse, HttpResponseRedirect
from django.core.mail import send_mail
from django.contrib import messages
from django.conf import settings
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.contrib.auth.tokens import default_token_generator
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils.html import strip_tags
from django.utils.safestring import mark_safe
import requests
import logging
import traceback
from datetime import datetime, timedelta
import jwt
from django.db.models import Avg, Min, Max, Count, Sum
from django.conf import settings
from decimal import Decimal
from django.db.models import Prefetch
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
import io

logger = logging.getLogger(__name__)

User = get_user_model()  # Obtener el modelo de usuario personalizado

def generate_registration_link(request):
    if request.user.is_superuser:
        new_link = RegistrationLink.objects.create(creator=request.user)
        return JsonResponse({'link': str(new_link.uuid)})
    else:
        return JsonResponse({'error': 'Unauthorized'}, status=403)

def enviar_correo_activacion(request, user):
    """
    Envía un correo de activación de cuenta al usuario registrado.
    """
    token = default_token_generator.make_token(user)
    uid = urlsafe_base64_encode(force_bytes(user.pk))
    activation_link = request.build_absolute_uri(
        reverse('activate-account', kwargs={'uidb64': uid, 'token': token})
    )
    subject = 'Activa tu cuenta en Gryphos Consulting'
    message = f"""
Hola {user.username},

Gracias por registrarte en Gryphos Consulting.

Por favor, haz clic en el siguiente enlace para activar tu cuenta:
{activation_link}

Si no solicitaste este registro, puedes ignorar este correo.

Saludos,
El equipo de Gryphos
"""
    send_mail(
        subject=subject,
        message=message,
        from_email=settings.DEFAULT_FROM_EMAIL if hasattr(settings, 'DEFAULT_FROM_EMAIL') else 'contacto@gryphos.cl',
        recipient_list=[user.email],
        fail_silently=False,
    )


def registration(request):
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            # Verificar si el email ya está registrado
            email = form.cleaned_data['email']
            
            # Verificar si MULTI_ACCOUNT_SIGNIN está habilitado
            if not getattr(settings, 'MULTI_ACCOUNT_SIGNIN', False):
                # Si MULTI_ACCOUNT_SIGNIN es False, no permitir múltiples cuentas con el mismo email
                if User.objects.filter(email=email).exists():
                    messages.error(request, 'Ya existe una cuenta registrada con este correo electrónico.')
                    return render(request, 'accounts/sign-up.html', {'form': form})
            else:
                # Si MULTI_ACCOUNT_SIGNIN es True, permitir múltiples cuentas pero mostrar advertencia
                existing_users = User.objects.filter(email=email)
                if existing_users.exists():
                    messages.warning(request, f'Ya existe una cuenta registrada con este correo electrónico. Se creará una nueva cuenta.')
            
            user = form.save(commit=False)
            user.is_active = False  # Usuario inactivo hasta confirmar correo
            user.save()
            enviar_correo_activacion(request, user)
            messages.success(request, '¡Registro exitoso! Revisa tu correo para activar tu cuenta.')
            return redirect('/accounts/login/')
        else:
            logger.error("Registration failed!")
            logger.error("Form errors:", form.errors)
            for field_name, errors in form.errors.items():
                logger.error(f"Field {field_name}: {errors}")
    else:
        form = RegistrationForm()
    context = {'form': form}
    return render(request, 'accounts/sign-up.html', context)


def activate_account(request, uidb64, token):
    """
    Vista para activar la cuenta de usuario desde el link de activación.
    """
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None

    if user is not None and default_token_generator.check_token(user, token):
        user.is_active = True
        user.save()
        messages.success(request, '¡Cuenta activada correctamente! Ya puedes iniciar sesión.')
        return redirect('/accounts/login/')
    else:
        messages.error(request, 'El enlace de activación no es válido o ha expirado.')
        return render(request, 'accounts/activation_invalid.html')

class UserLoginView(auth_views.LoginView):
    template_name = 'accounts/sign-in.html'
    form_class = LoginForm
    success_url = '/portal-cliente/'
    
    def get(self, request, *args, **kwargs):
        # Verificar si la sesión expiró
        if request.GET.get('expired'):
            messages.warning(request, 'Tu sesión ha expirado por inactividad. Por favor, inicia sesión nuevamente.')
        return super().get(request, *args, **kwargs)
    
    def form_valid(self, form):
        """Override para agregar logging al login exitoso"""
        response = super().form_valid(form)
        
        # Log del login exitoso
        user = form.get_user()
        logger.info(f"Login exitoso para usuario: {user.username}")
        logger.info(f"Session key después del login: {self.request.session.session_key}")
        logger.info(f"Usuario autenticado: {self.request.user.is_authenticated}")
        
        # Inicializar la última actividad en la sesión
        self.request.session['last_activity'] = time.time()
        self.request.session.save()
        
        return response
    
    def form_invalid(self, form):
        """Override para agregar logging al login fallido"""
        username = form.cleaned_data.get('username', 'N/A') if form.cleaned_data else 'N/A'
        logger.warning(f"Login fallido para usuario: {username}")
        logger.warning(f"Errores del formulario: {form.errors}")
        
        # Intentar identificar el problema específico
        if 'username' in form.errors:
            logger.warning("Error en el campo username")
        if 'password' in form.errors:
            logger.warning("Error en el campo password")
        
        return super().form_invalid(form)

class UserPasswordResetView(auth_views.PasswordResetView):
    template_name = 'accounts/password_reset.html'
    form_class = UserPasswordResetForm
    email_template_name = 'registration/password_reset_email.html'
    subject_template_name = 'registration/password_reset_subject.txt'
    
    def form_valid(self, form):
        email = form.cleaned_data['email']
        logger.info(f"Iniciando proceso de recuperación de contraseña para: {email}")
        
        try:
            # Verificar si el email existe en la base de datos
            if not User.objects.filter(email=email).exists():
                logger.warning(f"Intento de recuperación de contraseña para email inexistente: {email}")
                messages.error(self.request, 'No existe una cuenta registrada con este correo electrónico.')
                return self.form_invalid(form)
            
            # Si el email existe, proceder con el envío
            logger.info(f"Enviando correo de recuperación a: {email}")
            
            # Obtener el usuario
            user = User.objects.get(email=email)
            logger.info(f"Usuario encontrado: {user.username}")
            
            # Generar token y URL
            from django.contrib.auth.tokens import default_token_generator
            from django.utils.http import urlsafe_base64_encode
            from django.utils.encoding import force_bytes
            from django.template.loader import render_to_string
            from django.core.mail import send_mail
            import traceback
            
            token = default_token_generator.make_token(user)
            uid = urlsafe_base64_encode(force_bytes(user.pk))
            
            logger.info(f"Token generado: {token}")
            logger.info(f"UID generado: {uid}")
            
            # Construir URL de reset
            reset_url = f"https://gryphos.cl/accounts/password-reset-confirm/{uid}/{token}/"
            logger.info(f"URL de reset generada: {reset_url}")
            
            # Preparar contexto para las plantillas
            context = {
                'user': user,
                'protocol': 'https',
                'domain': 'gryphos.cl',
                'uid': uid,
                'token': token,
            }
            
            # Renderizar plantilla HTML
            html_message = render_to_string('registration/password_reset_email.html', context)
            logger.info("✅ Plantilla HTML renderizada correctamente")
            
            # Renderizar asunto
            subject = render_to_string('registration/password_reset_subject.txt', context).strip()
            logger.info(f"✅ Asunto renderizado: {subject}")
            
            # Mensaje en texto plano (fallback)
            message = f"""
            Hola {user.get_full_name() or user.username},
            
            Has solicitado restablecer tu contraseña en Gryphos Consulting.
            
            Para continuar con el proceso, visita el siguiente enlace:
            {reset_url}
            
            Si no solicitaste este cambio, puedes ignorar este correo.
            
            Este enlace es válido por 15 minutos.
            
            Saludos,
            Equipo Gryphos Consulting
            """
            
            logger.info("Enviando correo con send_mail...")
            
            try:
                # Enviar correo usando send_mail directamente
                result = send_mail(
                    subject=subject,
                    message=message,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[user.email],
                    html_message=html_message,
                    fail_silently=False,
                )
            except Exception as e:
                logger.error(f"Error al enviar correo de recuperación a {email}: {str(e)}")
                logger.error(f"Tipo de error: {type(e).__name__}")
                
                logger.error(f"Traceback completo: {traceback.format_exc()}")
                messages.error(self.request, 'Error al enviar el correo de recuperación. Por favor, intenta nuevamente.')
                return self.form_invalid(form)

            logger.info(f"✅ Correo enviado exitosamente con send_mail. Resultado: {result}")
            
            # Redirigir a la página de confirmación
            from django.shortcuts import redirect
            return redirect('password_reset_done')
            
        except Exception as e:
            logger.error(f"Error al enviar correo de recuperación a {email}: {str(e)}")
            logger.error(f"Tipo de error: {type(e).__name__}")
            import traceback
            logger.error(f"Traceback completo: {traceback.format_exc()}")
            messages.error(self.request, 'Error al enviar el correo de recuperación. Por favor, intenta nuevamente.')
            return self.form_invalid(form)
    
    def form_invalid(self, form):
        logger.warning(f"Formulario de recuperación de contraseña inválido: {form.errors}")
        return super().form_invalid(form)


class UserPasswordResetConfirmView(auth_views.PasswordResetConfirmView):
    template_name = 'accounts/password_reset_confirm.html'
    form_class = UserSetPasswordForm
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Siempre mostrar el formulario, validar en POST
        context['token_valid'] = True
        return context
    
    def get(self, request, *args, **kwargs):
        try:
            # Siempre mostrar el formulario
            return super().get(request, *args, **kwargs)
        except Exception as e:
            logger.error(f"Error en vista de confirmación de contraseña: {str(e)}")
            messages.error(request, 'Error al procesar la solicitud. Por favor, intenta nuevamente.')
            return self.render_to_response(self.get_context_data())
    
    def post(self, request, *args, **kwargs):
        try:
            logger.info(f"Iniciando proceso de restablecimiento de contraseña")
            logger.info(f"Token válido: {self.validlink}")
            
            if not self.validlink:
                logger.warning("Intento de restablecimiento con token inválido")
                messages.error(request, 'El enlace de recuperación de contraseña no es válido o ha expirado.')
                return self.render_to_response(self.get_context_data())
            
            # Obtener el usuario
            user = self.get_user(self.kwargs['uidb64'])
            if user:
                logger.info(f"Usuario encontrado para restablecimiento: {user.username}")
            else:
                logger.error("Usuario no encontrado para restablecimiento")
                messages.error(request, 'Usuario no encontrado.')
                return self.render_to_response(self.get_context_data())
            
            # Procesar el formulario
            form = self.get_form()
            if form.is_valid():
                logger.info("Formulario válido, cambiando contraseña")
                form.save()
                logger.info(f"Contraseña cambiada exitosamente para usuario: {user.username}")
                messages.success(request, 'Tu contraseña ha sido restablecida exitosamente. Ya puedes iniciar sesión con tu nueva contraseña.')
                
                # Redirigir a la página de login
                from django.shortcuts import redirect
                return redirect('login')
            else:
                logger.warning(f"Formulario inválido: {form.errors}")
                return self.render_to_response(self.get_context_data())
                
        except Exception as e:
            logger.error(f"Error en POST de confirmación de contraseña: {str(e)}")
            logger.error(f"Tipo de error: {type(e).__name__}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            messages.error(request, 'Error al restablecer la contraseña. Por favor, intenta nuevamente.')
            return self.render_to_response(self.get_context_data())


class UserPasswordChangeView(auth_views.PasswordChangeView):
    template_name = 'accounts/password_change.html'
    form_class = UserPasswordChangeForm

def user_logout_view(request):
    # Limpiar todos los mensajes de la sesión antes de hacer logout
    from django.contrib import messages
    storage = messages.get_messages(request)
    storage.used = True  # Marcar todos los mensajes como usados para limpiarlos
    
    logout(request)
    return redirect('/')

def index(request):
    # Si el usuario está logueado, redirigir al portal de miembros
    if request.user.is_authenticated:
        return redirect('user_space')
    
    # Obtener todos los cursos activos para mostrar en la página principal
    cursos_activos = Curso.objects.filter(activo=True).order_by('nombre')
    context = {
        'cursos_activos': cursos_activos,
    }
    return render(request, 'pages/index.html', context)

def que_hacemos(request):
    # Si el usuario está logueado, redirigir al portal de miembros
    if request.user.is_authenticated:
        return redirect('user_space')
    return render(request, 'pages/que-hacemos.html')

def quienes_somos(request):
    # Si el usuario está logueado, redirigir al portal de miembros
    if request.user.is_authenticated:
        return redirect('user_space')
    
    if request.method == 'POST':
        # Obtener datos del formulario
        nombre = request.POST.get('nombre', '')
        email = request.POST.get('email', '')
        mensaje = request.POST.get('message', '')
        
        if nombre and email and mensaje:
            try:
                # Enviar correo
                subject = f'Nuevo mensaje de contacto de {nombre}'
                message = f"""
                Nuevo mensaje de contacto recibido:
                
                Nombre: {nombre}
                Email: {email}
                Mensaje: {mensaje}
                
                Este mensaje fue enviado desde el formulario de contacto de la página "Quiénes somos".
                """
                logger.info("Sending email to: contacto@gryphos.cl")
                # Enviar a la dirección de contacto de Gryphos
                send_mail(
                    subject=subject,
                    message=message,
                    from_email=settings.DEFAULT_FROM_EMAIL if hasattr(settings, 'DEFAULT_FROM_EMAIL') else 'contacto@gryphos.cl',
                    recipient_list=['contacto@gryphos.cl'],
                    fail_silently=False,
                )
                logger.info("Email sent successfully!")
                
                messages.success(request, '¡Mensaje enviado exitosamente! Nos pondremos en contacto contigo pronto.')
                
            except Exception as e:
                messages.error(request, 'Error al enviar el mensaje. Por favor, intenta nuevamente.')
                logger.error(f"Error enviando email: {e}")
        else:
            messages.error(request, 'Por favor, completa todos los campos del formulario.')
    
    return render(request, 'pages/quienes-somos.html')

@login_required
def portal_cliente(request):
    logger.info(f"Acceso al portal del cliente - Usuario: {request.user.username}")
    
    try:
        cursos_usuario = request.user.cursos.all()
        logger.debug(f"Usuario {request.user.username} tiene {cursos_usuario.count()} cursos")
        
        context = {
            'cursos_usuario': cursos_usuario,
            'request': request
        }
        
        response = render(request, 'pages/portal-cliente.html', context)
        logger.info(f"Portal del cliente renderizado exitosamente para {request.user.username}")
        return response
        
    except Exception as e:
        logger.error(f"Error en portal_cliente para usuario {request.user.username}: {str(e)}")
        logger.error(f"Tipo de error: {type(e).__name__}")
        logger.error(f"Traceback completo: {traceback.format_exc()}")
        raise


def debug_session(request):
    """
    Vista de debug para verificar el estado de la sesión
    """
    if not settings.DEBUG:
        return HttpResponse("Debug solo disponible en modo desarrollo", status=403)
    
    # Información básica de la sesión
    session_info = {
        'session_key': request.session.session_key,
        'session_exists': request.session.session_key is not None,
        'session_data': dict(request.session),
        'session_modified': request.session.modified,
    }
    
    # Información del usuario
    user_info = {
        'user_authenticated': request.user.is_authenticated,
        'user': str(request.user),
        'user_id': request.user.id if request.user.is_authenticated else None,
        'user_backend': getattr(request.user, '_auth_user_backend', None),
    }
    
    # Información de cookies
    cookie_info = {
        'sessionid': request.COOKIES.get('sessionid'),
        'csrftoken': request.COOKIES.get('csrftoken'),
        'all_cookies': dict(request.COOKIES),
    }
    
    # Información de headers
    headers_info = {
        'user_agent': request.META.get('HTTP_USER_AGENT'),
        'referer': request.META.get('HTTP_REFERER'),
        'host': request.META.get('HTTP_HOST'),
    }
    
    debug_info = {
        'session': session_info,
        'user': user_info,
        'cookies': cookie_info,
        'headers': headers_info,
        'timestamp': str(datetime.now()),
    }
    
    return JsonResponse(debug_info, json_dumps_params={'indent': 2})


@login_required
def mailcow_sso(request):
    """
    Vista para crear sesión SSO en mailcow y redirigir a Soho
    """
    try:
        # Obtener configuración de mailcow
        mailcow_config = getattr(settings, 'MAILCOW_CONFIG', {})
        sso_config = getattr(settings, 'MAILCOW_SSO_CONFIG', {})
        
        if not mailcow_config.get('API_KEY'):
            logger.error("API_KEY no configurado en MAILCOW_CONFIG")
            messages.error(request, "Error de configuración del sistema de correo.")
            return redirect('user_space')
        
        # Verificar que el usuario tenga email
        if not request.user.email:
            logger.error(f"Usuario {request.user.username} no tiene email configurado")
            messages.error(request, "Tu cuenta no tiene un correo electrónico configurado.")
            return redirect('user_space')
        
        # Preparar datos para la API de mailcow (endpoint para usuarios alumnos)
        sso_data = {
            'username': request.user.email,
            'password': 'sso_session',  # Mailcow maneja la autenticación internamente
            'active': 1,
            'quota': 1024,  # 1GB para alumnos
            'domain': request.user.email.split('@')[1] if '@' in request.user.email else 'gryphos.cl',
            'local_part': request.user.email.split('@')[0] if '@' in request.user.email else request.user.username,
            'sogo_access': 1,  # Habilitar acceso a SOGo
            'sogo_admin': 0,   # No es admin (es alumno)
            'sogo_admin_username': request.user.email,
            'sogo_admin_password': 'sso_session'
        }
        
        # Construir URL del endpoint SSO
        api_url = f"{mailcow_config['API_URL']}{mailcow_config['SSO_ENDPOINT']}"
        
        # Headers para la API
        headers = {
            'X-API-Key': mailcow_config['API_KEY'],
            'Content-Type': 'application/json'
        }
        
        logger.info(f"Creando sesión SSO para usuario {request.user.username} ({request.user.email})")
        
        # Hacer petición a la API de mailcow
        response = requests.post(
            api_url,
            json=sso_data,
            headers=headers,
            timeout=30
        )
        
        if response.status_code == 200:
            logger.info(f"Sesión SSO creada exitosamente para {request.user.username}")
            
            # Redirigir a Soho
            soho_url = mailcow_config.get('SOHO_URL', 'https://mail.gryphos.cl/SOGo')
            return HttpResponseRedirect(soho_url)
        else:
            logger.error(f"Error en API de mailcow: {response.status_code} - {response.text}")
            messages.error(request, "Error al acceder al correo electrónico. Inténtalo de nuevo.")
            return redirect('user_space')
        
    except requests.exceptions.RequestException as e:
        logger.error(f"Error de conexión con mailcow para usuario {request.user.username}: {str(e)}")
        messages.error(request, "Error de conexión con el servidor de correo. Inténtalo de nuevo.")
        return redirect('user_space')
    except Exception as e:
        logger.error(f"Error en mailcow_sso para usuario {request.user.username}: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        messages.error(request, "Error al acceder al correo electrónico. Inténtalo de nuevo.")
        return redirect('user_space')


def enviar_correo_instrucciones_pago(request, inscripcion):
    """
    Función para enviar correo con instrucciones de pago al interesado
    """
    from django.template.loader import render_to_string
    from django.utils.html import strip_tags
    
    curso_nombre = inscripcion.curso.nombre
    fecha_solicitud = inscripcion.fecha_solicitud.strftime('%d/%m/%Y %H:%M')
    
    # Formatear el precio con separación de miles usando punto (formato chileno)
    curso_precio_formateado = f"{inscripcion.curso.precio:,.0f}".replace(",", ".")
    
    # Verificar si el usuario ya tiene una cuenta activa
    usuario_existente = User.objects.filter(email=inscripcion.correo_contacto, is_active=True).first()
    
    # Generar URLs para recuperación de contraseña y soporte
    password_reset_url = request.build_absolute_uri('/accounts/password_reset/')
    support_url = request.build_absolute_uri('/portal-cliente/')
    
    # Renderizar el template HTML
    html_message = render_to_string('emails/instrucciones_pago.html', {
        'nombre_interesado': inscripcion.nombre_interesado,
        'nombre_empresa': inscripcion.nombre_empresa,
        'curso_nombre': curso_nombre,
        'fecha_solicitud': fecha_solicitud,
        'correo_contacto': inscripcion.correo_contacto,
        'curso_precio': inscripcion.curso.precio,
        'curso_precio_formateado': curso_precio_formateado,
        'dias_plazo_pago': inscripcion.curso.dias_plazo_pago,
        'usuario_existente': usuario_existente,
        'password_reset_url': password_reset_url,
        'support_url': support_url,
    })

    logger.debug(f"HTML message: {html_message}")
    logger.debug(f"Preparando envío de correo electrónico a {inscripcion.correo_contacto}")
    
    # Crear versión de texto plano
    plain_message = strip_tags(html_message)
    
    subject = f'Instrucciones de Pago - Curso {curso_nombre} - Gryphos Consulting'
    
    try:
        send_mail(
            subject=subject,
            message=plain_message,
            from_email=settings.DEFAULT_FROM_EMAIL if hasattr(settings, 'DEFAULT_FROM_EMAIL') else 'contacto@gryphos.cl',
            recipient_list=[inscripcion.correo_contacto],
            html_message=html_message,
            fail_silently=False,
        )
        logger.info(f"Correo de instrucciones de pago enviado exitosamente a {inscripcion.correo_contacto}")
        return True

    except Exception as e:
        logger.error(f"Error enviando correo de instrucciones de pago: {e}")
        return False


def enviar_correo_bienvenida(request, user, password_temp, curso_nombre):
    """
    Función para enviar correo de bienvenida con credenciales al nuevo usuario
    """
    from django.template.loader import render_to_string
    from django.utils.html import strip_tags
    from .models import Curso
    
    # Obtener el curso completo para acceder a sus videollamadas
    try:
        curso = Curso.objects.get(nombre=curso_nombre)
        videollamadas = curso.videollamadas.filter(activa=True).order_by('dia_semana', 'hora_inicio')
        
        # Formatear información de horarios
        horarios_info = []
        for videollamada in videollamadas:
            horarios_info.append({
                'dia': videollamada.get_dia_semana_display(),
                'hora_inicio': videollamada.hora_inicio.strftime('%H:%M'),
                'hora_fin': videollamada.hora_fin.strftime('%H:%M'),
                'descripcion': videollamada.descripcion
            })
    except Curso.DoesNotExist:
        curso = None
        horarios_info = []
    
    # Generar URLs
    login_url = request.build_absolute_uri('/accounts/login/')
    change_password_url = request.build_absolute_uri('/accounts/password_change/')
    
    # Renderizar el template HTML
    html_message = render_to_string('emails/bienvenida_usuario.html', {
        'nombre_usuario': user.get_full_name() or user.username,
        'username': user.username,
        'password_temp': password_temp,
        'email': user.email,
        'curso_nombre': curso_nombre,
        'curso': curso,
        'horarios_info': horarios_info,
        'login_url': login_url,
        'change_password_url': change_password_url,
    })
    
    # Crear versión de texto plano
    plain_message = strip_tags(html_message)
    
    subject = f'¡Bienvenido a Gryphos Consulting! - Acceso a {curso_nombre}'
    
    try:
        send_mail(
            subject=subject,
            message=plain_message,
            from_email=settings.DEFAULT_FROM_EMAIL if hasattr(settings, 'DEFAULT_FROM_EMAIL') else 'contacto@gryphos.cl',
            recipient_list=[user.email],
            html_message=html_message,
            fail_silently=False,
        )
        logger.info(f"Correo de bienvenida enviado exitosamente a {user.email}")
        return True
    except Exception as e:
        logger.error(f"Error enviando correo de bienvenida: {e}")
        return False


def enviar_correo_inscripcion(nombre_interesado, nombre_empresa, telefono_contacto, correo_contacto, curso_interes):
    """
    Función para enviar correo de inscripción al curso de capacitación
    """
    # Obtener el nombre legible del curso
    try:
        curso = Curso.objects.get(id=curso_interes)
        curso_nombre = curso.nombre
    except Curso.DoesNotExist:
        curso_nombre = curso_interes
    
    # Manejar teléfono opcional
    telefono_info = telefono_contacto if telefono_contacto else "No proporcionado"
    
    subject = f'Nueva inscripción para curso de capacitación - {nombre_interesado}'
    message = f"""
    Se ha recibido una nueva inscripción para el curso de capacitación:
    
    Nombre del interesado: {nombre_interesado}
    Nombre de la empresa: {nombre_empresa}
    Teléfono de contacto: {telefono_info}
    Correo de contacto: {correo_contacto}
    Curso de interés: {curso_nombre}
    
    Este mensaje fue enviado desde el formulario de inscripción de cursos de capacitación.
    """
    
    try:
        send_mail(
            subject=subject,
            message=message,
            from_email=settings.DEFAULT_FROM_EMAIL if hasattr(settings, 'DEFAULT_FROM_EMAIL') else 'contacto@gryphos.cl',
            recipient_list=['contacto@gryphos.cl'],
            fail_silently=False,
        )
        return True
    except Exception as e:
        logger.error(f"Error enviando correo de inscripción: {e}")
        return False


def inscripcion_curso(request):
    """
    Vista para el formulario de inscripción al curso de capacitación
    """
    from .models import InscripcionCurso
    
    # Si el usuario está logueado, usar proceso simplificado
    if request.user.is_authenticated:
        curso_id = request.GET.get('curso')
        if curso_id:
            try:
                curso = Curso.objects.get(id=curso_id, activo=True)
                
                # Verificar si ya está inscrito en este curso
                # Verificar si tiene acceso al curso (ya pagó)
                tiene_acceso = request.user.cursos.filter(id=curso_id).exists()
                
                # Verificar si tiene inscripción pendiente
                inscripcion_existente = InscripcionCurso.objects.filter(
                    usuario_creado=request.user,
                    curso=curso,
                    estado__in=['pendiente', 'confirmada', 'en_proceso']
                ).first()
                
                if tiene_acceso or inscripcion_existente:
                    messages.warning(request, f'Ya estás inscrito en el curso "{curso.nombre}".')
                    return redirect('cursos_list')
                
                # Crear inscripción automática para usuario logueado
                inscripcion = InscripcionCurso.objects.create(
                    nombre_interesado=f"{request.user.first_name} {request.user.last_name}".strip(),
                    nombre_empresa=request.user.company_name or "No especificada",
                    telefono_contacto=request.user.phone_number or "",
                    correo_contacto=request.user.email,
                    curso=curso,
                    estado='pendiente',
                    usuario_creado=request.user
                )
                
                # Enviar correo con instrucciones de pago
                if enviar_correo_instrucciones_pago(request, inscripcion):
                    # También enviar notificación al administrador
                    enviar_correo_inscripcion(
                        inscripcion.nombre_interesado,
                        inscripcion.nombre_empresa,
                        inscripcion.telefono_contacto,
                        inscripcion.correo_contacto,
                        curso.nombre
                    )
                    
                    messages.success(request, f'¡Solicitud de inscripción enviada exitosamente para el curso "{curso.nombre}"! Revisa tu correo para las instrucciones de pago.')
                    logger.info(f"Inscripción automática enviada para usuario {request.user.username} al curso {curso.nombre}")
                    return redirect('cursos_list')
                else:
                    messages.error(request, 'Hubo un error al enviar las instrucciones de pago. Por favor, intenta nuevamente.')
                    inscripcion.delete()
                    
            except Curso.DoesNotExist:
                messages.error(request, 'El curso seleccionado no existe.')
                return redirect('cursos_list')
            except Exception as e:
                messages.error(request, f'Error al procesar la inscripción: {str(e)}')
                return redirect('cursos_list')
        else:
            messages.warning(request, 'No se especificó ningún curso para inscribirse.')
            return redirect('cursos_list')
    
    # Proceso normal para usuarios no logueados
    # Obtener el curso pre-seleccionado desde la URL
    curso_id = request.GET.get('curso')
    curso_seleccionado = None
    
    if curso_id:
        try:
            curso_seleccionado = Curso.objects.get(id=curso_id, activo=True)
        except Curso.DoesNotExist:
            messages.warning(request, 'El curso seleccionado no está disponible.')
    
    if request.method == 'POST':
        form = CursoCapacitacionForm(request.POST)
        if form.is_valid():
            nombre_interesado = form.cleaned_data['nombre_interesado']
            nombre_empresa = form.cleaned_data['nombre_empresa']
            telefono_contacto = form.cleaned_data['telefono_contacto']
            correo_contacto = form.cleaned_data['correo_contacto']
            curso_interes = form.cleaned_data['curso_interes']
            
            try:
                # Obtener el curso
                curso = Curso.objects.get(id=curso_interes)
                
                # Crear la inscripción en la base de datos
                inscripcion = InscripcionCurso.objects.create(
                    nombre_interesado=nombre_interesado,
                    nombre_empresa=nombre_empresa,
                    telefono_contacto=telefono_contacto,
                    correo_contacto=correo_contacto,
                    curso=curso,
                    estado='pendiente'
                )
                
                # Enviar correo con instrucciones de pago al interesado
                if enviar_correo_instrucciones_pago(request, inscripcion):
                    # También enviar notificación al administrador
                    enviar_correo_inscripcion(nombre_interesado, nombre_empresa, telefono_contacto, correo_contacto, curso_interes)
                    
                    messages.success(request, '¡Inscripción enviada exitosamente! Revisa tu correo para las instrucciones de pago. Si no lo encuentras, revisa tu spam.')
                    logger.info(f"Inscripción enviada exitosamente a {inscripcion.correo_contacto}")
                    return redirect('inscripcion-curso')
                else:
                    # Si falla el envío del correo, eliminar la inscripción
                    messages.error(request, 'Hubo un error al enviar las instrucciones de pago. Por favor, intenta nuevamente.')
                    logger.error(f"Error al enviar las instrucciones de pago: de {inscripcion.correo_contacto} : {e}")
                    inscripcion.delete()
                    
            except Curso.DoesNotExist:
                messages.error(request, 'El curso seleccionado no existe.')
            except Exception as e:
                messages.error(request, f'Error al procesar la inscripción: {str(e)}')
        else:
            logger.error(f"Error al procesar la inscripción: {form.errors}")
    else:
        # Si hay un curso pre-seleccionado, inicializar el formulario con ese curso
        if curso_seleccionado:
            form = CursoCapacitacionForm(initial={'curso_interes': curso_seleccionado.id})
        else:
            form = CursoCapacitacionForm()
    
    context = {
        'form': form,
        'curso_seleccionado': curso_seleccionado
    }
    return render(request, 'pages/inscripcion-curso.html', context)


def forum_list(request):
    """
    Vista para listar todos los posts del foro, filtrando por curso si corresponde
    """
    posts = Post.objects.filter(is_active=True)
    cursos_usuario = None
    curso_id = request.GET.get('curso_id')
    curso_especifico = None
    
    if request.user.is_authenticated:
        cursos_usuario = request.user.cursos.all()
        posts = posts.filter(curso__in=cursos_usuario)
        if curso_id:
            posts = posts.filter(curso_id=curso_id)
            try:
                curso_especifico = Curso.objects.get(id=curso_id)
            except Curso.DoesNotExist:
                pass
    else:
        if curso_id:
            posts = posts.filter(curso_id=curso_id)
            try:
                curso_especifico = Curso.objects.get(id=curso_id)
            except Curso.DoesNotExist:
                pass
    
    category_filter = request.GET.get('category')
    if category_filter:
        posts = posts.filter(category=category_filter)
    
    context = {
        'posts': posts,
        'categories': Post.CATEGORY_CHOICES,
        'current_category': category_filter,
        'cursos_usuario': cursos_usuario,
        'curso_id': curso_id,
        'curso_especifico': curso_especifico
    }
    return render(request, 'forum/forum_list.html', context)


def forum_post_detail(request, post_id):
    """
    Vista para ver un post individual y sus comentarios
    """
    try:
        post = Post.objects.get(id=post_id, is_active=True)
        
        # Verificar que el usuario esté inscrito en el curso del post
        if request.user.is_authenticated and post.curso:
            if post.curso not in request.user.cursos.all():
                messages.error(request, 'No tienes acceso a este post. Debes estar inscrito en el curso correspondiente.')
                return redirect('forum_list')
        
        # Incrementar contador de vistas
        post.views += 1
        post.save()
        
        comments = post.comments.filter(is_active=True)
        
        if request.method == 'POST':
            comment_form = CommentForm(request.POST)
            if comment_form.is_valid():
                comment = comment_form.save(commit=False)
                comment.post = post
                comment.author = request.user
                comment.save()
                messages.success(request, 'Comentario publicado exitosamente.')
                return redirect('forum_post_detail', post_id=post.id)
        else:
            comment_form = CommentForm()
        
        context = {
            'post': post,
            'comments': comments,
            'comment_form': comment_form,
            'curso_especifico': post.curso
        }
        return render(request, 'forum/forum_post_detail.html', context)
        
    except Post.DoesNotExist:
        messages.error(request, 'El post no existe.')
        return redirect('forum_list')


@login_required
def forum_create_post(request):
    """
    Vista para crear un nuevo post
    """
    curso_especifico = None
    curso_id = request.GET.get('curso_id')
    
    if curso_id:
        try:
            curso_especifico = Curso.objects.get(id=curso_id)
            if curso_especifico not in request.user.cursos.all():
                messages.error(request, 'No tienes acceso a este curso.')
                return redirect('forum_list')
        except Curso.DoesNotExist:
            messages.error(request, 'Curso no encontrado.')
            return redirect('forum_list')
    
    if request.method == 'POST':
        form = PostForm(request.POST)
        if curso_especifico:
            # Si es un curso específico, no mostrar el campo curso
            form.fields.pop('curso', None)
        else:
            form.fields['curso'].queryset = request.user.cursos.all()
        
        if form.is_valid():
            post = form.save(commit=False)
            post.author = request.user
            if curso_especifico:
                post.curso = curso_especifico
            
            # Limpiar y validar HTML
            content = post.content
            allowed_tags = ['strong', 'em', 'u', 'b', 'i', 'br', 'p', 'div', 'span', 'h1', 'h2', 'h3', 'h4', 'h5', 'h6', 'ul', 'ol', 'li', 'blockquote', 'code', 'pre']
            allowed_attrs = ['class', 'style']
            import bleach
            try:
                content = bleach.clean(content, tags=allowed_tags, attributes=allowed_attrs, strip=True)
            except ImportError:
                import re
                pattern = re.compile(r'<(?!\/?(?:' + '|'.join(allowed_tags) + r')\b)[^>]+>')
                content = pattern.sub('', content)
            post.content = content
            post.save()
            messages.success(request, 'Post creado exitosamente.')
            return redirect('forum_post_detail', post_id=post.id)
    else:
        form = PostForm()
        if curso_especifico:
            # Si es un curso específico, no mostrar el campo curso
            form.fields.pop('curso', None)
        else:
            form.fields['curso'].queryset = request.user.cursos.all()
    
    context = {'form': form, 'curso_especifico': curso_especifico}
    return render(request, 'forum/forum_create_post.html', context)


@login_required
def forum_delete_post(request, post_id):
    """
    Vista para eliminar un post (solo el autor puede eliminarlo)
    """
    try:
        post = Post.objects.get(id=post_id, author=request.user)
        post.is_active = False
        post.save()
        messages.success(request, 'Post eliminado exitosamente.')
    except Post.DoesNotExist:
        messages.error(request, 'No tienes permisos para eliminar este post.')
    
    return redirect('forum_list')


@login_required
def forum_delete_comment(request, comment_id):
    """
    Vista para eliminar un comentario (solo el autor puede eliminarlo)
    """
    try:
        comment = Comment.objects.get(id=comment_id, author=request.user)
        post_id = comment.post.id
        comment.is_active = False
        comment.save()
        messages.success(request, 'Comentario eliminado exitosamente.')
        return redirect('forum_post_detail', post_id=post_id)
    except Comment.DoesNotExist:
        messages.error(request, 'No tienes permisos para eliminar este comentario.')
        return redirect('forum_list')

def test_auth(request):
    """
    Vista de prueba para verificar el estado de autenticación
    """
    context = {
        'user_authenticated': request.user.is_authenticated,
        'user': request.user,
        'session_key': request.session.session_key,
        'session_exists': request.session.session_key is not None,
    }
    return render(request, 'pages/test_auth.html', context)

def test_registration_form(request):
    """
    Vista de prueba para el formulario de registro
    """
    if not settings.DEBUG:
        return HttpResponse("Debug solo disponible en modo desarrollo", status=403)
    
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            # No crear el usuario, solo mostrar que es válido
            return JsonResponse({
                'valid': True,
                'message': 'Formulario válido - Usuario no creado (modo prueba)'
            })
        else:
            return JsonResponse({
                'valid': False,
                'errors': form.errors
            })
    else:
        form = RegistrationForm()
    
    return render(request, 'pages/test_registration.html', {'form': form})

def blog_list(request):
    """
    Vista para listar todos los posts del blog
    """
    posts = BlogPost.objects.filter(is_active=True)
    category_filter = request.GET.get('category')
    if category_filter:
        posts = posts.filter(category=category_filter)
    
    context = {
        'posts': posts,
        'categories': BlogPost.CATEGORY_CHOICES,
        'current_category': category_filter
    }
    return render(request, 'blog/blog_list.html', context)


def blog_post_detail(request, post_id):
    """
    Vista para ver un post individual del blog
    """
    try:
        post = BlogPost.objects.get(id=post_id, is_active=True)
        # Incrementar contador de vistas
        post.views += 1
        post.save()
        
        context = {
            'post': post
        }
        return render(request, 'blog/blog_post_detail.html', context)
        
    except BlogPost.DoesNotExist:
        messages.error(request, 'El artículo no existe.')
        return redirect('blog_list')


@login_required
def blog_create_post(request):
    """
    Vista para crear un nuevo post del blog (solo administradores)
    """
    if not request.user.is_staff:
        messages.error(request, 'No tienes permisos para crear artículos del blog.')
        return redirect('blog_list')
    
    if request.method == 'POST':
        form = BlogPostForm(request.POST, request.FILES)
        if form.is_valid():
            post = form.save(commit=False)
            post.author = request.user
            
            # Limpiar y validar HTML
            content = post.content
            allowed_tags = ['strong', 'em', 'u', 'b', 'i', 'br', 'p', 'div', 'span', 'h1', 'h2', 'h3', 'h4', 'h5', 'h6', 'ul', 'ol', 'li', 'blockquote', 'code', 'pre', 'img']
            allowed_attrs = ['class', 'style', 'src', 'alt', 'width', 'height']
            
            import bleach
            try:
                content = bleach.clean(content, tags=allowed_tags, attributes=allowed_attrs, strip=True)
            except ImportError:
                import re
                pattern = re.compile(r'<(?!\/?(?:' + '|'.join(allowed_tags) + r')\b)[^>]+>')
                content = pattern.sub('', content)
            
            post.content = content
            post.save()
            messages.success(request, 'Artículo publicado exitosamente.')
            return redirect('blog_post_detail', post_id=post.id)
    else:
        form = BlogPostForm()
    
    context = {'form': form}
    return render(request, 'blog/blog_create_post.html', context)


@login_required
def blog_delete_post(request, post_id):
    """
    Vista para eliminar un post del blog (solo administradores)
    """
    if not request.user.is_staff:
        messages.error(request, 'No tienes permisos para eliminar artículos del blog.')
        return redirect('blog_list')
    
    try:
        post = BlogPost.objects.get(id=post_id)
        post.is_active = False
        post.save()
        messages.success(request, 'Artículo eliminado exitosamente.')
    except BlogPost.DoesNotExist:
        messages.error(request, 'El artículo no existe.')
    
    return redirect('blog_list')


@login_required
def curso_detail(request, curso_id):
    """
    Vista para mostrar el detalle completo de un curso
    """
    try:
        curso = Curso.objects.get(id=curso_id, activo=True)
        logger.debug(request)
        
        # Verificar que el usuario esté inscrito en el curso
        if curso not in request.user.cursos.all():
            messages.error(request, 'No tienes acceso a este curso. Debes estar inscrito para ver su contenido.')
            return redirect('user_space')
        
        context = {
            'curso': curso,
        }
        return render(request, 'pages/curso_detail.html', context)
        
    except Curso.DoesNotExist:
        messages.error(request, 'El curso no existe o no está disponible.')
        return redirect('user_space')


@login_required
def plataforma_aprendizaje(request, curso_id):
    """
    Vista para la plataforma de aprendizaje de un curso específico
    """
    try:
        curso = Curso.objects.get(id=curso_id, activo=True)
        
        # Verificar que el usuario esté inscrito en el curso
        if curso not in request.user.cursos.all():
            messages.error(request, 'No tienes acceso a este curso. Debes estar inscrito para ver su contenido.')
            return redirect('user_space')
        
        # Obtener la sección activa y acciones desde la URL
        seccion_activa = request.GET.get('seccion', 'inicio')
        action = request.GET.get('action')
        post_id = request.GET.get('post_id')
        
        # Definir el contexto base
        context = {
            'curso': curso,
            'seccion_activa': seccion_activa,
        }
        
        # Si hay una acción específica, manejar según el tipo
        if action == 'crear_post':
            # Cargar el formulario de crear post en el contexto
            from .forms import PostForm
            form = PostForm()
            # Ocultar el campo curso ya que se asigna automáticamente
            form.fields.pop('curso', None)
            context['form'] = form
            context['action'] = 'crear_post'
        elif action == 'ver_post' and post_id:
            # Cargar el post específico en el contexto
            try:
                post = Post.objects.get(id=post_id, curso=curso, is_active=True)
                
                # Manejar envío de comentarios
                if request.method == 'POST':
                    comment_form = CommentForm(request.POST)
                    if comment_form.is_valid():
                        comment = comment_form.save(commit=False)
                        comment.post = post
                        comment.author = request.user
                        comment.save()
                        messages.success(request, 'Comentario publicado exitosamente.')
                        # Redirigir de vuelta al mismo post para evitar reenvío del formulario
                        from django.urls import reverse
                        redirect_url = f"{reverse('plataforma_aprendizaje', kwargs={'curso_id': curso_id})}?seccion=foro&action=ver_post&post_id={post_id}"
                        return HttpResponseRedirect(redirect_url)
                else:
                    comment_form = CommentForm()
                
                # Incrementar contador de vistas solo en GET
                # if request.method == 'GET':
                #     post.views += 1
                #     post.save()
                
                comments = post.comments.filter(is_active=True)
                
                context['post'] = post
                context['comments'] = comments
                context['comment_form'] = comment_form
                context['action'] = 'ver_post'
            except Post.DoesNotExist:
                messages.error(request, 'El post no existe o no está disponible.')
                return redirect('plataforma_aprendizaje', curso_id=curso_id)
        
        return render(request, 'pages/plataforma_aprendizaje.html', context)
        
    except Curso.DoesNotExist:
        messages.error(request, 'El curso no existe o no está disponible.')
        return redirect('user_space')


@login_required
def plataforma_foro(request, curso_id):
    """
    Vista para el foro dentro de la plataforma de aprendizaje
    """
    try:
        curso = Curso.objects.get(id=curso_id, activo=True)
        
        # Verificar que el usuario esté inscrito en el curso
        if curso not in request.user.cursos.all():
            messages.error(request, 'No tienes acceso a este curso. Debes estar inscrito para ver su contenido.')
            return redirect('user_space')
        
        # Obtener posts del curso
        posts = Post.objects.filter(curso=curso, is_active=True).order_by('-created_at')
        
        # Filtros
        category_filter = request.GET.get('category')
        if category_filter:
            posts = posts.filter(category=category_filter)
        
        context = {
            'curso': curso,
            'posts': posts,
            'categories': Post.CATEGORY_CHOICES,
            'current_category': category_filter,
        }
        
        # Si es una petición AJAX, devolver solo el contenido
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            html = render_to_string('pages/plataforma_foro_content.html', context, request=request)
            return JsonResponse({'html': html})
        
        # Si no es AJAX, renderizar la página completa
        return render(request, 'pages/plataforma_foro.html', context)
        
    except Curso.DoesNotExist:
        messages.error(request, 'El curso no existe o no está disponible.')
        return redirect('user_space')


@login_required
def plataforma_foro_post_detail(request, curso_id, post_id):
    """
    Vista para ver un post individual dentro de la plataforma
    """
    try:
        curso = Curso.objects.get(id=curso_id, activo=True)
        
        # Verificar que el usuario esté inscrito en el curso
        if curso not in request.user.cursos.all():
            messages.error(request, 'No tienes acceso a este curso. Debes estar inscrito para ver su contenido.')
            return redirect('user_space')
        
        post = Post.objects.get(id=post_id, curso=curso, is_active=True)
        
        # Incrementar contador de vistas
        post.views += 1
        post.save()
        
        comments = post.comments.filter(is_active=True)
        
        if request.method == 'POST':
            comment_form = CommentForm(request.POST)
            if comment_form.is_valid():
                comment = comment_form.save(commit=False)
                comment.post = post
                comment.author = request.user
                comment.save()
                messages.success(request, 'Comentario publicado exitosamente.')
                return redirect('plataforma_foro_post_detail', curso_id=curso.id, post_id=post.id)
        else:
            comment_form = CommentForm()
        
        context = {
            'curso': curso,
            'post': post,
            'comments': comments,
            'comment_form': comment_form,
        }
        return render(request, 'pages/plataforma_foro_post_detail.html', context)
        
    except (Curso.DoesNotExist, Post.DoesNotExist):
        messages.error(request, 'El curso o post no existe.')
        return redirect('user_space')


@login_required
def plataforma_foro_create_post(request, curso_id):
    """
    Vista para crear un nuevo post dentro de la plataforma
    """
    try:
        curso = Curso.objects.get(id=curso_id, activo=True)
        
        # Verificar que el usuario esté inscrito en el curso
        if curso not in request.user.cursos.all():
            messages.error(request, 'No tienes acceso a este curso. Debes estar inscrito para ver su contenido.')
            return redirect('user_space')
        
        if request.method == 'POST':
            form = PostForm(request.POST)
            if form.is_valid():
                post = form.save(commit=False)
                post.author = request.user
                post.curso = curso
                
                # Limpiar y validar HTML
                content = post.content
                allowed_tags = ['strong', 'em', 'u', 'b', 'i', 'br', 'p', 'div', 'span', 'h1', 'h2', 'h3', 'h4', 'h5', 'h6', 'ul', 'ol', 'li', 'blockquote', 'code', 'pre']
                allowed_attrs = ['class', 'style']
                import bleach
                try:
                    content = bleach.clean(content, tags=allowed_tags, attributes=allowed_attrs, strip=True)
                except ImportError:
                    import re
                    pattern = re.compile(r'<(?!\/?(?:' + '|'.join(allowed_tags) + r')\b)[^>]+>')
                    content = pattern.sub('', content)
                post.content = content
                post.save()
                messages.success(request, 'Post creado exitosamente.')
                from django.urls import reverse
                redirect_url = f"{reverse('plataforma_aprendizaje', kwargs={'curso_id': curso.id})}?seccion=foro"
                return redirect(redirect_url)
        else:
            form = PostForm()
            # Ocultar el campo curso ya que se asigna automáticamente
            form.fields.pop('curso', None)
        
        context = {'form': form, 'curso': curso}
        return render(request, 'pages/plataforma_foro_create_post.html', context)
        
    except Curso.DoesNotExist:
        messages.error(request, 'El curso no existe o no está disponible.')
        return redirect('user_space')


@login_required
def plataforma_foro_delete_post(request, curso_id, post_id):
    """
    Vista para eliminar un post dentro de la plataforma (solo el autor puede eliminarlo)
    """
    try:
        curso = Curso.objects.get(id=curso_id, activo=True)
        
        # Verificar que el usuario esté inscrito en el curso
        if curso not in request.user.cursos.all():
            messages.error(request, 'No tienes acceso a este curso. Debes estar inscrito para ver su contenido.')
            return redirect('user_space')
        
        # Buscar el post y verificar que el usuario sea el autor
        post = Post.objects.get(id=post_id, curso=curso, author=request.user, is_active=True)
        
        # Marcar como inactivo en lugar de eliminar físicamente
        post.is_active = False
        post.save()
        
        messages.success(request, 'Post eliminado exitosamente.')
        
    except Curso.DoesNotExist:
        messages.error(request, 'El curso no existe o no está disponible.')
    except Post.DoesNotExist:
        messages.error(request, 'No tienes permisos para eliminar este post o el post no existe.')
    
    # Redirigir de vuelta a la plataforma del curso en la sección foro
    from django.urls import reverse
    redirect_url = f"{reverse('plataforma_aprendizaje', kwargs={'curso_id': curso_id})}?seccion=foro"
    return redirect(redirect_url)


# @login_required
# def plataforma_foro_delete_comment(request, curso_id, comment_id):
#     """
#     Vista para eliminar un comentario dentro de la plataforma (solo el autor puede eliminarlo)
#     """
#     try:
#         curso = Curso.objects.get(id=curso_id, activo=True)
        
#         # Verificar que el usuario esté inscrito en el curso
#         if curso not in request.user.cursos.all():
#             messages.error(request, 'No tienes acceso a este curso. Debes estar inscrito para ver su contenido.')
#             return redirect('user_space')
        
#         # Buscar el comentario y verificar que el usuario sea el autor
#         comment = Comment.objects.get(id=comment_id, author=request.user, is_active=True)
        
#         # Verificar que el comentario pertenezca a un post del curso
#         if comment.post.curso != curso:
#             messages.error(request, 'No tienes permisos para eliminar este comentario.')
#             return redirect('plataforma_aprendizaje', curso_id=curso_id)
        
#         # Marcar como inactivo en lugar de eliminar físicamente
#         comment.is_active = False
#         comment.save()
        
#         messages.success(request, 'Comentario eliminado exitosamente.')
        
#     except Curso.DoesNotExist:
#         messages.error(request, 'El curso no existe o no está disponible.')
#     except Comment.DoesNotExist:
#         messages.error(request, 'No tienes permisos para eliminar este comentario o el comentario no existe.')
    
#     # Redirigir de vuelta a la plataforma del curso
#     return redirect('plataforma_aprendizaje', curso_id=curso_id)


def cursos_list(request):
    """
    Vista para mostrar todos los cursos de capacitación disponibles
    """
    from .models import InscripcionCurso
    
    cursos_activos = Curso.objects.filter(activo=True).order_by('nombre')
    
    # Si el usuario está logueado, obtener sus inscripciones
    inscripciones_usuario = []
    inscripciones_pendientes = []
    if request.user.is_authenticated:
        # Obtener cursos donde el usuario está inscrito (ya pagó y tiene acceso)
        cursos_inscritos = request.user.cursos.all().values_list('id', flat=True)
        
        # Obtener inscripciones pendientes
        inscripciones_pendientes = InscripcionCurso.objects.filter(
            usuario_creado=request.user,
            estado__in=['pendiente', 'confirmada', 'en_proceso']
        ).values_list('curso_id', flat=True)
        
        # Solo incluir cursos donde realmente está inscrito (ya pagó)
        inscripciones_usuario = list(cursos_inscritos)
    
    context = {
        'cursos': cursos_activos,
        'inscripciones_usuario': inscripciones_usuario,
        'inscripciones_pendientes': inscripciones_pendientes,
    }
    return render(request, 'pages/cursos_list.html', context)


@login_required
def mi_perfil(request):
    """
    Vista para gestionar el perfil del usuario
    """
    if request.method == 'POST':
        # Obtener datos del formulario
        first_name = request.POST.get('first_name', '')
        last_name = request.POST.get('last_name', '')
        email = request.POST.get('email', '')
        company_name = request.POST.get('company_name', '')
        
        # Validar que el email no esté en uso por otro usuario
        if email != request.user.email:
            if User.objects.filter(email=email).exclude(id=request.user.id).exists():
                messages.error(request, 'Este correo electrónico ya está en uso por otro usuario.')
                return redirect('mi_perfil')
        
        # Actualizar datos del usuario
        request.user.first_name = first_name
        request.user.last_name = last_name
        request.user.email = email
        request.user.company_name = company_name
        
        # Manejar subida de foto de perfil
        if 'profile_photo' in request.FILES:
            request.user.profile_photo = request.FILES['profile_photo']
        
        # Manejar eliminación de foto de perfil
        if request.POST.get('delete_photo') == 'true':
            if request.user.profile_photo:
                request.user.profile_photo.delete(save=False)
            request.user.profile_photo = None
        
        request.user.save()
        messages.success(request, 'Perfil actualizado exitosamente.')
        return redirect('mi_perfil')
    
    # Obtener cursos completados del usuario (puedes ajustar la lógica según tu modelo)
    cursos_completados = request.user.cursos.all()  # Ajusta según tu lógica de cursos completados
    
    context = {
        'cursos_completados': cursos_completados,
    }
    return render(request, 'pages/mi_perfil.html', context)


def curso_detail_public(request, curso_id):
    """
    Vista pública para mostrar el detalle de un curso (sin restricción de login)
    """
    from .models import InscripcionCurso
    
    try:
        curso = Curso.objects.get(id=curso_id, activo=True)
        
        # Verificar si el usuario está inscrito en este curso
        usuario_inscrito = False
        tiene_inscripcion_pendiente = False
        if request.user.is_authenticated:
            # Verificar si el usuario tiene acceso al curso (ya pagó)
            tiene_acceso = request.user.cursos.filter(id=curso_id).exists()
            
            # Verificar si tiene inscripción pendiente
            tiene_inscripcion_pendiente = InscripcionCurso.objects.filter(
                usuario_creado=request.user,
                curso=curso,
                estado__in=['pendiente', 'confirmada', 'en_proceso']
            ).exists()
            
            # Solo considerar como inscrito si realmente tiene acceso al curso
            usuario_inscrito = tiene_acceso
        
        context = {
            'curso': curso,
            'usuario_inscrito': usuario_inscrito,
            'tiene_inscripcion_pendiente': tiene_inscripcion_pendiente,
        }
        return render(request, 'pages/curso_detail_public.html', context)
        
    except Curso.DoesNotExist:
        messages.error(request, 'El curso no existe o no está disponible.')
        return redirect('cursos_list')


def custom_404(request, exception):
    """
    Vista personalizada para error 404
    """
    return render(request, 'pages/404.html', status=404)


def custom_500(request):
    """
    Vista personalizada para error 500
    """
    return render(request, 'pages/500.html', status=500)

@login_required
def extend_session(request):
    """
    Vista para extender la sesión del usuario
    """
    if request.method == 'POST':
        # Actualizar la última actividad en la sesión
        request.session['last_activity'] = time.time()
        request.session.modified = True
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error'}, status=400)


@login_required
def admin_inscripciones(request):
    """
    Panel de administrador para gestionar inscripciones a cursos
    """
    if not request.user.is_staff:
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('index')
    
    # Obtener todas las inscripciones ordenadas por fecha
    inscripciones = InscripcionCurso.objects.all()
    
    # Filtros
    estado_filter = request.GET.get('estado')
    curso_filter = request.GET.get('curso')
    
    if estado_filter:
        inscripciones = inscripciones.filter(estado=estado_filter)
    
    if curso_filter:
        inscripciones = inscripciones.filter(curso_id=curso_filter)
    
    # Obtener cursos para el filtro
    cursos = Curso.objects.filter(activo=True)
    
    context = {
        'inscripciones': inscripciones,
        'cursos': cursos,
        'estados': InscripcionCurso.ESTADO_CHOICES,
        'estado_filter': estado_filter,
        'curso_filter': curso_filter,
    }
    
    return render(request, 'admin/inscripciones_list.html', context)


@login_required
def admin_inscripcion_detail(request, inscripcion_id):
    """
    Vista detallada de una inscripción específica
    """
    if not request.user.is_staff:
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('index')
    
    try:
        inscripcion = InscripcionCurso.objects.get(id=inscripcion_id)
    except InscripcionCurso.DoesNotExist:
        messages.error(request, 'La inscripción no existe.')
        return redirect('admin-inscripciones')
    
    context = {
        'inscripcion': inscripcion,
    }
    
    return render(request, 'admin/inscripcion_detail.html', context)


@login_required
def admin_marcar_pagado(request, inscripcion_id):
    """
    Marcar una inscripción como pagada y crear el usuario o reutilizar uno existente
    """
    if not request.user.is_staff:
        messages.error(request, 'No tienes permisos para realizar esta acción.')
        return JsonResponse({'success': False, 'message': 'Sin permisos'})
    
    try:
        inscripcion = InscripcionCurso.objects.get(id=inscripcion_id)
        
        if inscripcion.estado == 'pendiente':
            # Marcar como pagado y crear/reutilizar usuario
            user, password_temp = inscripcion.marcar_como_pagado()
            
            if user:
                if password_temp:
                    # Usuario nuevo creado - procesar correo y dirección
                    from django.utils import timezone
                    
                    # Actualizar tracking
                    inscripcion.intentos_procesamiento += 1
                    inscripcion.fecha_ultimo_intento = timezone.now()
                    
                    # Paso 1: Intentar crear dirección gryphos
                    direccion_creada = crear_direccion_gryphos(request, user, password_temp)
                    
                    if not direccion_creada:
                        # Error al crear dirección de correo
                        inscripcion.error_creacion_correo = f"Error al crear dirección de correo para {user.username}@gryphos.cl"
                        inscripcion.error_envio_bienvenida = None
                        inscripcion.save()
                        
                        return JsonResponse({
                            'success': False,
                            'message': 'Error al crear dirección de correo electrónico',
                            'error_type': 'creacion_correo',
                            'username': user.username,
                            'usuario_nuevo': True,
                            'password_temp': password_temp,
                            'can_retry': True
                        })
                    
                    # Paso 2: Si se creó la dirección, intentar enviar correo de bienvenida
                    correo_enviado = enviar_correo_bienvenida(request, user, inscripcion.password_temp, inscripcion.curso.nombre)
                    
                    if not correo_enviado:
                        # Error al enviar correo de bienvenida
                        inscripcion.error_creacion_correo = None
                        inscripcion.error_envio_bienvenida = f"Error al enviar correo de bienvenida a {user.email}"
                        inscripcion.save()
                        
                        return JsonResponse({
                            'success': False,
                            'message': 'Se creó el correo electrónico pero falló el envío del correo de bienvenida',
                            'error_type': 'envio_bienvenida',
                            'username': user.username,
                            'usuario_nuevo': True,
                            'password_temp': password_temp,
                            'can_retry': True
                        })
                    
                    # Éxito completo
                    inscripcion.error_creacion_correo = None
                    inscripcion.error_envio_bienvenida = None
                    inscripcion.save()
                    
                    messages.success(request, f'Inscripción marcada como pagada. Usuario creado: {user.username}')
                    logger.info(f"Correo de bienvenida enviado exitosamente a {user.email}")
                    
                    return JsonResponse({
                        'success': True, 
                        'message': 'Inscripción marcada como pagada',
                        'username': user.username,
                        'usuario_nuevo': True
                    })
                    
                else:
                    # Usuario existente reutilizado - solo enviar correo de bienvenida
                    correo_enviado = enviar_correo_bienvenida_usuario_existente(request, user, inscripcion.curso.nombre)
                    
                    if not correo_enviado:
                        inscripcion.error_envio_bienvenida = f"Error al enviar correo de bienvenida a usuario existente {user.email}"
                        inscripcion.save()
                        
                        return JsonResponse({
                            'success': False,
                            'message': 'Error al enviar correo de bienvenida a usuario existente',
                            'error_type': 'envio_bienvenida_existente',
                            'username': user.username,
                            'usuario_nuevo': False,
                            'can_retry': True
                        })
                    
                    inscripcion.error_envio_bienvenida = None
                    inscripcion.save()
                    
                    messages.success(request, f'Inscripción marcada como pagada. Usuario existente reutilizado: {user.username}')
                    logger.info(f"Correo de bienvenida a usuario existente enviado exitosamente a {user.email}")
                    
                    return JsonResponse({
                        'success': True, 
                        'message': 'Inscripción marcada como pagada',
                        'username': user.username,
                        'usuario_nuevo': False
                    })
            else:
                return JsonResponse({
                    'success': False, 
                    'message': 'Error al procesar la inscripción'
                })
        else:
            return JsonResponse({
                'success': False, 
                'message': 'La inscripción ya no está pendiente'
            })
            
    except InscripcionCurso.DoesNotExist:
        return JsonResponse({
            'success': False, 
            'message': 'Inscripción no encontrada'
        })
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'message': f'Error: {str(e)}'
        })


@login_required
def admin_cambiar_estado(request, inscripcion_id):
    """
    Cambiar el estado de una inscripción
    """
    if not request.user.is_staff:
        messages.error(request, 'No tienes permisos para realizar esta acción.')
        return JsonResponse({'success': False, 'message': 'Sin permisos'})
    
    if request.method == 'POST':
        nuevo_estado = request.POST.get('estado')
        observaciones = request.POST.get('observaciones', '')
        
        if nuevo_estado not in dict(InscripcionCurso.ESTADO_CHOICES):
            return JsonResponse({
                'success': False, 
                'message': 'Estado inválido'
            })
        
        try:
            inscripcion = InscripcionCurso.objects.get(id=inscripcion_id)
            inscripcion.estado = nuevo_estado
            inscripcion.observaciones = observaciones
            inscripcion.save()
            
            messages.success(request, f'Estado de inscripción actualizado a: {inscripcion.get_estado_display()}')
            return JsonResponse({
                'success': True, 
                'message': 'Estado actualizado correctamente'
            })
            
        except InscripcionCurso.DoesNotExist:
            return JsonResponse({
                'success': False, 
                'message': 'Inscripción no encontrada'
            })
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'message': f'Error: {str(e)}'
            })
    
    return JsonResponse({
        'success': False, 
        'message': 'Método no permitido'
    })


@login_required
def admin_reenviar_correo(request, inscripcion_id):
    """
    Vista para reenviar el correo de instrucciones de pago
    """
    if not request.user.is_staff:
        messages.error(request, 'No tienes permisos para realizar esta acción.')
        return JsonResponse({'success': False, 'message': 'Sin permisos'})
    
    if request.method == 'POST':
        try:
            inscripcion = InscripcionCurso.objects.get(id=inscripcion_id)
            
            # Reenviar correo de instrucciones de pago
            if enviar_correo_instrucciones_pago(request, inscripcion):
                #messages.success(request, f'Correo de instrucciones de pago reenviado exitosamente a {inscripcion.correo_contacto}')
                logger.info(f"Correo de instrucciones de pago reenviado exitosamente a {inscripcion.correo_contacto}")
                return JsonResponse({
                    'success': True, 
                    'message': f'Correo reenviado exitosamente a {inscripcion.correo_contacto}'
                })
            else:
                messages.error(request, 'Error al reenviar el correo de instrucciones de pago')
                return JsonResponse({
                    'success': False, 
                    'message': 'Error al reenviar el correo de instrucciones de pago'
                })
                
        except InscripcionCurso.DoesNotExist:
            return JsonResponse({
                'success': False, 
                'message': 'Inscripción no encontrada'
            })
        except Exception as e:
            logger.error(f"Error al reenviar correo: {e}")
            return JsonResponse({
                'success': False, 
                'message': f'Error al reenviar correo: {str(e)}'
            })
    
    return JsonResponse({
        'success': False, 
        'message': 'Método no permitido'
    })


@login_required
def admin_reintentar_procesamiento(request, inscripcion_id):
    """
    Reintentar el procesamiento de una inscripción pagada que tuvo errores
    """
    if not request.user.is_staff:
        messages.error(request, 'No tienes permisos para realizar esta acción.')
        return JsonResponse({'success': False, 'message': 'Sin permisos'})
    
    try:
        inscripcion = InscripcionCurso.objects.get(id=inscripcion_id)
        
        if inscripcion.estado != 'pagado':
            return JsonResponse({
                'success': False, 
                'message': 'Solo se puede reintentar procesamiento de inscripciones pagadas'
            })
        
        if not inscripcion.usuario_creado:
            return JsonResponse({
                'success': False, 
                'message': 'No hay usuario asociado a esta inscripción'
            })
        
        user = inscripcion.usuario_creado
        
        # Verificar si es usuario nuevo o existente
        from django.contrib.auth.hashers import check_password
        is_new_user = not user.has_usable_password() or user.password.startswith('pbkdf2_sha256$')
        
        if is_new_user:
            # Usuario nuevo - reintentar creación de dirección y correo de bienvenida
            from django.utils import timezone
            
            # Actualizar tracking
            inscripcion.intentos_procesamiento += 1
            inscripcion.fecha_ultimo_intento = timezone.now()
            
            # Obtener la contraseña temporal guardada en la inscripción
            password_temp = inscripcion.password_temp
            
            # Paso 1: Intentar crear dirección gryphos
            direccion_creada = crear_direccion_gryphos(request, user, password_temp)
            
            if not direccion_creada:
                inscripcion.error_creacion_correo = f"Error al crear dirección de correo para {user.username}@gryphos.cl"
                inscripcion.error_envio_bienvenida = None
                inscripcion.save()
                
                return JsonResponse({
                    'success': False,
                    'message': 'Error al crear dirección de correo electrónico',
                    'error_type': 'creacion_correo',
                    'username': user.username,
                    'usuario_nuevo': True,
                    'can_retry': True
                })
            
            # Paso 2: Si se creó la dirección, intentar enviar correo de bienvenida
            correo_enviado = enviar_correo_bienvenida(request, user, inscripcion.password_temp, inscripcion.curso.nombre)
            
            if not correo_enviado:
                inscripcion.error_creacion_correo = None
                inscripcion.error_envio_bienvenida = f"Error al enviar correo de bienvenida a {user.email}"
                inscripcion.save()
                
                return JsonResponse({
                    'success': False,
                    'message': 'Se creó el correo electrónico pero falló el envío del correo de bienvenida',
                    'error_type': 'envio_bienvenida',
                    'username': user.username,
                    'usuario_nuevo': True,
                    'can_retry': True
                })
            
            # Éxito completo
            inscripcion.error_creacion_correo = None
            inscripcion.error_envio_bienvenida = None
            inscripcion.save()
            
            messages.success(request, f'Procesamiento completado exitosamente para usuario: {user.username}')
            
            return JsonResponse({
                'success': True, 
                'message': 'Procesamiento completado exitosamente',
                'username': user.username,
                'usuario_nuevo': True
            })
            
        else:
            # Usuario existente - solo reintentar correo de bienvenida
            correo_enviado = enviar_correo_bienvenida_usuario_existente(request, user, inscripcion.curso.nombre)
            
            if not correo_enviado:
                inscripcion.error_envio_bienvenida = f"Error al enviar correo de bienvenida a usuario existente {user.email}"
                inscripcion.save()
                
                return JsonResponse({
                    'success': False,
                    'message': 'Error al enviar correo de bienvenida a usuario existente',
                    'error_type': 'envio_bienvenida_existente',
                    'username': user.username,
                    'usuario_nuevo': False,
                    'can_retry': True
                })
            
            inscripcion.error_envio_bienvenida = None
            inscripcion.save()
            
            messages.success(request, f'Correo de bienvenida enviado exitosamente a usuario existente: {user.username}')
            
            return JsonResponse({
                'success': True, 
                'message': 'Correo de bienvenida enviado exitosamente',
                'username': user.username,
                'usuario_nuevo': False
            })
            
    except InscripcionCurso.DoesNotExist:
        return JsonResponse({
            'success': False, 
            'message': 'Inscripción no encontrada'
        })
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'message': f'Error: {str(e)}'
        })


def crear_direccion_gryphos(request, user, password_temp):
    url = "https://mail.gryphos.cl/api/v1/add/mailbox"
    headers = {
        "Accept": "application/json",
        "X-API-Key": f"{settings.API_KEY_MAILCOW}",
        "Content-Type": "application/json"
    }
    data = {
    "active": "1",
    "domain": "gryphos.cl",
    "local_part": user.username,
    "name": user.get_full_name() or user.username,
    "authsource": "mailcow",
    "password": password_temp,
    "password2": password_temp,
    "quota": "3072",
    "force_pw_update": "0",
    "tls_enforce_in": "1",
    "tls_enforce_out": "1",
    "tags": [
        "estudiante"
    ]
    }
    try:
        response = requests.post(url, headers=headers, json=data)
        if response.status_code == 200:
            logger.info(f"Dirección de correo creada exitosamente: {response.json()}")
            return True
        else:
            logger.error(f"Error al crear la dirección de correo: {response.json()}")
            return False
    except Exception as e:
        logger.error(f"Error al crear la dirección de correo: {e}")
        return False


def generate_jitsi_token(request, room_name=None, user=None):
    """
    Genera un token JWT para Jitsi Meet
    
    Args:
        request: HttpRequest object
        room_name: Nombre específico de la sala (opcional)
        user: Usuario específico (opcional, por defecto usa request.user)
    
    Returns:
        JsonResponse con el token JWT
    """
    logger.info(f"Iniciando generación de JWT - Usuario: {request.user.username}, Sala: {room_name or 'general'}")
    
    if not request.user.is_authenticated:
        logger.warning(f"Intento de generar JWT sin autenticación desde IP: {request.META.get('REMOTE_ADDR', 'desconocida')}")
        return JsonResponse({"error": "No autenticado"}, status=403)

    # Usar el usuario proporcionado o el usuario de la request
    current_user = user or request.user
    logger.debug(f"Generando JWT para usuario: {current_user.username} (staff: {current_user.is_staff})")
    
    try:
        payload = {
            "iss": "gryphos",  # Debe coincidir con el issuer en Jitsi
            "aud": "jitsi",  # Audience estándar para Jitsi
            "sub": "meet.gryphos.cl",
            "room": room_name or "*",  # Sala específica o cualquier sala
            "exp": datetime.utcnow() + timedelta(hours=2),  # Expira en 2 horas
            "context": {
                "user": {
                    "name": current_user.get_full_name() or current_user.username,
                    "email": current_user.email,
                    "avatar": getattr(current_user, 'profile', None) and current_user.profile.avatar_url or "",
                    "moderator": current_user.is_staff,  # Los staff son moderadores
                },
            }
        }
        
        logger.debug(f"Payload JWT generado para sala: {payload['room']}")
        
        # Log detallado de tiempos
        import pytz
        
        # Obtener tiempo actual en UTC
        now_utc = datetime.utcnow()
        exp_utc = payload['exp']
        
        # Convertir a zona horaria local para logs
        try:
            local_tz = pytz.timezone('America/Santiago')  # Zona horaria de Chile
            now_local = now_utc.replace(tzinfo=pytz.UTC).astimezone(local_tz)
            exp_local = exp_utc.replace(tzinfo=pytz.UTC).astimezone(local_tz)
        except:
            now_local = now_utc
            exp_local = exp_utc
        
        logger.debug(f"=== TIEMPOS JWT ===")
        logger.debug(f"Tiempo actual (UTC): {now_utc}")
        logger.debug(f"Tiempo actual (Local): {now_local}")
        logger.debug(f"Tiempo expiración (UTC): {exp_utc}")
        logger.debug(f"Tiempo expiración (Local): {exp_local}")
        logger.debug(f"Duración del token: {exp_utc - now_utc}")
        logger.debug(f"==================")
        
        token = jwt.encode(payload, settings.JITSI_JWT_SECRET, algorithm="HS256")
        logger.debug(f"JWT generado exitosamente para usuario {current_user.username} en sala {payload['room']}")
        
        # Log detallado del token
        logger.debug(f"=== TOKEN JWT COMPLETO ===")
        logger.debug(f"Token: {token}")
        logger.debug(f"Payload: {payload}")
        logger.debug(f"Secret key length: {len(settings.JITSI_JWT_SECRET)}")
        logger.debug(f"==========================")
        
        return JsonResponse({"token": token})
        
    except Exception as e:
        logger.error(f"Error generando JWT para usuario {current_user.username}: {str(e)}")
        logger.error(f"Tipo de error: {type(e).__name__}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        return JsonResponse({"error": "Error interno del servidor"}, status=500)


@login_required
def join_meeting(request, videollamada_id):
    """
    Vista para unirse a una videollamada específica con JWT token
    """
    from .models import Videollamada
    
    logger.info(f"Intento de unirse a videollamada {videollamada_id} - Usuario: {request.user.username}")
    logger.debug(f"IP del usuario: {request.META.get('REMOTE_ADDR', 'desconocida')}")
    
    try:
        # Obtener la videollamada
        logger.debug(f"Buscando videollamada con ID: {videollamada_id}")
        videollamada = Videollamada.objects.get(id=videollamada_id, activa=True)
        logger.debug(f"Videollamada encontrada: {videollamada} (Curso: {videollamada.curso.nombre})")
        
        # Verificar que el usuario esté inscrito en el curso de la videollamada
        logger.debug(f"Verificando acceso del usuario {request.user.username} al curso {videollamada.curso.nombre}")
        if videollamada.curso not in request.user.cursos.all():
            logger.warning(f"Acceso denegado: Usuario {request.user.username} no está inscrito en el curso {videollamada.curso.nombre}")
            messages.error(request, 'No tienes acceso a esta videollamada. Debes estar inscrito en el curso correspondiente.')
            return redirect('user_space')
        
        logger.debug(f"Acceso verificado: Usuario {request.user.username} tiene acceso al curso {videollamada.curso.nombre}")
        
        # Verificar que la videollamada esté activa ahora
        logger.debug(f"Verificando si la videollamada {videollamada_id} está activa ahora")
        if not videollamada.esta_activa_ahora():
            logger.warning(f"Videollamada {videollamada_id} no está activa en este momento para usuario {request.user.username}")
            messages.warning(request, 'Esta videollamada no está activa en este momento.')
            return redirect('user_space')
        
        logger.debug(f"Videollamada {videollamada_id} está activa y disponible")
        
        # Verificar que tenga un enlace configurado
        logger.debug(f"Verificando enlace de videollamada {videollamada_id}")
        if not videollamada.link_videollamada:
            logger.error(f"Videollamada {videollamada_id} no tiene enlace configurado")
            messages.error(request, 'Esta videollamada no tiene un enlace configurado.')
            return redirect('user_space')
        
        logger.debug(f"Enlace de videollamada {videollamada_id} verificado: {videollamada.link_videollamada}")
        
        # Generar JWT token para la videollamada usando la función existente
        try:
            # Extraer el nombre de la sala de la URL de la videollamada
            from urllib.parse import urlparse
            parsed_url = urlparse(videollamada.link_videollamada)
            # Obtener la última parte de la URL (después del último /)
            room_name = parsed_url.path.strip('/').split('/')[-1]
            logger.debug(f"URL de videollamada: {videollamada.link_videollamada}")
            logger.debug(f"Nombre de sala extraído: {room_name}")
            
            # Usar la función generate_jitsi_token para generar el token
            logger.debug(f"Llamando a generate_jitsi_token para sala {room_name}")
            token_response = generate_jitsi_token(request, room_name=room_name)
            
            if token_response.status_code != 200:
                logger.error(f"Error generando JWT para videollamada {videollamada_id}: {token_response.content}")
                messages.error(request, 'Error al generar el token de acceso a la videollamada.')
                return redirect('user_space')
            
            logger.debug(f"JWT generado exitosamente para videollamada {videollamada_id}")
            
            # Extraer el token del response
            import json
            token_data = json.loads(token_response.content)
            token = token_data.get('token')
            
            if not token:
                logger.error(f"No se pudo obtener el token JWT para videollamada {videollamada_id}")
                messages.error(request, 'Error al generar el token de acceso a la videollamada.')
                return redirect('user_space')
            
            logger.debug(f"Token JWT extraído exitosamente para videollamada {videollamada_id}")
            
            # Log del tiempo de procesamiento
            from datetime import datetime
            processing_time = datetime.utcnow()
            logger.info(f"=== PROCESAMIENTO JWT ===")
            logger.debug(f"Tiempo de procesamiento: {processing_time}")
            logger.debug(f"Token extraído y listo para uso")
            logger.debug(f"Token completo: {token}")
            logger.debug(f"========================")
            
            # Construir la URL de la videollamada con el token
            base_url = videollamada.link_videollamada.rstrip('/')
            logger.debug(f"URL base de videollamada: {base_url}")
            
            # Si la URL ya tiene parámetros, agregar el token, si no, agregar como primer parámetro
            if '?' in base_url:
                meeting_url = f"{base_url}&jwt={token}"
            else:
                meeting_url = f"{base_url}?jwt={token}"
            
            logger.debug(f"URL final de videollamada construida: {meeting_url[:100]}...")
            
            # Log de la acción exitosa con tiempo
            redirect_time = datetime.utcnow()
            logger.info(f"Usuario {request.user.username} se unió exitosamente a videollamada {videollamada.id} del curso {videollamada.curso.nombre}")
            logger.debug(f"Tiempo de redirección: {redirect_time}")
            logger.debug(f"Tiempo total de procesamiento: {redirect_time - processing_time}")
            logger.debug(f"Redirigiendo a: {meeting_url}")
            
            # Redirigir a la videollamada
            return redirect(meeting_url)
            
        except Exception as e:
            logger.error(f"Error generando JWT para videollamada {videollamada_id}: {str(e)}")
            messages.error(request, 'Error al generar el token de acceso a la videollamada.')
            return redirect('user_space')
            
    except Videollamada.DoesNotExist:
        logger.warning(f"Videollamada {videollamada_id} no encontrada para usuario {request.user.username}")
        messages.error(request, 'Videollamada no encontrada.')
        return redirect('user_space')
    except Exception as e:
        logger.error(f"Error inesperado al unirse a videollamada {videollamada_id}: {str(e)}")
        messages.error(request, 'Error inesperado al acceder a la videollamada.')
        return redirect('user_space')

@login_required
def plataforma_foro_ajax(request, curso_id):
    """
    Vista AJAX para cargar el contenido del foro dinámicamente
    """
    try:
        curso = Curso.objects.get(id=curso_id, activo=True)
        
        # Verificar que el usuario esté inscrito en el curso
        if curso not in request.user.cursos.all():
            return JsonResponse({'error': 'No tienes acceso a este curso'}, status=403)
        
        # Verificar si es una acción específica
        action = request.GET.get('action')
        
        if action == 'crear_post':
            # Cargar formulario de crear post
            from .forms import PostForm
            form = PostForm()
            # Ocultar el campo curso ya que se asigna automáticamente
            form.fields.pop('curso', None)
            
            context = {
                'curso': curso,
                'form': form,
            }
            
            html = render_to_string('pages/plataforma_foro_create_form.html', context, request=request)
            return JsonResponse({'html': html})
        
        elif action == 'ver_post':
            # Cargar post específico
            post_id = request.GET.get('post_id')
            if post_id:
                try:
                    post = Post.objects.get(id=post_id, curso=curso, is_active=True)
                    # Incrementar contador de vistas
                    post.views += 1
                    post.save()
                    
                    comments = post.comments.filter(is_active=True)
                    comment_form = CommentForm()
                    
                    context = {
                        'curso': curso,
                        'post': post,
                        'comments': comments,
                        'comment_form': comment_form,
                    }
                    
                    html = render_to_string('pages/plataforma_foro_post_detail_content.html', context, request=request)
                    return JsonResponse({'html': html})
                except Post.DoesNotExist:
                    return JsonResponse({'error': 'El post no existe'}, status=404)
            else:
                return JsonResponse({'error': 'ID de post no proporcionado'}, status=400)
        
        # Obtener posts del curso
        posts = Post.objects.filter(curso=curso, is_active=True).order_by('-created_at')
        
        # Filtros
        category_filter = request.GET.get('category')
        if category_filter:
            posts = posts.filter(category=category_filter)
        
        context = {
            'curso': curso,
            'posts': posts,
            'categories': Post.CATEGORY_CHOICES,
            'current_category': category_filter,
        }
        
        # Renderizar solo el contenido del foro
        html = render_to_string('pages/plataforma_foro_content.html', context, request=request)
        return JsonResponse({'html': html})
        
    except Curso.DoesNotExist:
        return JsonResponse({'error': 'El curso no existe'}, status=404)

# ============================================================================
# VISTAS DEL SISTEMA DE CALIFICACIONES
# ============================================================================

@login_required
def plataforma_calificaciones(request, curso_id):
    """
    Vista principal de calificaciones para la plataforma de aprendizaje
    """
    curso = get_object_or_404(Curso, id=curso_id)
    
    # Verificar que el usuario esté inscrito en el curso
    if not request.user.cursos.filter(id=curso_id).exists():
        messages.error(request, 'No tienes acceso a este curso.')
        return redirect('user_space')
    
    context = {
        'curso': curso,
        'user': request.user,
    }
    
    if request.user.is_staff:
        # Vista para Staff/Admin
        evaluaciones = Evaluacion.objects.filter(curso=curso).order_by('-fecha_creacion')
        context['evaluaciones'] = evaluaciones
        
        # Estadísticas generales del curso
        calificaciones_curso = Calificacion.objects.filter(evaluacion__curso=curso, nota__isnull=False)
        if calificaciones_curso.exists():
            estadisticas = {
                'promedio_general': calificaciones_curso.aggregate(Avg('nota'))['nota__avg'],
                'nota_minima': calificaciones_curso.aggregate(Min('nota'))['nota__min'],
                'nota_maxima': calificaciones_curso.aggregate(Max('nota'))['nota__max'],
                'total_estudiantes': curso.usuarios.filter(is_staff=False, is_superuser=False).count(),
            }
            context['estadisticas'] = estadisticas
        
        # Estadísticas de entregas por evaluación
        for evaluacion in evaluaciones:
            total_estudiantes = curso.usuarios.filter(is_staff=False, is_superuser=False).count()
            estudiantes_con_entregas = evaluacion.entregas.values('estudiante').distinct().count()
            estudiantes_calificados = evaluacion.calificaciones.count()
            
            evaluacion.stats_entregas = {
                'total_estudiantes': total_estudiantes,
                'estudiantes_con_entregas': estudiantes_con_entregas,
                'estudiantes_sin_entregas': total_estudiantes - estudiantes_con_entregas,
                'estudiantes_calificados': estudiantes_calificados,
                'estudiantes_pendientes_calificacion': estudiantes_con_entregas - estudiantes_calificados
            }
        
        # Agregar información sobre la nueva funcionalidad
        context['info_entregas'] = {
            'mensaje': '💡 Solo se pueden calificar estudiantes que hayan entregado su trabajo. Desde la primera entrega se puede acceder a calificar una evaluación.',
            'total_evaluaciones': evaluaciones.count(),
            'evaluaciones_con_entregas': sum(1 for e in evaluaciones if e.stats_entregas['estudiantes_con_entregas'] > 0)
            #'evaluaciones_sin_entregas': sum(1 for e in evaluaciones if e.stats_entregas['estudiantes_con_entregas'] == 0)
        }
        
    else:
        # Vista para Estudiantes
        calificaciones_usuario = Calificacion.objects.filter(
            evaluacion__curso=curso,
            estudiante=request.user
        ).order_by('-fecha_calificacion')
        context['calificaciones_usuario'] = calificaciones_usuario
        
        # Verificar si se solicita ver rúbricas
        mostrar_rubricas = request.GET.get('seccion') == 'rubricas'
        context['mostrar_rubricas'] = mostrar_rubricas
        
        if mostrar_rubricas:
            # Obtener todas las evaluaciones del curso que tengan rúbricas
            evaluaciones_con_rubricas = []
            evaluaciones = Evaluacion.objects.filter(curso=curso, activa=True).order_by('fecha_inicio')
            
            for evaluacion in evaluaciones:
                try:
                    rubrica = evaluacion.rubrica
                    if rubrica and rubrica.activa:
                        evaluaciones_con_rubricas.append({
                            'evaluacion': evaluacion,
                            'rubrica': rubrica
                        })
                except:
                    continue
            
            context['evaluaciones_con_rubricas'] = evaluaciones_con_rubricas
        
        # Estadísticas personales del estudiante
        calificaciones_con_nota = calificaciones_usuario.filter(nota__isnull=False)
        total_evaluaciones = Evaluacion.objects.filter(curso=curso).count()
        
        if calificaciones_con_nota.exists():
            # Calcular promedio ponderado
            suma_ponderada = 0
            suma_ponderaciones = 0
            evaluaciones_calificadas = 0
            
            for calificacion in calificaciones_con_nota:
                # Calcular nota ponderada: nota * ponderacion
                nota_ponderada = calificacion.nota * calificacion.evaluacion.ponderacion
                suma_ponderada += nota_ponderada
                suma_ponderaciones += calificacion.evaluacion.ponderacion
                evaluaciones_calificadas += 1
            
            # Calcular promedio ponderado solo si todas las evaluaciones están calificadas y la suma de ponderaciones es 100%
            promedio_ponderado = None
            if evaluaciones_calificadas == total_evaluaciones and suma_ponderaciones == 100:
                promedio_ponderado = suma_ponderada / 100
            
            estadisticas_estudiante = {
                'promedio_ponderado': promedio_ponderado,
                'evaluaciones_calificadas': evaluaciones_calificadas,
                'total_evaluaciones': total_evaluaciones,
                'suma_ponderaciones': suma_ponderaciones,
            }
            context['estadisticas_estudiante'] = estadisticas_estudiante
        
        # Promedios por tipo de evaluación
        promedios_por_tipo = {}
        for calificacion in calificaciones_con_nota:
            tipo = calificacion.evaluacion.get_tipo_display()
            if tipo not in promedios_por_tipo:
                promedios_por_tipo[tipo] = {
                    'notas': [],
                    'ponderaciones': []
                }
            promedios_por_tipo[tipo]['notas'].append(calificacion.nota)
            promedios_por_tipo[tipo]['ponderaciones'].append(calificacion.evaluacion.ponderacion)
        
        # Calcular promedios
        for tipo, datos in promedios_por_tipo.items():
            if datos['notas']:
                datos['promedio'] = sum(datos['notas']) / len(datos['notas'])
                datos['ponderacion_promedio'] = sum(datos['ponderaciones']) / len(datos['ponderaciones'])
        
        context['promedios_por_tipo'] = promedios_por_tipo
    
    return render(request, 'pages/plataforma_calificaciones.html', context)

@login_required
def crear_evaluacion(request, curso_id):
    """
    Vista para crear una nueva evaluación (solo staff/admin)
    """
    curso = get_object_or_404(Curso, id=curso_id)
    
    # Verificar permisos
    if not request.user.is_staff:
        messages.error(request, 'No tienes permisos para crear evaluaciones.')
        return redirect('plataforma_calificaciones', curso_id=curso_id)
    
    if request.method == 'POST':
        form = EvaluacionForm(request.POST, curso=curso)
        if form.is_valid():
            evaluacion = form.save(commit=False)
            evaluacion.curso = curso
            evaluacion.creado_por = request.user
            evaluacion.save()
            messages.success(request, f'Evaluación "{evaluacion.nombre}" creada exitosamente.')
            return redirect('plataforma_calificaciones', curso_id=curso_id)
    else:
        form = EvaluacionForm(curso=curso)
    
    context = {
        'curso': curso,
        'form': form,
        'user': request.user,
    }
    return render(request, 'pages/crear_evaluacion.html', context)

@login_required
def calificar_estudiante(request, curso_id, evaluacion_id):
    """
    Vista para calificar estudiantes en una evaluación específica (solo staff/admin)
    Permite calificar estudiantes que tienen entregas para esta evaluación
    También permite editar calificaciones existentes
    """
    curso = get_object_or_404(Curso, id=curso_id)
    evaluacion = get_object_or_404(Evaluacion, id=evaluacion_id, curso=curso)
    
    # Verificar permisos
    if not request.user.is_staff:
        messages.error(request, 'No tienes permisos para calificar estudiantes.')
        return redirect('plataforma_calificaciones', curso_id=curso_id)
    
    # Obtener estudiantes del curso que tienen entregas para esta evaluación (no staff/admin)
    estudiantes_con_entregas = User.objects.filter(
        cursos=curso,
        is_staff=False,
        is_superuser=False,
        entregas__evaluacion=evaluacion
    ).distinct().order_by('first_name', 'last_name', 'username')
    
    # Para el formulario: solo estudiantes no calificados
    estudiantes_calificados_ids = evaluacion.calificaciones.values_list('estudiante_id', flat=True)
    estudiantes_disponibles_para_calificar = estudiantes_con_entregas.exclude(id__in=estudiantes_calificados_ids)
    
    # Si no hay estudiantes con entregas, mostrar mensaje
    if not estudiantes_con_entregas.exists():
        messages.warning(request, f'No hay estudiantes con entregas para la evaluación "{evaluacion.nombre}". Solo se pueden calificar estudiantes que hayan entregado su trabajo.')
        return redirect('plataforma_calificaciones', curso_id=curso_id)
    
    if request.method == 'POST':
        form = CalificacionForm(request.POST, curso=curso, evaluacion=evaluacion, estudiantes_con_entregas=estudiantes_disponibles_para_calificar)
        if form.is_valid():
            # Verificar que el estudiante tenga una entrega para esta evaluación
            estudiante = form.cleaned_data.get('estudiante')
            if estudiante:
                tiene_entrega = estudiante.entregas.filter(evaluacion=evaluacion).exists()
                if not tiene_entrega:
                    messages.error(request, f'El estudiante {estudiante.get_full_name()} no tiene entregas para la evaluación "{evaluacion.nombre}". Solo se pueden calificar estudiantes que hayan entregado su trabajo.')
                    return redirect('calificar_estudiante', curso_id=curso_id, evaluacion_id=evaluacion_id)
            
            # Verificar que la evaluación tenga una rúbrica
            if not hasattr(evaluacion, 'rubrica') or not evaluacion.rubrica:
                messages.error(request, f'La evaluación "{evaluacion.nombre}" no tiene una rúbrica asociada. Debes crear una rúbrica antes de calificar.')
                return redirect('plataforma_calificaciones', curso_id=curso_id)
            
            calificacion = form.save(commit=False)
            calificacion.evaluacion = evaluacion
            calificacion.calificado_por = request.user
            
            # Verificar que no exista ya una calificación para este estudiante en esta evaluación
            calificacion_existente = Calificacion.objects.filter(
                evaluacion=evaluacion,
                estudiante=calificacion.estudiante
            ).first()
            
            # Limpiar caracteres de control de línea en la retroalimentación
            if calificacion.retroalimentacion:
                # Reemplazar caracteres de control con saltos de línea normales
                retroalimentacion_limpia = re.sub(r'\\u000D\\u000A', '\n', calificacion.retroalimentacion)
                retroalimentacion_limpia = re.sub(r'\\u000D', '\n', retroalimentacion_limpia)
                retroalimentacion_limpia = re.sub(r'\\u000A', '\n', retroalimentacion_limpia)
                retroalimentacion_limpia = re.sub(r'\\n', '\n', retroalimentacion_limpia)
                retroalimentacion_limpia = re.sub(r'\\r', '\n', retroalimentacion_limpia)
            else:
                retroalimentacion_limpia = calificacion.retroalimentacion
            
            try:
                if calificacion_existente:
                    # Actualizar calificación existente
                    calificacion_existente.retroalimentacion = retroalimentacion_limpia
                    calificacion_existente.calificado_por = request.user
                    calificacion_existente.save()
                    
                    # Actualizar o crear ResultadoRubrica
                    resultado_rubrica, created = ResultadoRubrica.objects.get_or_create(
                        rubrica=evaluacion.rubrica,
                        estudiante=calificacion_existente.estudiante,
                        defaults={
                            'evaluador': request.user,
                            'puntaje_total': 0
                        }
                    )
                    
                    # Actualizar puntajes de criterios
                    puntaje_total = 0
                    for criterio in evaluacion.rubrica.criterios.all():
                        field_name = f'criterio_{criterio.id}'
                        esperable_id = form.cleaned_data.get(field_name)
                        
                        if esperable_id:
                            esperable = Esperable.objects.get(id=esperable_id)
                            puntaje_criterio, created = PuntajeCriterio.objects.get_or_create(
                                resultado_rubrica=resultado_rubrica,
                                criterio=criterio,
                                defaults={
                                    'esperable_seleccionado': esperable,
                                    'puntaje_obtenido': esperable.puntaje
                                }
                            )
                            
                            if not created:
                                puntaje_criterio.esperable_seleccionado = esperable
                                puntaje_criterio.puntaje_obtenido = esperable.puntaje
                                puntaje_criterio.save()
                            
                            puntaje_total += esperable.puntaje
                    
                    # Calcular la nota según la fórmula
                    suma_puntajes_maximos = sum(criterio.puntaje for criterio in evaluacion.rubrica.criterios.all())
                    if suma_puntajes_maximos > 0:
                        nota_calculada = (evaluacion.nota_maxima * puntaje_total) / suma_puntajes_maximos
                        # Usar la nota del formulario si está disponible, sino usar la calculada
                        nota_final = form.cleaned_data.get('nota')
                        if nota_final is not None and nota_final != '':
                            calificacion_existente.nota = round(float(nota_final), 1)
                        else:
                            calificacion_existente.nota = round(nota_calculada, 1)
                        calificacion_existente.save()
                    
                    # Actualizar puntaje total
                    resultado_rubrica.puntaje_total = puntaje_total
                    resultado_rubrica.save()
                    
                    # Recalcular puntaje total usando el método del modelo
                    resultado_rubrica.calcular_puntaje_total()
                    
                    messages.success(request, f'Calificación actualizada para {calificacion.estudiante.get_full_name()}.')
                else:
                    # Crear nueva calificación
                    calificacion.retroalimentacion = retroalimentacion_limpia
                    calificacion.save()
                    
                    # Crear ResultadoRubrica
                    resultado_rubrica = ResultadoRubrica.objects.create(
                        rubrica=evaluacion.rubrica,
                        estudiante=calificacion.estudiante,
                        evaluador=request.user,
                        puntaje_total=0
                    )
                    
                    # Crear puntajes de criterios
                    puntaje_total = 0
                    for criterio in evaluacion.rubrica.criterios.all():
                        field_name = f'criterio_{criterio.id}'
                        esperable_id = form.cleaned_data.get(field_name)
                        
                        if esperable_id:
                            esperable = Esperable.objects.get(id=esperable_id)
                            PuntajeCriterio.objects.create(
                                resultado_rubrica=resultado_rubrica,
                                criterio=criterio,
                                esperable_seleccionado=esperable,
                                puntaje_obtenido=esperable.puntaje
                            )
                            puntaje_total += esperable.puntaje
                    
                    # Calcular la nota según la fórmula
                    suma_puntajes_maximos = sum(criterio.puntaje for criterio in evaluacion.rubrica.criterios.all())
                    if suma_puntajes_maximos > 0:
                        nota_calculada = (evaluacion.nota_maxima * puntaje_total) / suma_puntajes_maximos
                        # Usar la nota del formulario si está disponible, sino usar la calculada
                        nota_final = form.cleaned_data.get('nota')
                        if nota_final is not None and nota_final != '':
                            calificacion.nota = round(float(nota_final), 1)
                        else:
                            calificacion.nota = round(nota_calculada, 1)
                        calificacion.save()
                    
                    # Actualizar puntaje total
                    resultado_rubrica.puntaje_total = puntaje_total
                    resultado_rubrica.save()
                    
                    # Recalcular puntaje total usando el método del modelo
                    resultado_rubrica.calcular_puntaje_total()
                    
                    messages.success(request, f'Calificación registrada para {calificacion.estudiante.get_full_name()}.')
                
                return redirect('plataforma_calificaciones', curso_id=curso_id)
                
            except Exception as e:
                messages.error(request, f'Error al guardar la calificación: {str(e)}')
                return redirect('calificar_estudiante', curso_id=curso_id, evaluacion_id=evaluacion_id)
    else:
        form = CalificacionForm(curso=curso, evaluacion=evaluacion, estudiantes_con_entregas=estudiantes_disponibles_para_calificar)
        
        # Si se especifica un estudiante en la URL, precargarlo en el formulario
        estudiante_id = request.GET.get('estudiante')
        if estudiante_id:
            try:
                estudiante = estudiantes_disponibles_para_calificar.get(id=estudiante_id)
                if estudiante:
                    form.initial['estudiante'] = estudiante
                    logger.debug(f"Estudiante precargado: {estudiante.get_full_name()} (ID: {estudiante.id})")
                else:
                    logger.debug(f"Estudiante con ID {estudiante_id} no está disponible para calificar")
            except User.DoesNotExist:
                logger.debug(f"Estudiante no encontrado con ID: {estudiante_id}")
                pass
    
    # Obtener todos los estudiantes del curso para mostrar en la lista
    todos_estudiantes = User.objects.filter(
        cursos=curso,
        is_staff=False,
        is_superuser=False
    ).order_by('first_name', 'last_name', 'username')
    

    
    # Obtener estudiantes ya calificados para mostrar estadísticas
    estudiantes_calificados_ids = evaluacion.calificaciones.values_list('estudiante_id', flat=True)
    estudiantes_ya_calificados = estudiantes_con_entregas.filter(id__in=estudiantes_calificados_ids)
    

    
    # Estadísticas para mostrar en el template
    stats = {
        'total_estudiantes_con_entregas': estudiantes_con_entregas.count(),
        'estudiantes_disponibles_para_calificar': estudiantes_disponibles_para_calificar.count(),
        'estudiantes_ya_calificados': estudiantes_ya_calificados.count(),
        'total_estudiantes_curso': todos_estudiantes.count(),
    }
    
    context = {
        'curso': curso,
        'evaluacion': evaluacion,
        'form': form,
        'estudiantes': estudiantes_disponibles_para_calificar,  # Para el formulario (solo no calificados)
        'todos_estudiantes': todos_estudiantes,  # Para la lista completa
        'estudiantes_ya_calificados': estudiantes_ya_calificados,  # Para mostrar cuáles ya están calificados
        'stats': stats,
        'user': request.user,
    }
    
    # Agregar criterios de rúbrica al contexto si existe
    if evaluacion and hasattr(evaluacion, 'rubrica') and evaluacion.rubrica:
        context['criterios_rubrica'] = form.criterios_rubrica if hasattr(form, 'criterios_rubrica') else evaluacion.rubrica.criterios.all()
        if hasattr(form, 'criterios_info'):
            context['criterios_info'] = form.criterios_info
    return render(request, 'pages/calificar_estudiante.html', context)

@login_required
def ver_calificacion_detalle(request, curso_id, calificacion_id):
    """
    Vista para ver el detalle de una calificación específica
    """
    curso = get_object_or_404(Curso, id=curso_id)
    calificacion = get_object_or_404(Calificacion, id=calificacion_id)
    
    # Verificar que el usuario tenga acceso a esta calificación
    if not request.user.is_staff and calificacion.estudiante != request.user:
        messages.error(request, 'No tienes acceso a esta calificación.')
        return redirect('plataforma_calificaciones', curso_id=curso_id)
    
    # Verificar que la calificación pertenezca al curso
    if calificacion.evaluacion.curso != curso:
        messages.error(request, 'La calificación no pertenece a este curso.')
        return redirect('plataforma_calificaciones', curso_id=curso_id)
    
    # Obtener la rúbrica asociada a la evaluación si existe
    rubrica = None
    resultado_rubrica = None
    criterios_con_puntajes = []
    
    try:
        rubrica = calificacion.evaluacion.rubrica
        if rubrica:
            # Obtener el resultado de la rúbrica para este estudiante
            try:
                resultado_rubrica = ResultadoRubrica.objects.get(
                    rubrica=rubrica,
                    estudiante=calificacion.estudiante
                )
                
                # Obtener los criterios con sus puntajes
                for criterio in rubrica.criterios.all():
                    try:
                        puntaje_criterio = PuntajeCriterio.objects.get(
                            resultado_rubrica=resultado_rubrica,
                            criterio=criterio
                        )
                        
                        # Determinar el esperable correcto basado en el puntaje obtenido
                        esperable_correcto = None
                        puntaje_obtenido = puntaje_criterio.puntaje_obtenido
                        
                        # Buscar coincidencia exacta primero
                        for esperable in criterio.esperables.all():
                            if esperable.puntaje == puntaje_obtenido:
                                esperable_correcto = esperable
                                break
                        
                        # Si no hay coincidencia exacta, buscar el esperable más cercano
                        if esperable_correcto is None:
                            esperables_ordenados = list(criterio.esperables.all().order_by('puntaje'))
                            for i, esperable in enumerate(esperables_ordenados):
                                if esperable.puntaje >= puntaje_obtenido:
                                    esperable_correcto = esperable
                                    break
                            # Si no se encontró ninguno mayor o igual, usar el último
                            if esperable_correcto is None and esperables_ordenados:
                                esperable_correcto = esperables_ordenados[-1]
                        
                        criterios_con_puntajes.append({
                            'criterio': criterio,
                            'puntaje_obtenido': puntaje_criterio.puntaje_obtenido,
                            'esperable_seleccionado': esperable_correcto,
                            'comentarios': puntaje_criterio.comentarios,
                        })
                    except PuntajeCriterio.DoesNotExist:
                        criterios_con_puntajes.append({
                            'criterio': criterio,
                            'puntaje_obtenido': 0,
                            'esperable_seleccionado': None,
                            'comentarios': '',
                        })
            except ResultadoRubrica.DoesNotExist:
                # Si no hay resultado de rúbrica, mostrar solo los criterios sin puntajes
                for criterio in rubrica.criterios.all():
                    criterios_con_puntajes.append({
                        'criterio': criterio,
                        'puntaje_obtenido': 0,
                        'esperable_seleccionado': None,
                        'comentarios': '',
                    })
    except:
        pass
    
    context = {
        'curso': curso,
        'calificacion': calificacion,
        'rubrica': rubrica,
        'resultado_rubrica': resultado_rubrica,
        'criterios_con_puntajes': criterios_con_puntajes,
        'user': request.user,
    }
    return render(request, 'pages/calificacion_detalle.html', context)

@login_required
def estadisticas_curso(request, curso_id):
    """
    Vista para ver estadísticas detalladas del curso (solo staff/admin)
    """
    curso = get_object_or_404(Curso, id=curso_id)
    
    # Verificar permisos
    if not request.user.is_staff:
        messages.error(request, 'No tienes permisos para ver estadísticas del curso.')
        return redirect('plataforma_calificaciones', curso_id=curso_id)
    
    # Obtener todas las calificaciones del curso
    calificaciones = Calificacion.objects.filter(evaluacion__curso=curso, nota__isnull=False)
    
    # Estadísticas generales
    estadisticas = {
        'total_estudiantes': curso.usuarios.filter(is_staff=False, is_superuser=False).count(),
        'total_evaluaciones': Evaluacion.objects.filter(curso=curso).count(),
        'total_calificaciones': calificaciones.count(),
    }
    
    if calificaciones.exists():
        estadisticas.update({
            'promedio_general': calificaciones.aggregate(Avg('nota'))['nota__avg'],
            'nota_minima': calificaciones.aggregate(Min('nota'))['nota__min'],
            'nota_maxima': calificaciones.aggregate(Max('nota'))['nota__max'],
        })
    
    # Estadísticas por evaluación
    evaluaciones = Evaluacion.objects.filter(curso=curso).order_by('fecha_inicio')
    estadisticas_evaluaciones = []
    
    for evaluacion in evaluaciones:
        calificaciones_eval = calificaciones.filter(evaluacion=evaluacion)
        stats = {
            'evaluacion': evaluacion,
            'total_calificaciones': calificaciones_eval.count(),
            'promedio': calificaciones_eval.aggregate(Avg('nota'))['nota__avg'] if calificaciones_eval.exists() else None,
            'nota_minima': calificaciones_eval.aggregate(Min('nota'))['nota__min'] if calificaciones_eval.exists() else None,
            'nota_maxima': calificaciones_eval.aggregate(Max('nota'))['nota__max'] if calificaciones_eval.exists() else None,
        }
        estadisticas_evaluaciones.append(stats)
    
    # Estadísticas por estudiante
    estudiantes = curso.usuarios.filter(is_staff=False, is_superuser=False).order_by('first_name', 'last_name')
    estadisticas_estudiantes = []
    
    for estudiante in estudiantes:
        calificaciones_est = calificaciones.filter(estudiante=estudiante)
        stats = {
            'estudiante': estudiante,
            'total_calificaciones': calificaciones_est.count(),
            'promedio': calificaciones_est.aggregate(Avg('nota'))['nota__avg'] if calificaciones_est.exists() else None,
            'nota_minima': calificaciones_est.aggregate(Min('nota'))['nota__min'] if calificaciones_est.exists() else None,
            'nota_maxima': calificaciones_est.aggregate(Max('nota'))['nota__max'] if calificaciones_est.exists() else None,
        }
        estadisticas_estudiantes.append(stats)
    
    context = {
        'curso': curso,
        'estadisticas': estadisticas,
        'estadisticas_evaluaciones': estadisticas_evaluaciones,
        'estadisticas_estudiantes': estadisticas_estudiantes,
        'user': request.user,
    }
    return render(request, 'pages/estadisticas_curso.html', context)

@login_required
def exportar_calificaciones_excel(request, curso_id):
    """
    Vista para exportar todas las calificaciones del curso a un archivo Excel
    con información detallada de rúbricas, criterios y esperables
    """
    
    curso = get_object_or_404(Curso, id=curso_id)
    
    # Verificar permisos
    if not request.user.is_staff:
        messages.error(request, 'No tienes permisos para exportar calificaciones.')
        return redirect('plataforma_calificaciones', curso_id=curso_id)
    
    # Obtener todas las calificaciones del curso con información relacionada
    calificaciones = Calificacion.objects.filter(
        evaluacion__curso=curso
    ).select_related(
        'estudiante', 
        'evaluacion', 
        'calificado_por'
    ).prefetch_related(
        'evaluacion__rubrica__criterios__esperables'
    ).order_by('estudiante__first_name', 'estudiante__last_name', 'evaluacion__fecha_inicio')
    
    # Crear un nuevo libro de trabajo
    wb = Workbook()
    
    # Hoja 1: Resumen de calificaciones
    ws_resumen = wb.active
    ws_resumen.title = "Resumen Calificaciones"
    
    # Definir estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Definir encabezados para el resumen
    headers_resumen = [
        'ID Estudiante',
        'Nombre Completo',
        'Email',
        'ID Evaluación',
        'Nombre Evaluación',
        'Tipo Evaluación',
        'Fecha Inicio',
        'Fecha Fin',
        'Nota Máxima',
        'Ponderación (%)',
        'Nota Obtenida',
        'Porcentaje Obtenido',
        'Nota Ponderada',
        'Retroalimentación',
        'Calificado Por',
        'Fecha Calificación',
        'Fecha Modificación'
    ]
    
    # Escribir encabezados del resumen
    for col, header in enumerate(headers_resumen, 1):
        cell = ws_resumen.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Escribir datos del resumen
    row = 2
    for calificacion in calificaciones:
        # Información del estudiante
        ws_resumen.cell(row=row, column=1, value=calificacion.estudiante.id)
        ws_resumen.cell(row=row, column=2, value=calificacion.estudiante.get_full_name() or calificacion.estudiante.username)
        ws_resumen.cell(row=row, column=3, value=calificacion.estudiante.email)
        
        # Información de la evaluación
        ws_resumen.cell(row=row, column=4, value=calificacion.evaluacion.id)
        ws_resumen.cell(row=row, column=5, value=calificacion.evaluacion.nombre)
        ws_resumen.cell(row=row, column=6, value=calificacion.evaluacion.get_tipo_display())
        ws_resumen.cell(row=row, column=7, value=calificacion.evaluacion.fecha_inicio.strftime('%d/%m/%Y') if calificacion.evaluacion.fecha_inicio else '')
        ws_resumen.cell(row=row, column=8, value=calificacion.evaluacion.fecha_fin.strftime('%d/%m/%Y') if calificacion.evaluacion.fecha_fin else '')
        ws_resumen.cell(row=row, column=9, value=float(calificacion.evaluacion.nota_maxima))
        ws_resumen.cell(row=row, column=10, value=float(calificacion.evaluacion.ponderacion))
        
        # Información de la calificación
        ws_resumen.cell(row=row, column=11, value=float(calificacion.nota) if calificacion.nota else '')
        
        # Calcular porcentaje obtenido
        if calificacion.nota and calificacion.evaluacion.nota_maxima:
            porcentaje = (calificacion.nota / calificacion.evaluacion.nota_maxima) * 100
            ws_resumen.cell(row=row, column=12, value=round(porcentaje, 2))
        else:
            ws_resumen.cell(row=row, column=12, value='')
        
        # Calcular nota ponderada
        if calificacion.nota and calificacion.evaluacion.ponderacion:
            nota_ponderada = (calificacion.nota / calificacion.evaluacion.nota_maxima) * calificacion.evaluacion.ponderacion
            ws_resumen.cell(row=row, column=13, value=round(nota_ponderada, 2))
        else:
            ws_resumen.cell(row=row, column=13, value='')
        
        # Información adicional
        ws_resumen.cell(row=row, column=14, value=calificacion.retroalimentacion or '')
        ws_resumen.cell(row=row, column=15, value=calificacion.calificado_por.get_full_name() or calificacion.calificado_por.username)
        ws_resumen.cell(row=row, column=16, value=calificacion.fecha_calificacion.strftime('%d/%m/%Y %H:%M'))
        ws_resumen.cell(row=row, column=17, value=calificacion.fecha_modificacion.strftime('%d/%m/%Y %H:%M'))
        
        row += 1
    
    # Ajustar ancho de columnas del resumen
    for column in ws_resumen.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_resumen.column_dimensions[column_letter].width = adjusted_width
    
    # Hoja 2: Detalle de rúbricas y criterios
    ws_detalle = wb.create_sheet("Detalle Rúbricas")
    
    # Encabezados para el detalle de rúbricas
    headers_detalle = [
        'ID Estudiante',
        'Nombre Estudiante',
        'Email Estudiante',
        'ID Evaluación',
        'Nombre Evaluación',
        'ID Rúbrica',
        'Nombre Rúbrica',
        'Objetivo Rúbrica',
        'Aprendizaje Esperado',
        'ID Criterio',
        'Nombre Criterio',
        'Objetivo Criterio',
        'Puntaje Máximo Criterio',
        'ID Esperable',
        'Nivel Esperable',
        'Descripción Esperable',
        'Puntaje Esperable',
        'Esperable Seleccionado',
        'Puntaje Obtenido',
        'Comentarios Criterio',
        'Evaluador',
        'Fecha Evaluación'
    ]
    
    # Escribir encabezados del detalle
    for col, header in enumerate(headers_detalle, 1):
        cell = ws_detalle.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Escribir datos del detalle
    row_detalle = 2
    
    for calificacion in calificaciones:
        evaluacion = calificacion.evaluacion
        
        # Verificar si la evaluación tiene rúbrica
        if hasattr(evaluacion, 'rubrica') and evaluacion.rubrica:
            rubrica = evaluacion.rubrica
            
            # Obtener resultado de rúbrica para este estudiante
            try:
                resultado_rubrica = ResultadoRubrica.objects.get(
                    rubrica=rubrica,
                    estudiante=calificacion.estudiante
                )
                
                # Para cada criterio en la rúbrica
                for criterio in rubrica.criterios.all():
                    # Obtener puntaje del criterio para este estudiante
                    try:
                        puntaje_criterio = PuntajeCriterio.objects.get(
                            resultado_rubrica=resultado_rubrica,
                            criterio=criterio
                        )
                        esperable_seleccionado = puntaje_criterio.esperable_seleccionado
                        puntaje_obtenido = puntaje_criterio.puntaje_obtenido
                        comentarios = puntaje_criterio.comentarios
                    except PuntajeCriterio.DoesNotExist:
                        esperable_seleccionado = None
                        puntaje_obtenido = 0
                        comentarios = ''
                    
                    # Para cada esperable del criterio
                    for esperable in criterio.esperables.all():
                        # Información del estudiante
                        ws_detalle.cell(row=row_detalle, column=1, value=calificacion.estudiante.id)
                        ws_detalle.cell(row=row_detalle, column=2, value=calificacion.estudiante.get_full_name() or calificacion.estudiante.username)
                        ws_detalle.cell(row=row_detalle, column=3, value=calificacion.estudiante.email)
                        
                        # Información de la evaluación
                        ws_detalle.cell(row=row_detalle, column=4, value=evaluacion.id)
                        ws_detalle.cell(row=row_detalle, column=5, value=evaluacion.nombre)
                        
                        # Información de la rúbrica
                        ws_detalle.cell(row=row_detalle, column=6, value=rubrica.id)
                        ws_detalle.cell(row=row_detalle, column=7, value=rubrica.nombre)
                        ws_detalle.cell(row=row_detalle, column=8, value=rubrica.objetivo)
                        ws_detalle.cell(row=row_detalle, column=9, value=rubrica.aprendizaje_esperado)
                        
                        # Información del criterio
                        ws_detalle.cell(row=row_detalle, column=10, value=criterio.id)
                        ws_detalle.cell(row=row_detalle, column=11, value=criterio.nombre)
                        ws_detalle.cell(row=row_detalle, column=12, value=criterio.objetivo)
                        ws_detalle.cell(row=row_detalle, column=13, value=float(criterio.puntaje))
                        
                        # Información del esperable
                        ws_detalle.cell(row=row_detalle, column=14, value=esperable.id)
                        ws_detalle.cell(row=row_detalle, column=15, value=esperable.nivel)
                        ws_detalle.cell(row=row_detalle, column=16, value=esperable.descripcion)
                        ws_detalle.cell(row=row_detalle, column=17, value=float(esperable.puntaje))
                        
                        # Información del esperable seleccionado
                        if esperable_seleccionado:
                            ws_detalle.cell(row=row_detalle, column=18, value=f"{esperable_seleccionado.nivel} - {esperable_seleccionado.descripcion}")
                        else:
                            ws_detalle.cell(row=row_detalle, column=18, value='No seleccionado')
                        
                        ws_detalle.cell(row=row_detalle, column=19, value=float(puntaje_obtenido))
                        ws_detalle.cell(row=row_detalle, column=20, value=comentarios or '')
                        
                        # Información del evaluador
                        ws_detalle.cell(row=row_detalle, column=21, value=resultado_rubrica.evaluador.get_full_name() or resultado_rubrica.evaluador.username)
                        ws_detalle.cell(row=row_detalle, column=22, value=resultado_rubrica.fecha_evaluacion.strftime('%d/%m/%Y %H:%M'))
                        
                        row_detalle += 1
                        
            except ResultadoRubrica.DoesNotExist:
                # Si no hay resultado de rúbrica, crear filas con información básica
                for criterio in rubrica.criterios.all():
                    for esperable in criterio.esperables.all():
                        # Información del estudiante
                        ws_detalle.cell(row=row_detalle, column=1, value=calificacion.estudiante.id)
                        ws_detalle.cell(row=row_detalle, column=2, value=calificacion.estudiante.get_full_name() or calificacion.estudiante.username)
                        ws_detalle.cell(row=row_detalle, column=3, value=calificacion.estudiante.email)
                        
                        # Información de la evaluación
                        ws_detalle.cell(row=row_detalle, column=4, value=evaluacion.id)
                        ws_detalle.cell(row=row_detalle, column=5, value=evaluacion.nombre)
                        
                        # Información de la rúbrica
                        ws_detalle.cell(row=row_detalle, column=6, value=rubrica.id)
                        ws_detalle.cell(row=row_detalle, column=7, value=rubrica.nombre)
                        ws_detalle.cell(row=row_detalle, column=8, value=rubrica.objetivo)
                        ws_detalle.cell(row=row_detalle, column=9, value=rubrica.aprendizaje_esperado)
                        
                        # Información del criterio
                        ws_detalle.cell(row=row_detalle, column=10, value=criterio.id)
                        ws_detalle.cell(row=row_detalle, column=11, value=criterio.nombre)
                        ws_detalle.cell(row=row_detalle, column=12, value=criterio.objetivo)
                        ws_detalle.cell(row=row_detalle, column=13, value=float(criterio.puntaje))
                        
                        # Información del esperable
                        ws_detalle.cell(row=row_detalle, column=14, value=esperable.id)
                        ws_detalle.cell(row=row_detalle, column=15, value=esperable.nivel)
                        ws_detalle.cell(row=row_detalle, column=16, value=esperable.descripcion)
                        ws_detalle.cell(row=row_detalle, column=17, value=float(esperable.puntaje))
                        
                        # Sin esperable seleccionado
                        ws_detalle.cell(row=row_detalle, column=18, value='No evaluado')
                        ws_detalle.cell(row=row_detalle, column=19, value=0)
                        ws_detalle.cell(row=row_detalle, column=20, value='')
                        ws_detalle.cell(row=row_detalle, column=21, value='')
                        ws_detalle.cell(row=row_detalle, column=22, value='')
                        
                        row_detalle += 1
    
    # Ajustar ancho de columnas del detalle
    for column in ws_detalle.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_detalle.column_dimensions[column_letter].width = adjusted_width
    
    # Crear respuesta HTTP con encoding correcto
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Crear nombre de archivo seguro
    nombre_archivo = f"calificaciones_detalladas_{curso.nombre.replace(' ', '_').replace('/', '_').replace(chr(92), '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    
    # Guardar el archivo en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Escribir el contenido en la respuesta
    response.write(output.getvalue())
    
    return response

@login_required
def eliminar_evaluacion(request, curso_id, evaluacion_id):
    """
    Vista para eliminar una evaluación (solo staff/admin)
    """
    curso = get_object_or_404(Curso, id=curso_id)
    evaluacion = get_object_or_404(Evaluacion, id=evaluacion_id, curso=curso)
    
    # Verificar permisos
    if not request.user.is_staff:
        messages.error(request, 'No tienes permisos para eliminar evaluaciones.')
        return redirect('plataforma_calificaciones', curso_id=curso_id)
    
    # Verificar si la evaluación tiene calificaciones
    tiene_calificaciones = evaluacion.calificaciones.exists()
    
    if request.method == 'POST':
        # Si no tiene calificaciones, eliminar directamente
        if not tiene_calificaciones:
            nombre_evaluacion = evaluacion.nombre
            evaluacion.delete()
            messages.success(request, f'Evaluación "{nombre_evaluacion}" eliminada exitosamente.')
            return redirect('plataforma_calificaciones', curso_id=curso_id)
        
        # Si tiene calificaciones, verificar confirmación
        confirmacion = request.POST.get('confirmacion', '').strip()
        if confirmacion != 'ELIMINAR':
            messages.error(request, 'Debes escribir "ELIMINAR" para confirmar la eliminación.')
            context = {
                'curso': curso,
                'evaluacion': evaluacion,
                'user': request.user,
                'tiene_calificaciones': tiene_calificaciones,
            }
            return render(request, 'pages/eliminar_evaluacion.html', context)
        
        # Eliminar la evaluación y todas sus calificaciones asociadas
        nombre_evaluacion = evaluacion.nombre
        evaluacion.delete()
        messages.success(request, f'Evaluación "{nombre_evaluacion}" eliminada exitosamente.')
        return redirect('plataforma_calificaciones', curso_id=curso_id)
    
    # Si es GET, mostrar página de confirmación
    context = {
        'curso': curso,
        'evaluacion': evaluacion,
        'user': request.user,
        'tiene_calificaciones': tiene_calificaciones,
    }
    return render(request, 'pages/eliminar_evaluacion.html', context)

@login_required
def editar_evaluacion(request, curso_id, evaluacion_id):
    """
    Vista para editar una evaluación (solo staff/admin)
    """
    curso = get_object_or_404(Curso, id=curso_id)
    evaluacion = get_object_or_404(Evaluacion, id=evaluacion_id, curso=curso)
    
    # Verificar permisos
    if not request.user.is_staff:
        messages.error(request, 'No tienes permisos para editar evaluaciones.')
        return redirect('plataforma_calificaciones', curso_id=curso_id)
    
    if request.method == 'POST':
        form = EvaluacionForm(request.POST, instance=evaluacion, curso=curso)
        if form.is_valid():
            # Para editar, guardar con commit=False para asignar el curso
            evaluacion = form.save(commit=False)
            evaluacion.curso = curso
            evaluacion.save()
            messages.success(request, f'Evaluación "{evaluacion.nombre}" actualizada exitosamente.')
            return redirect('plataforma_calificaciones', curso_id=curso_id)
    else:
        form = EvaluacionForm(instance=evaluacion, curso=curso)
    
    context = {
        'curso': curso,
        'evaluacion': evaluacion,
        'form': form,
        'user': request.user,
    }
    return render(request, 'pages/editar_evaluacion.html', context)

@login_required
def editar_calificacion(request):
    """
    Vista para editar una calificación existente (solo staff/admin)
    """
    if request.method != 'POST':
        messages.error(request, 'Método no permitido.')
        return redirect('user_space')
    
    # Verificar permisos
    if not request.user.is_staff:
        messages.error(request, 'No tienes permisos para editar calificaciones.')
        return redirect('user_space')
    
    # Obtener datos del formulario
    calificacion_id = request.POST.get('calificacion_id')
    estudiante_id = request.POST.get('estudiante_id')
    evaluacion_id = request.POST.get('evaluacion_id')
    nueva_nota = request.POST.get('nota')
    retroalimentacion = request.POST.get('retroalimentacion', '')
    
    try:
        # Obtener objetos
        calificacion = get_object_or_404(Calificacion, id=calificacion_id)
        estudiante = get_object_or_404(User, id=estudiante_id)
        evaluacion = get_object_or_404(Evaluacion, id=evaluacion_id)
        
        # Verificar que la calificación pertenezca a la evaluación correcta
        if calificacion.evaluacion != evaluacion:
            messages.error(request, 'La calificación no pertenece a esta evaluación.')
            return redirect('plataforma_calificaciones', curso_id=evaluacion.curso.id)
        
        # Verificar que el estudiante esté inscrito en el curso
        if estudiante not in evaluacion.curso.usuarios.all():
            messages.error(request, 'El estudiante no está inscrito en este curso.')
            return redirect('plataforma_calificaciones', curso_id=evaluacion.curso.id)
        
        # Validar la nota
        try:
            nueva_nota = float(nueva_nota)
            if nueva_nota < 0 or nueva_nota > evaluacion.nota_maxima:
                messages.error(request, f'La nota debe estar entre 0 y {evaluacion.nota_maxima}.')
                return redirect('calificar_estudiante', curso_id=evaluacion.curso.id, evaluacion_id=evaluacion.id)
        except (ValueError, TypeError):
            messages.error(request, 'La nota debe ser un número válido.')
            return redirect('calificar_estudiante', curso_id=evaluacion.curso.id, evaluacion_id=evaluacion.id)
        
        # Limpiar caracteres de control de línea en la retroalimentación
        if retroalimentacion:
            # Reemplazar caracteres de control con saltos de línea normales
            retroalimentacion_limpia = re.sub(r'\\u000D\\u000A', '\n', retroalimentacion)
            retroalimentacion_limpia = re.sub(r'\\u000D', '\n', retroalimentacion_limpia)
            retroalimentacion_limpia = re.sub(r'\\u000A', '\n', retroalimentacion_limpia)
            retroalimentacion_limpia = re.sub(r'\\n', '\n', retroalimentacion_limpia)
            retroalimentacion_limpia = re.sub(r'\\r', '\n', retroalimentacion_limpia)
        else:
            retroalimentacion_limpia = retroalimentacion
        
        # Verificar que la evaluación tenga rúbrica
        if not hasattr(evaluacion, 'rubrica') or not evaluacion.rubrica:
            messages.error(request, f'La evaluación "{evaluacion.nombre}" no tiene una rúbrica asociada.')
            return redirect('calificar_estudiante', curso_id=evaluacion.curso.id, evaluacion_id=evaluacion.id)
        
        # Obtener o crear ResultadoRubrica
        resultado_rubrica, created = ResultadoRubrica.objects.get_or_create(
            rubrica=evaluacion.rubrica,
            estudiante=estudiante,
            defaults={
                'evaluador': request.user,
                'puntaje_total': 0
            }
        )
        
        if not created:
            resultado_rubrica.evaluador = request.user
            resultado_rubrica.save()
        
        # Actualizar puntajes de criterios
        puntaje_total = 0
        for criterio in evaluacion.rubrica.criterios.all():
            field_name = f'criterio_{criterio.id}'
            esperable_id = request.POST.get(field_name)
            
            if esperable_id:
                esperable = Esperable.objects.get(id=esperable_id)
                puntaje_criterio, created = PuntajeCriterio.objects.get_or_create(
                    resultado_rubrica=resultado_rubrica,
                    criterio=criterio,
                    defaults={
                        'esperable_seleccionado': esperable,
                        'puntaje_obtenido': esperable.puntaje
                    }
                )
                
                if not created:
                    puntaje_criterio.esperable_seleccionado = esperable
                    puntaje_criterio.puntaje_obtenido = esperable.puntaje
                    puntaje_criterio.save()
                
                puntaje_total += esperable.puntaje
        
        # Calcular la nota según la fórmula
        suma_puntajes_maximos = sum(criterio.puntaje for criterio in evaluacion.rubrica.criterios.all())
        if suma_puntajes_maximos > 0:
            nota_calculada = (evaluacion.nota_maxima * puntaje_total) / suma_puntajes_maximos
            # Usar la nota calculada automáticamente
            calificacion.nota = round(nota_calculada, 1)
        else:
            calificacion.nota = nueva_nota
        
        # Actualizar la calificación
        calificacion.retroalimentacion = retroalimentacion_limpia
        calificacion.calificado_por = request.user
        calificacion.save()
        
        # Actualizar puntaje total en ResultadoRubrica
        resultado_rubrica.puntaje_total = puntaje_total
        resultado_rubrica.nota_final = calificacion.nota
        resultado_rubrica.save()
        
        # Recalcular puntaje total usando el método del modelo
        resultado_rubrica.calcular_puntaje_total()
        
        messages.success(request, f'Calificación actualizada exitosamente para {estudiante.get_full_name()}.')
        return redirect('calificar_estudiante', curso_id=evaluacion.curso.id, evaluacion_id=evaluacion.id)
        
    except Exception as e:
        messages.error(request, f'Error al actualizar la calificación: {str(e)}')
        return redirect('plataforma_calificaciones', curso_id=evaluacion.curso.id)

@login_required
def limpiar_retroalimentaciones(request):
    """
    Vista temporal para limpiar caracteres de control en retroalimentaciones existentes
    SOLO EJECUTAR UNA VEZ
    """
    if not request.user.is_superuser:
        messages.error(request, 'Solo superusuarios pueden ejecutar esta acción.')
        return redirect('user_space')
    
    from .models import Calificacion
    
    # Obtener todas las calificaciones con retroalimentación
    calificaciones = Calificacion.objects.filter(retroalimentacion__isnull=False).exclude(retroalimentacion='')
    
    contador = 0
    for calificacion in calificaciones:
        if calificacion.retroalimentacion:
            # Limpiar caracteres de control
            retroalimentacion_original = calificacion.retroalimentacion
            retroalimentacion_limpia = re.sub(r'\\u000D\\u000A', '\n', calificacion.retroalimentacion)
            retroalimentacion_limpia = re.sub(r'\\u000D', '\n', retroalimentacion_limpia)
            retroalimentacion_limpia = re.sub(r'\\u000A', '\n', retroalimentacion_limpia)
            retroalimentacion_limpia = re.sub(r'\\n', '\n', retroalimentacion_limpia)
            retroalimentacion_limpia = re.sub(r'\\r', '\n', retroalimentacion_limpia)
            
            # Solo actualizar si hay cambios
            if retroalimentacion_original != retroalimentacion_limpia:
                calificacion.retroalimentacion = retroalimentacion_limpia
                calificacion.save()
                contador += 1
    
    messages.success(request, f'Se limpiaron {contador} retroalimentaciones con caracteres de control.')
    return redirect('user_space')

@login_required
def plataforma_entregas(request, curso_id):
    """
    Vista principal de entregas para la plataforma de aprendizaje
    """
    curso = get_object_or_404(Curso, id=curso_id)
    
    # Verificar que el usuario esté inscrito en el curso
    if not request.user.cursos.filter(id=curso_id).exists():
        messages.error(request, 'No tienes acceso a este curso.')
        return redirect('user_space')
    
    context = {
        'curso': curso,
        'user': request.user,
    }
    
    return render(request, 'pages/plataforma_entregas.html', context)

@login_required
def plataforma_entregas_ajax(request, curso_id):
    """Devuelve el HTML de la tabla de entregas y el formulario de subida para el usuario actual en el curso."""
    from django.template.loader import render_to_string
    from django.db.models import Prefetch, Q
    curso = get_object_or_404(Curso, id=curso_id)
    user = request.user
    
    # Manejar petición POST (subir entrega)
    if request.method == 'POST':
        logger.debug("DEBUG: Procesando petición POST para subir entrega")
        evaluacion_id = request.POST.get('evaluacion')
        logger.debug(f"DEBUG: Evaluación ID: {evaluacion_id}")
        evaluacion = get_object_or_404(Evaluacion, id=evaluacion_id)
        logger.debug(f"DEBUG: Evaluación encontrada: {evaluacion.nombre}")
        
        # Debug de archivos
        if request.FILES:
            for field_name, uploaded_file in request.FILES.items():
                logger.debug(f"DEBUG: Archivo recibido - Campo: {field_name}, Nombre: {uploaded_file.name}, Tamaño: {uploaded_file.size} bytes ({uploaded_file.size / (1024*1024):.2f} MB)")
        else:
            logger.debug("DEBUG: No se recibieron archivos")
        
        form = EntregaForm(request.POST, request.FILES, evaluacion=evaluacion, estudiante=user)
        
        if form.is_valid():
            logger.debug("DEBUG: Formulario válido, guardando entrega...")
            entrega = form.save(commit=False)
            entrega.evaluacion = evaluacion
            entrega.estudiante = user
            entrega.save()
            logger.debug("DEBUG: Entrega guardada exitosamente")
            messages.success(request, 'Entrega subida correctamente.')
        else:
            logger.debug("DEBUG: Formulario inválido")
            logger.debug(f"DEBUG: Errores del formulario: {form.errors}")
            # Para peticiones AJAX, incluir los errores en la respuesta
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                # Recrear el formulario con errores para que se muestren en el template
                form = EntregaForm(request.POST, request.FILES, evaluacion=evaluacion, estudiante=user)
            else:
                messages.error(request, 'Error al subir la entrega. Verifica los datos.')
    
    # Si es admin/staff, mostrar todas las entregas agrupadas por evaluación
    if user.is_staff or user.is_superuser:
        # Obtener todos los estudiantes inscritos en el curso (no staff/admin)
        estudiantes_curso = User.objects.filter(
            cursos=curso,
            is_staff=False,
            is_superuser=False
        ).order_by('first_name', 'last_name', 'username')
        
        evaluaciones_con_entregas = Evaluacion.objects.filter(
            curso=curso
        ).prefetch_related(
            Prefetch(
                'entregas',
                queryset=Entrega.objects.select_related('estudiante').order_by('-fecha_entrega')
            )
        ).order_by('-fecha_inicio')
        
        # Para cada evaluación, crear una lista de todos los estudiantes con su información
        for evaluacion in evaluaciones_con_entregas:
            # Crear una lista de estudiantes con su información de entrega y calificación
            estudiantes_con_info = []
            
            for estudiante in estudiantes_curso:
                # Buscar si el estudiante tiene entrega para esta evaluación
                entrega = evaluacion.entregas.filter(estudiante=estudiante).first()
                
                # Verificar si existe una calificación para este estudiante en esta evaluación
                tiene_calificacion = Calificacion.objects.filter(
                    evaluacion=evaluacion,
                    estudiante=estudiante
                ).exists()
                
                # Crear un objeto con toda la información del estudiante
                info_estudiante = {
                    'estudiante': estudiante,
                    'entrega': entrega,
                    'tiene_calificacion': tiene_calificacion,
                    'tiene_entrega': entrega is not None
                }
                
                estudiantes_con_info.append(info_estudiante)
            
            # Calcular estadísticas
            total_estudiantes = len(estudiantes_con_info)
            estudiantes_con_entrega = sum(1 for info in estudiantes_con_info if info['tiene_entrega'])
            estudiantes_sin_entregar = total_estudiantes - estudiantes_con_entrega
            estudiantes_calificados = sum(1 for info in estudiantes_con_info if info['tiene_calificacion'])
            
            # Agregar esta información a la evaluación
            evaluacion.estudiantes_con_info = estudiantes_con_info
            evaluacion.stats = {
                'total_estudiantes': total_estudiantes,
                'estudiantes_con_entrega': estudiantes_con_entrega,
                'estudiantes_sin_entregar': estudiantes_sin_entregar,
                'estudiantes_calificados': estudiantes_calificados
            }
        
        html = render_to_string('pages/plataforma_entregas_admin_content.html', {
            'evaluaciones_con_entregas': evaluaciones_con_entregas,
            'curso': curso,
            'user': user,
        }, request=request)
    else:
        # Solo entregas del usuario actual
        entregas = Entrega.objects.filter(evaluacion__curso=curso, estudiante=user).select_related('evaluacion').order_by('-fecha_entrega')
        
        # Evaluaciones activas para subir entrega
        from django.utils import timezone
        fecha_actual = timezone.now().date()
        
        # Obtener evaluaciones que están activas, dentro del rango de fechas y donde el usuario no ha entregado
        evaluaciones_activas = Evaluacion.objects.filter(
            curso=curso,
            activa=True
        ).filter(
            # Solo evaluaciones que estén realmente dentro del rango de fechas actual
            Q(
                fecha_inicio__lte=fecha_actual,
                fecha_fin__gte=fecha_actual
            )
        ).exclude(
            # Excluir evaluaciones donde el usuario ya tiene entregas
            entregas__estudiante=user
        )
        
        # Debug de evaluaciones disponibles
        evaluaciones_en_rango = Evaluacion.objects.filter(
            curso=curso,
            activa=True,
            fecha_inicio__lte=fecha_actual,
            fecha_fin__gte=fecha_actual
        )
        logger.debug(f"DEBUG: Evaluaciones en rango de fechas: {evaluaciones_en_rango.count()}")
        
        evaluaciones_sin_entregas = evaluaciones_en_rango.exclude(entregas__estudiante=user)
        logger.debug(f"DEBUG: Evaluaciones en rango sin entregas del usuario: {evaluaciones_sin_entregas.count()}")
        
        # Debug: imprimir información detallada
        logger.debug(f"DEBUG: Usuario: {user.username}")
        logger.debug(f"DEBUG: Fecha actual: {fecha_actual}")
        logger.debug(f"DEBUG: Total evaluaciones en el curso: {Evaluacion.objects.filter(curso=curso).count()}")
        logger.debug(f"DEBUG: Evaluaciones activas: {Evaluacion.objects.filter(curso=curso, activa=True).count()}")
        logger.debug(f"DEBUG: Evaluaciones con fechas válidas: {Evaluacion.objects.filter(curso=curso, activa=True).filter(Q(fecha_inicio__isnull=True, fecha_fin__isnull=True) | Q(fecha_inicio__lte=fecha_actual, fecha_fin__gte=fecha_actual)).count()}")
        logger.debug(f"DEBUG: Entregas del usuario: {Entrega.objects.filter(evaluacion__curso=curso, estudiante=user).count()}")
        logger.debug(f"DEBUG: Evaluaciones finales disponibles: {evaluaciones_activas.count()}")
        
        # Mostrar detalles de cada evaluación
        for eval in Evaluacion.objects.filter(curso=curso):
            logger.debug(f"DEBUG: Evaluación '{eval.nombre}' - Activa: {eval.activa}, Inicio: {eval.fecha_inicio}, Fin: {eval.fecha_fin}")
            tiene_entrega = eval.entregas.filter(estudiante=user).exists()
            logger.debug(f"DEBUG: - Usuario tiene entrega: {tiene_entrega}")
            if eval.fecha_inicio and eval.fecha_fin:
                en_rango = eval.fecha_inicio <= fecha_actual <= eval.fecha_fin
                logger.debug(f"DEBUG: - En rango de fechas: {en_rango}")
        
        form = None
        evaluacion_para_entregar = None
        if evaluaciones_activas.exists():
            evaluacion_para_entregar = evaluaciones_activas.first()
            logger.debug(f"DEBUG: Evaluación seleccionada para entregar: {evaluacion_para_entregar.nombre}")
            # Crear el formulario siempre que haya una evaluación disponible
            form = EntregaForm(evaluacion=evaluacion_para_entregar, estudiante=user)
            logger.debug(f"DEBUG: Formulario creado: {form is not None}")
        else:
            logger.debug("DEBUG: No hay evaluaciones activas para entregar")
        logger.debug(f"DEBUG: Variables para template - entregas: {entregas.count()}, form: {form is not None}, evaluacion_para_entregar: {evaluacion_para_entregar.nombre if evaluacion_para_entregar else None}")
        
        html = render_to_string('pages/plataforma_entregas_content.html', {
            'entregas': entregas,
            'form': form,
            'evaluacion_para_entregar': evaluacion_para_entregar,
            'curso': curso,
        }, request=request)
    
    return JsonResponse({'html': html})

@login_required
def reemplazar_archivo_entrega(request):
    """
    Vista para reemplazar el archivo de una entrega existente.
    Solo accesible para staff/superuser.
    """
    if not request.user.is_staff and not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'No tienes permisos para realizar esta acción'})
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    
    try:
        entrega_id = request.POST.get('entrega_id')
        nuevo_archivo = request.FILES.get('archivo')
        comentario = request.POST.get('comentario', '')
        
        if not entrega_id or not nuevo_archivo:
            return JsonResponse({'success': False, 'error': 'Datos incompletos'})
        
        # Obtener la entrega
        entrega = get_object_or_404(Entrega, id=entrega_id)
        
        # Eliminar el archivo anterior si existe
        if entrega.archivo:
            if os.path.exists(entrega.archivo.path):
                os.remove(entrega.archivo.path)
        
        # Actualizar la entrega con el nuevo archivo
        entrega.archivo = nuevo_archivo
        if comentario:
            entrega.comentario = comentario
        entrega.save()
        
        return JsonResponse({'success': True, 'message': 'Archivo reemplazado correctamente'})
        
    except Exception as e:
        logger.error(f"Error al reemplazar archivo de entrega: {str(e)}")
        return JsonResponse({'success': False, 'error': f'Error al reemplazar el archivo: {str(e)}'})

def enviar_correo_bienvenida_usuario_existente(request, user, curso_nombre):
    """
    Función para enviar correo de bienvenida a un usuario existente que se inscribe a un nuevo curso
    """
    from django.template.loader import render_to_string
    from django.utils.html import strip_tags
    from .models import Curso
    
    # Obtener el curso completo para acceder a sus videollamadas
    try:
        curso = Curso.objects.get(nombre=curso_nombre)
        videollamadas = curso.videollamadas.filter(activa=True).order_by('dia_semana', 'hora_inicio')
        
        # Formatear información de horarios
        horarios_info = []
        for videollamada in videollamadas:
            horarios_info.append({
                'dia': videollamada.get_dia_semana_display(),
                'hora_inicio': videollamada.hora_inicio.strftime('%H:%M'),
                'hora_fin': videollamada.hora_fin.strftime('%H:%M'),
                'descripcion': videollamada.descripcion
            })
    except Curso.DoesNotExist:
        curso = None
        horarios_info = []
    
    # Generar URLs
    login_url = request.build_absolute_uri('/accounts/login/')
    
    # Renderizar el template HTML
    html_message = render_to_string('emails/nueva_inscripcion.html', {
        'nombre_usuario': user.get_full_name() or user.username,
        'username': user.username,
        'email': user.email,
        'curso_nombre': curso_nombre,
        'curso': curso,
        'horarios_info': horarios_info,
        'login_url': login_url,
    })
    
    # Crear versión de texto plano
    plain_message = strip_tags(html_message)
    
    subject = f'¡Bienvenido a tu nuevo curso! - {curso_nombre}'
    
    try:
        send_mail(
            subject=subject,
            message=plain_message,
            from_email=settings.DEFAULT_FROM_EMAIL if hasattr(settings, 'DEFAULT_FROM_EMAIL') else 'contacto@gryphos.cl',
            recipient_list=[user.email],
            html_message=html_message,
            fail_silently=False,
        )
        logger.info(f"Correo de bienvenida a usuario existente enviado exitosamente a {user.email}")
        return True
    except Exception as e:
        logger.error(f"Error enviando correo de bienvenida a usuario existente: {e}")
        return False

# ============================================================================
# VISTAS DEL SISTEMA DE TICKETS DE SOPORTE
# ============================================================================

@login_required
def plataforma_soporte(request, curso_id):
    """
    Vista principal de soporte para la plataforma de aprendizaje
    """
    curso = get_object_or_404(Curso, id=curso_id)
    
    # Verificar que el usuario esté inscrito en el curso
    if not request.user.cursos.filter(id=curso_id).exists():
        messages.error(request, 'No tienes acceso a este curso.')
        return redirect('user_space')
    
    context = {
        'curso': curso,
        'user': request.user,
    }
    
    return render(request, 'pages/plataforma_soporte.html', context)

@login_required
def plataforma_soporte_ajax(request, curso_id):
    """
    Vista AJAX para cargar el contenido del sistema de soporte dinámicamente
    """
    try:
        curso = Curso.objects.get(id=curso_id, activo=True)
        
        # Verificar que el usuario esté inscrito en el curso
        if curso not in request.user.cursos.all():
            return JsonResponse({'error': 'No tienes acceso a este curso'}, status=403)
        
        # Verificar si es una acción específica
        action = request.GET.get('action')
        
        logger.debug(f"=== DEBUG PLATAFORMA SOPORTE AJAX ===")
        logger.debug(f"Action recibida: '{action}'")
        logger.debug(f"URL completa: {request.get_full_path()}")
        logger.debug(f"Parámetros GET: {dict(request.GET)}")
        
        if action == 'crear_ticket':
            # Verificar si el usuario es admin/staff
            if request.user.is_staff or request.user.is_superuser:
                # Para usuarios admin/staff, mostrar el panel de administración
                tickets = TicketSoporte.objects.filter(curso=curso).order_by('-fecha_creacion')
                
                context = {
                    'curso': curso,
                    'tickets': tickets,
                    'estados': TicketSoporte.ESTADO_CHOICES,
                    'current_estado': None,
                    'is_admin': True,
                }
                
                html = render_to_string('pages/plataforma_soporte_admin.html', context, request=request)
                return JsonResponse({'html': html})
            else:
                # Para usuarios normales, mostrar el formulario de crear ticket
                from .forms import TicketSoporteForm
                
                logger.debug("=== DEBUG FORMULARIO TICKET ===")
                logger.debug(f"Usuario: {request.user.username}")
                logger.debug(f"Curso: {curso.nombre} (ID: {curso.id})")
                
                # Verificar clasificaciones en la base de datos
                from .models import ClasificacionTicket, SubclasificacionTicket
                clasificaciones_db = ClasificacionTicket.objects.filter(activa=True)
                logger.debug(f"Clasificaciones activas en DB: {clasificaciones_db.count()}")
                for clas in clasificaciones_db:
                    logger.debug(f"  - {clas.nombre} (ID: {clas.id})")
                    subclas = clas.subclasificaciones.filter(activa=True)
                    logger.debug(f"    Subclasificaciones: {subclas.count()}")
                    for sub in subclas:
                        logger.debug(f"      - {sub.nombre}")
                
                # Instanciar formulario
                form = TicketSoporteForm()
                
                # Verificar opciones del formulario
                logger.debug(f"Opciones de clasificación en form: {len(form.fields['clasificacion'].choices)}")
                for choice in form.fields['clasificacion'].choices:
                    logger.debug(f"  - {choice[1]} (valor: {choice[0]})")
                
                logger.debug(f"Opciones de subclasificación en form widget: {len(form.fields['subclasificacion'].widget.choices)}")
                for choice in form.fields['subclasificacion'].widget.choices:
                    logger.debug(f"  - {choice[1]} (valor: {choice[0]})")
                
                context = {
                    'curso': curso,
                    'form': form,
                }
                
                # Renderizar template
                html = render_to_string('pages/plataforma_soporte_create_ticket.html', context, request=request)
                
                # Verificar si el HTML contiene las opciones
                if 'clasificacion-select' in html:
                    logger.debug("✓ Campo clasificacion-select encontrado en HTML")
                else:
                    logger.debug("✗ Campo clasificacion-select NO encontrado en HTML")
                
                if 'subclasificacion-select' in html:
                    logger.debug("✓ Campo subclasificacion-select encontrado en HTML")
                else:
                    logger.debug("✗ Campo subclasificacion-select NO encontrado en HTML")
                
                # Buscar opciones en el HTML
                import re
                option_pattern = r'<option[^>]*value="([^"]*)"[^>]*>([^<]*)</option>'
                options_in_html = re.findall(option_pattern, html)
                logger.debug(f"Opciones encontradas en HTML: {len(options_in_html)}")
                for value, text in options_in_html[:10]:  # Mostrar las primeras 10
                    logger.debug(f"  - {text} (valor: {value})")
                
                # Verificar específicamente las opciones de clasificación
                clasificacion_options = re.findall(r'<option[^>]*value="([^"]*)"[^>]*>([^<]*)</option>', html)
                clasificacion_options = [opt for opt in clasificacion_options if opt[0] != '' and 'Selecciona' not in opt[1]]
                logger.debug(f"Opciones de clasificación válidas en HTML: {len(clasificacion_options)}")
                for value, text in clasificacion_options:
                    logger.debug(f"  - {text} (valor: {value})")
                
                logger.debug("=== FIN DEBUG FORMULARIO TICKET ===")
                
                return JsonResponse({'html': html})
        
        elif action == 'ver_ticket':
            # Cargar ticket específico
            ticket_id = request.GET.get('ticket_id')
            logger.debug(f"=== DEBUG VER TICKET ===")
            logger.debug(f"Ticket ID: {ticket_id}")
            logger.debug(f"Usuario: {request.user.username}")
            logger.debug(f"Es staff: {request.user.is_staff}")
            logger.debug(f"Es superuser: {request.user.is_superuser}")
            
            if ticket_id:
                try:
                    ticket = TicketSoporte.objects.get(id=ticket_id, curso=curso)
                    logger.debug(f"Ticket encontrado: {ticket.titulo}")
                    
                    # Verificar permisos
                    if not (request.user == ticket.usuario or request.user.is_staff or request.user.is_superuser):
                        logger.debug("Usuario no tiene permisos para ver este ticket")
                        return JsonResponse({'error': 'No tienes permisos para ver este ticket'}, status=403)
                    
                    # Filtrar comentarios según permisos del usuario
                    if request.user.is_staff or request.user.is_superuser:
                        comentarios = ticket.comentarios.all()
                        logger.debug(f"Comentarios para admin: {comentarios.count()}")
                    else:
                        comentarios = ticket.comentarios.filter(es_interno=False)
                        logger.debug(f"Comentarios para usuario: {comentarios.count()}")
                    
                    comentario_form = ComentarioTicketForm()
                    admin_form = TicketSoporteAdminForm(instance=ticket) if request.user.is_staff else None
                    
                    # Obtener usuarios staff para reasignación
                    usuarios_staff = CustomUser.objects.filter(
                        models.Q(is_staff=True) | models.Q(is_superuser=True)
                    ).exclude(id=request.user.id).order_by('first_name', 'last_name', 'username')
                    
                    context = {
                        'curso': curso,
                        'ticket': ticket,
                        'comentarios': comentarios,
                        'comentario_form': comentario_form,
                        'admin_form': admin_form,
                        'usuarios_staff': usuarios_staff,
                    }
                    
                    logger.debug(f"Contexto preparado con {len(comentarios)} comentarios")
                    
                    # Usar template diferente según el tipo de usuario
                    if request.user.is_staff or request.user.is_superuser:
                        logger.debug("Usando template de admin")
                        html = render_to_string('pages/plataforma_soporte_ticket_detail.html', context, request=request)
                    else:
                        logger.debug("Usando template de usuario")
                        html = render_to_string('pages/plataforma_soporte_ticket_detail_user.html', context, request=request)
                    
                    logger.debug(f"HTML generado: {len(html)} caracteres")
                    return JsonResponse({'html': html})
                except TicketSoporte.DoesNotExist:
                    logger.debug(f"Ticket {ticket_id} no existe")
                    return JsonResponse({'error': 'El ticket no existe'}, status=404)
                except Exception as e:
                    logger.debug(f"Error al cargar ticket {ticket_id}: {str(e)}")
                    import traceback
                    traceback.logger.debug_exc()
                    return JsonResponse({'error': f'Error interno del servidor: {str(e)}'}, status=500)
            else:
                logger.debug("No se proporcionó ticket_id")
                return JsonResponse({'error': 'ID de ticket no proporcionado'}, status=400)
        
        # Verificar si el usuario es admin/staff
        if request.user.is_staff or request.user.is_superuser:
            # Para usuarios admin/staff, mostrar el panel de administración
            tickets = TicketSoporte.objects.filter(curso=curso).order_by('-fecha_creacion')
            
            # Filtros
            estado_filter = request.GET.get('estado')
            if estado_filter:
                tickets = tickets.filter(estado=estado_filter)
            
            # Calcular estadísticas
            total_tickets = tickets.count()
            tickets_abiertos = tickets.filter(estado='abierto').count()
            tickets_en_proceso = tickets.filter(estado='en_proceso').count()
            tickets_resueltos = tickets.filter(estado='resuelto').count()
            tickets_cerrados = tickets.filter(estado='cerrado').count()
            
            # Obtener usuarios staff para reasignación (incluyendo al usuario actual para asignaciones iniciales)
            usuarios_staff = CustomUser.objects.filter(
                models.Q(is_staff=True) | models.Q(is_superuser=True)
            ).order_by('first_name', 'last_name', 'username')
            
            context = {
                'curso': curso,
                'tickets': tickets,
                'estados': TicketSoporte.ESTADO_CHOICES,
                'current_estado': estado_filter,
                'is_admin': True,
                'usuarios_staff': usuarios_staff,
                'stats': {
                    'total': total_tickets,
                    'abiertos': tickets_abiertos,
                    'en_proceso': tickets_en_proceso,
                    'resueltos': tickets_resueltos,
                    'cerrados': tickets_cerrados,
                }
            }
            
            html = render_to_string('pages/plataforma_soporte_admin.html', context, request=request)
            return JsonResponse({'html': html})
        else:
            # Para usuarios normales, mostrar la lista de sus tickets
            tickets = TicketSoporte.objects.filter(curso=curso, usuario=request.user).order_by('-fecha_creacion')
            
            # Filtros
            estado_filter = request.GET.get('estado')
            logger.debug(f"=== DEBUG FILTRO SOPORTE ===")
            logger.debug(f"Estado filtro recibido: '{estado_filter}'")
            logger.debug(f"Tickets antes del filtro: {tickets.count()}")
            
            # Por defecto, mostrar tickets abiertos si no se especifica un estado
            if estado_filter:
                tickets = tickets.filter(estado=estado_filter)
                logger.debug(f"Tickets después del filtro '{estado_filter}': {tickets.count()}")
            else:
                # Filtro por defecto: mostrar solo tickets abiertos
                tickets = tickets.filter(estado='abierto')
                estado_filter = 'abierto'  # Establecer el estado por defecto
                logger.debug(f"Tickets después del filtro por defecto 'abierto': {tickets.count()}")
            
            context = {
                'curso': curso,
                'tickets': tickets,
                'estados': TicketSoporte.ESTADO_CHOICES,
                'current_estado': estado_filter,
            }
            
            logger.debug(f"Estados disponibles: {TicketSoporte.ESTADO_CHOICES}")
            logger.debug("=== FIN DEBUG FILTRO SOPORTE ===")
            
            # Renderizar solo el contenido del soporte
            html = render_to_string('pages/plataforma_soporte_content.html', context, request=request)
            return JsonResponse({'html': html})
        
    except Curso.DoesNotExist:
        return JsonResponse({'error': 'El curso no existe'}, status=404)


@login_required
def crear_ticket_soporte(request, curso_id):
    """
    Vista para crear un nuevo ticket de soporte
    """
    if request.method == 'POST':
        logger.debug("=== DEBUG CREAR TICKET ===")
        logger.debug(f"Datos POST recibidos: {request.POST}")
        
        form = TicketSoporteForm(request.POST)
        logger.debug(f"Formulario válido: {form.is_valid()}")
        
        if form.is_valid():
            ticket = form.save(commit=False)
            ticket.usuario = request.user
            ticket.curso_id = curso_id
            ticket.save()
            
            logger.debug("✓ Ticket creado exitosamente")
            logger.debug("=== FIN DEBUG CREAR TICKET ===")
            
            messages.success(request, 'Ticket creado exitosamente. Recibirás una respuesta pronto.')
            return JsonResponse({'success': True, 'message': 'Ticket creado exitosamente'})
        else:
            logger.debug("✗ Formulario inválido")
            logger.debug(f"Errores del formulario: {form.errors}")
            logger.debug("=== FIN DEBUG CREAR TICKET ===")
            return JsonResponse({'success': False, 'errors': form.errors})
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def agregar_comentario_ticket(request, ticket_id):
    """
    Vista para agregar comentarios a un ticket
    """
    logger.debug(f"=== DEBUG AGREGAR COMENTARIO ===")
    logger.debug(f"Ticket ID: {ticket_id}")
    logger.debug(f"Usuario: {request.user.username}")
    logger.debug(f"Método: {request.method}")
    logger.debug(f"Datos POST: {request.POST}")
    
    try:
        ticket = TicketSoporte.objects.get(id=ticket_id)
        logger.debug(f"Ticket encontrado: {ticket.titulo}")
        
        # Verificar permisos
        puede_comentar = ticket.puede_comentar(request.user)
        logger.debug(f"Puede comentar: {puede_comentar}")
        
        if not puede_comentar:
            logger.debug("Usuario no tiene permisos para comentar")
            return JsonResponse({'error': 'No tienes permisos para comentar en este ticket'}, status=403)
        
        if request.method == 'POST':
            form = ComentarioTicketForm(request.POST)
            logger.debug(f"Formulario válido: {form.is_valid()}")
            
            if form.is_valid():
                comentario = form.save(commit=False)
                comentario.ticket = ticket
                comentario.autor = request.user
                comentario.save()
                
                logger.debug(f"Comentario guardado: {comentario.id}")
                
                # Actualizar fecha de actualización del ticket
                ticket.fecha_actualizacion = timezone.now()
                ticket.save()
                
                logger.debug("=== FIN DEBUG AGREGAR COMENTARIO ===")
                return JsonResponse({'success': True, 'message': 'Comentario agregado exitosamente'})
            else:
                logger.debug(f"Errores del formulario: {form.errors}")
                logger.debug("=== FIN DEBUG AGREGAR COMENTARIO ===")
                return JsonResponse({'success': False, 'errors': form.errors})
    
    except TicketSoporte.DoesNotExist:
        return JsonResponse({'error': 'El ticket no existe'}, status=404)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def actualizar_ticket_admin(request, ticket_id):
    """
    Vista para que los administradores actualicen el estado de un ticket
    """
    try:
        ticket = TicketSoporte.objects.get(id=ticket_id)
        
        # Verificar que sea admin/staff
        if not (request.user.is_staff or request.user.is_superuser):
            return JsonResponse({'error': 'No tienes permisos para actualizar tickets'}, status=403)
        
        if request.method == 'POST':
            form = TicketSoporteAdminForm(request.POST, instance=ticket)
            if form.is_valid():
                form.save()
                
                # Si el estado cambió a resuelto, actualizar fecha de resolución
                if ticket.estado == 'resuelto' and not ticket.fecha_resolucion:
                    ticket.fecha_resolucion = timezone.now()
                    ticket.save()
                
                return JsonResponse({'success': True, 'message': 'Ticket actualizado exitosamente'})
            else:
                return JsonResponse({'success': False, 'errors': form.errors})
    
    except TicketSoporte.DoesNotExist:
        return JsonResponse({'error': 'El ticket no existe'}, status=404)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def obtener_subclasificaciones(request):
    """
    Vista AJAX para obtener subclasificaciones basadas en la clasificación seleccionada
    """
    clasificacion = request.GET.get('clasificacion')
    
    logger.debug(f"=== DEBUG OBTENER SUBCLASIFICACIONES ===")
    logger.debug(f"Clasificación solicitada: {clasificacion}")
    
    if clasificacion:
        try:
            clasificacion_obj = ClasificacionTicket.objects.get(nombre=clasificacion, activa=True)
            logger.debug(f"Clasificación encontrada: {clasificacion_obj.nombre} (ID: {clasificacion_obj.id})")
            
            subclasificaciones = clasificacion_obj.subclasificaciones.filter(activa=True)
            logger.debug(f"Subclasificaciones encontradas: {subclasificaciones.count()}")
            for sub in subclasificaciones:
                logger.debug(f"  - {sub.nombre} (ID: {sub.id})")
            
            choices = [('', 'Selecciona una subclasificación')] + [
                (sub.nombre, sub.nombre) for sub in subclasificaciones
            ]
            
            logger.debug(f"Choices finales: {len(choices)}")
            for choice in choices:
                logger.debug(f"  - {choice[1]} (valor: {choice[0]})")
            
            logger.debug("=== FIN DEBUG OBTENER SUBCLASIFICACIONES ===")
            return JsonResponse({'choices': choices})
        except ClasificacionTicket.DoesNotExist:
            logger.debug(f"✗ Clasificación '{clasificacion}' no encontrada")
            logger.debug("=== FIN DEBUG OBTENER SUBCLASIFICACIONES ===")
            return JsonResponse({'choices': [('', 'No hay subclasificaciones disponibles')]})
    
    logger.debug("✗ No se proporcionó clasificación")
    logger.debug("=== FIN DEBUG OBTENER SUBCLASIFICACIONES ===")
    return JsonResponse({'choices': [('', 'Primero selecciona una clasificación')]})


def crear_clasificaciones_iniciales():
    """
    Función para crear las clasificaciones iniciales de tickets
    """
    clasificaciones = [
        {
            'nombre': 'Problemas Técnicos',
            'descripcion': 'Problemas relacionados con la plataforma, acceso, o funcionalidades técnicas',
            'subclasificaciones': [
                'Problemas de acceso',
                'Errores en la plataforma',
                'Problemas con archivos',
                'Problemas con videollamadas'
            ]
        },
        {
            'nombre': 'Contenido del Curso',
            'descripcion': 'Consultas sobre el contenido, materiales o recursos del curso',
            'subclasificaciones': [
                'Consultas sobre contenido',
                'Materiales faltantes',
                'Recursos adicionales',
                'Clarificación de conceptos'
            ]
        },
        {
            'nombre': 'Evaluaciones',
            'descripcion': 'Consultas sobre evaluaciones, calificaciones o entregas',
            'subclasificaciones': [
                'Consultas sobre evaluaciones',
                'Problemas con entregas',
                'Revisión de calificaciones',
                'Solicitud de retroalimentación'
            ]
        },
        {
            'nombre': 'General',
            'descripcion': 'Consultas generales sobre el curso o la plataforma',
            'subclasificaciones': [
                'Información general',
                'Sugerencias',
                'Otros'
            ]
        }
    ]
    
    for clasificacion_data in clasificaciones:
        clasificacion, created = ClasificacionTicket.objects.get_or_create(
            nombre=clasificacion_data['nombre'],
            defaults={
                'descripcion': clasificacion_data['descripcion'],
                'activa': True
            }
        )
        
        if created:
            logger.debug(f"Clasificación creada: {clasificacion.nombre}")
        
        for subclasificacion_nombre in clasificacion_data['subclasificaciones']:
            subclasificacion, created = SubclasificacionTicket.objects.get_or_create(
                clasificacion=clasificacion,
                nombre=subclasificacion_nombre,
                defaults={
                    'descripcion': f'Subclasificación de {clasificacion.nombre}',
                    'activa': True
                }
            )
            
            if created:
                logger.debug(f"  Subclasificación creada: {subclasificacion.nombre}")

@login_required
def crear_rubrica(request, curso_id, evaluacion_id):
    """
    Vista para crear una rúbrica para una evaluación específica
    """
    curso = get_object_or_404(Curso, id=curso_id)
    evaluacion = get_object_or_404(Evaluacion, id=evaluacion_id, curso=curso)
    
    # Verificar que el usuario tenga permisos (staff o creador de la evaluación)
    if not request.user.is_staff and evaluacion.creado_por != request.user:
        messages.error(request, 'No tienes permisos para crear rúbricas para esta evaluación.')
        return redirect('plataforma_calificaciones', curso_id=curso_id)
    
    # Verificar si ya existe una rúbrica para esta evaluación
    if hasattr(evaluacion, 'rubrica'):
        messages.warning(request, 'Esta evaluación ya tiene una rúbrica asociada.')
        return redirect('editar_rubrica', curso_id=curso_id, evaluacion_id=evaluacion_id)
    
    if request.method == 'POST':
        # Procesar el formulario de creación de rúbrica
        nombre = request.POST.get('nombre')
        descripcion = request.POST.get('descripcion')
        objetivo = request.POST.get('objetivo')
        aprendizaje_esperado = request.POST.get('aprendizaje_esperado')
        
        if nombre and objetivo and aprendizaje_esperado:
            # Crear la rúbrica
            rubrica = Rubrica.objects.create(
                evaluacion=evaluacion,
                nombre=nombre,
                descripcion=descripcion,
                objetivo=objetivo,
                aprendizaje_esperado=aprendizaje_esperado,
                creado_por=request.user
            )
            
            messages.success(request, 'Rúbrica creada exitosamente. Ahora puedes agregar criterios.')
            return redirect('editar_rubrica', curso_id=curso_id, evaluacion_id=evaluacion_id)
        else:
            messages.error(request, 'Por favor completa todos los campos obligatorios.')
    
    # Contexto necesario para el template que hereda de plataforma_aprendizaje.html
    context = {
        'curso': curso,
        'evaluacion': evaluacion,
        'user': request.user,
        'current_category': None,  # Para el foro
    }
    
    return render(request, 'pages/crear_rubrica.html', context)

@login_required
def editar_rubrica(request, curso_id, evaluacion_id):
    """
    Vista para editar una rúbrica existente
    """
    curso = get_object_or_404(Curso, id=curso_id)
    evaluacion = get_object_or_404(Evaluacion, id=evaluacion_id, curso=curso)
    
    # Verificar que el usuario tenga permisos
    if not request.user.is_staff and evaluacion.creado_por != request.user:
        messages.error(request, 'No tienes permisos para editar rúbricas para esta evaluación.')
        return redirect('plataforma_calificaciones', curso_id=curso_id)
    
    # Obtener la rúbrica
    try:
        rubrica = evaluacion.rubrica
    except Rubrica.DoesNotExist:
        messages.error(request, 'No se encontró una rúbrica para esta evaluación.')
        return redirect('plataforma_calificaciones', curso_id=curso_id)
    
    # Verificar si la evaluación ya comenzó (no se puede editar después de la fecha de inicio)
    from datetime import date
    hoy = date.today()
    puede_editar = True
    mensaje_restriccion = None
    
    if evaluacion.fecha_inicio and hoy >= evaluacion.fecha_inicio:
        puede_editar = False
        mensaje_restriccion = f'La evaluación comenzó el {evaluacion.fecha_inicio.strftime("%d/%m/%Y")}. La rúbrica no se puede editar después de esta fecha.'
    
    if request.method == 'POST' and puede_editar:
        # Procesar actualización de la rúbrica
        nombre = request.POST.get('nombre')
        descripcion = request.POST.get('descripcion')
        objetivo = request.POST.get('objetivo')
        aprendizaje_esperado = request.POST.get('aprendizaje_esperado')
        
        if nombre and objetivo and aprendizaje_esperado:
            rubrica.nombre = nombre
            rubrica.descripcion = descripcion
            rubrica.objetivo = objetivo
            rubrica.aprendizaje_esperado = aprendizaje_esperado
            rubrica.save()
            
            # Verificar si es una petición AJAX
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Rúbrica actualizada exitosamente.'
                })
            else:
                messages.success(request, 'Rúbrica actualizada exitosamente.')
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'error': 'Por favor completa todos los campos obligatorios.'
                }, status=400)
            else:
                messages.error(request, 'Por favor completa todos los campos obligatorios.')
    elif request.method == 'POST' and not puede_editar:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': mensaje_restriccion
            }, status=403)
        else:
            messages.error(request, mensaje_restriccion)
    
    # Contexto necesario para el template que hereda de plataforma_aprendizaje.html
    context = {
        'curso': curso,
        'evaluacion': evaluacion,
        'rubrica': rubrica,
        'user': request.user,
        'current_category': None,  # Para el foro
        'puede_editar': puede_editar,
        'mensaje_restriccion': mensaje_restriccion,
    }
    
    return render(request, 'pages/editar_rubrica.html', context)

@login_required
def agregar_criterio_rubrica(request, curso_id, evaluacion_id):
    """
    Vista AJAX para agregar un criterio a una rúbrica
    """
    logger.info(f"Agregando criterio - Curso: {curso_id}, Evaluación: {evaluacion_id}")
    
    if request.method != 'POST':
        logger.error("Método no permitido")
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        curso = get_object_or_404(Curso, id=curso_id)
        evaluacion = get_object_or_404(Evaluacion, id=evaluacion_id, curso=curso)
        
        # Verificar permisos
        if not request.user.is_staff and evaluacion.creado_por != request.user:
            logger.error(f"Usuario {request.user.id} no tiene permisos para evaluación {evaluacion_id}")
            return JsonResponse({'error': 'No tienes permisos para realizar esta acción'}, status=403)
        
        # Verificar si la evaluación ya comenzó (no se puede editar después de la fecha de inicio)
        from datetime import date
        hoy = date.today()
        if evaluacion.fecha_inicio and hoy >= evaluacion.fecha_inicio:
            logger.error(f"Evaluación {evaluacion_id} ya comenzó")
            return JsonResponse({
                'error': f'La evaluación comenzó el {evaluacion.fecha_inicio.strftime("%d/%m/%Y")}. No se pueden agregar criterios después de esta fecha.'
            }, status=403)
        
        try:
            rubrica = evaluacion.rubrica
        except Rubrica.DoesNotExist:
            logger.error(f"No se encontró rúbrica para evaluación {evaluacion_id}")
            return JsonResponse({'error': 'No se encontró una rúbrica para esta evaluación'}, status=404)
        
        # Obtener datos del formulario
        nombre = request.POST.get('nombre')
        objetivo = request.POST.get('objetivo')
        puntaje = request.POST.get('puntaje')
        esperables_data = request.POST.getlist('esperables[]')
        
        logger.info(f"Datos recibidos - Nombre: {nombre}, Objetivo: {objetivo}, Puntaje: {puntaje}, Esperables: {len(esperables_data)}")
        
        if not nombre or not objetivo or not puntaje:
            logger.error("Campos obligatorios faltantes")
            return JsonResponse({'error': 'Nombre, objetivo y puntaje son campos obligatorios'}, status=400)
        
        try:
            puntaje_decimal = float(puntaje)
            if puntaje_decimal < 0:
                logger.error(f"Puntaje negativo: {puntaje_decimal}")
                return JsonResponse({'error': 'El puntaje debe ser un número positivo'}, status=400)
        except ValueError:
            logger.error(f"Puntaje inválido: {puntaje}")
            return JsonResponse({'error': 'El puntaje debe ser un número válido'}, status=400)
        
        # Crear el criterio
        criterio = CriterioRubrica.objects.create(
            rubrica=rubrica,
            nombre=nombre,
            objetivo=objetivo,
            puntaje=puntaje_decimal,
            orden=rubrica.criterios.count() + 1
        )
        
        logger.info(f"Criterio creado con ID: {criterio.id}")
        
        # Crear los esperables
        esperables_creados = 0
        for i, esperable_data in enumerate(esperables_data):
            if esperable_data.strip():  # Solo crear si no está vacío
                # Parsear el esperable que viene en formato JSON
                try:
                    esperable_obj = json.loads(esperable_data)
                    Esperable.objects.create(
                        criterio=criterio,
                        nivel=esperable_obj.get('nivel', f"Nivel {i+1}"),
                        descripcion=esperable_obj.get('descripcion', ''),
                        puntaje=esperable_obj.get('puntaje', 0),
                        orden=i+1
                    )
                    esperables_creados += 1
                except (json.JSONDecodeError, TypeError) as e:
                    logger.warning(f"Error parsing esperable {i}: {e}")
                    # Fallback para formato antiguo
                    Esperable.objects.create(
                        criterio=criterio,
                        nivel=f"Nivel {i+1}",
                        descripcion=esperable_data.strip(),
                        puntaje=0,
                        orden=i+1
                    )
                    esperables_creados += 1
        
        logger.info(f"Esperables creados: {esperables_creados}")
        
        # Recalcular el puntaje total del criterio basado en los esperables
        puntaje_total = criterio.esperables.aggregate(total=Max('puntaje'))['total'] or 0
        criterio.puntaje = puntaje_total
        criterio.save()
        
        logger.info(f"Criterio {criterio.id} guardado exitosamente con puntaje total: {puntaje_total}")
        
        response_data = {
            'success': True,
            'criterio_id': criterio.id,
            'criterio_nombre': str(criterio.nombre),
            'criterio_objetivo': str(criterio.objetivo),
            'criterio_puntaje': float(criterio.puntaje),
            'esperables_count': criterio.esperables.count()
        }
        
        logger.info(f"Respuesta exitosa: {response_data}")
        return JsonResponse(response_data, safe=True)
        
    except Exception as e:
        logger.error(f"Error inesperado en agregar_criterio_rubrica: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        return JsonResponse({'error': 'Error interno del servidor. Por favor, intenta nuevamente.'}, status=500)

@login_required
def obtener_criterio_rubrica(request, curso_id, evaluacion_id, criterio_id):
    """
    Vista AJAX para obtener los datos de un criterio específico
    """
    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    curso = get_object_or_404(Curso, id=curso_id)
    evaluacion = get_object_or_404(Evaluacion, id=evaluacion_id, curso=curso)
    criterio = get_object_or_404(CriterioRubrica, id=criterio_id, rubrica__evaluacion=evaluacion)
    
    # Verificar permisos
    if not request.user.is_staff and evaluacion.creado_por != request.user:
        return JsonResponse({'error': 'No tienes permisos para realizar esta acción'}, status=403)
    
    # Obtener los esperables del criterio
    esperables = list(criterio.esperables.values('nivel', 'descripcion', 'puntaje'))
    
    return JsonResponse({
        'success': True,
        'criterio': {
            'id': criterio.id,
            'nombre': str(criterio.nombre),
            'objetivo': str(criterio.objetivo),
            'puntaje': float(criterio.puntaje),
            'esperables': [
                {
                    'nivel': str(e['nivel']),
                    'descripcion': str(e['descripcion']),
                    'puntaje': float(e['puntaje'])
                } for e in esperables
            ]
        }
    }, safe=True)

@login_required
def editar_criterio_rubrica(request, curso_id, evaluacion_id, criterio_id):
    """
    Vista AJAX para editar un criterio de una rúbrica
    """
    logger.info(f"Editando criterio - Curso: {curso_id}, Evaluación: {evaluacion_id}, Criterio: {criterio_id}")
    
    if request.method != 'POST':
        logger.error("Método no permitido")
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        curso = get_object_or_404(Curso, id=curso_id)
        evaluacion = get_object_or_404(Evaluacion, id=evaluacion_id, curso=curso)
        criterio = get_object_or_404(CriterioRubrica, id=criterio_id, rubrica__evaluacion=evaluacion)
        
        # Verificar permisos
        if not request.user.is_staff and evaluacion.creado_por != request.user:
            logger.error(f"Usuario {request.user.id} no tiene permisos para evaluación {evaluacion_id}")
            return JsonResponse({'error': 'No tienes permisos para realizar esta acción'}, status=403)
        
        # Verificar si la evaluación ya comenzó (no se puede editar después de la fecha de inicio)
        from datetime import date
        hoy = date.today()
        if evaluacion.fecha_inicio and hoy >= evaluacion.fecha_inicio:
            logger.error(f"Evaluación {evaluacion_id} ya comenzó")
            return JsonResponse({
                'error': f'La evaluación comenzó el {evaluacion.fecha_inicio.strftime("%d/%m/%Y")}. No se pueden editar criterios después de esta fecha.'
            }, status=403)
        
        # Obtener datos del formulario
        nombre = request.POST.get('nombre')
        objetivo = request.POST.get('objetivo')
        puntaje = request.POST.get('puntaje')
        esperables_data = request.POST.getlist('esperables[]')
        
        logger.info(f"Datos recibidos - Nombre: {nombre}, Objetivo: {objetivo}, Puntaje: {puntaje}, Esperables: {len(esperables_data)}")
        
        if not nombre or not objetivo or not puntaje:
            logger.error("Campos obligatorios faltantes")
            return JsonResponse({'error': 'Nombre, objetivo y puntaje son campos obligatorios'}, status=400)
        
        try:
            puntaje_decimal = float(puntaje)
            if puntaje_decimal < 0:
                logger.error(f"Puntaje negativo: {puntaje_decimal}")
                return JsonResponse({'error': 'El puntaje debe ser un número positivo'}, status=400)
        except ValueError:
            logger.error(f"Puntaje inválido: {puntaje}")
            return JsonResponse({'error': 'El puntaje debe ser un número válido'}, status=400)
        
        # Actualizar el criterio
        criterio.nombre = nombre
        criterio.objetivo = objetivo
        criterio.puntaje = puntaje_decimal
        criterio.save()
        
        logger.info(f"Criterio actualizado con ID: {criterio.id}")
        
        # Actualizar los esperables
        # Primero eliminar los existentes
        criterio.esperables.all().delete()
        
        # Crear los nuevos esperables
        esperables_creados = 0
        for i, esperable_data in enumerate(esperables_data):
            if esperable_data.strip():  # Solo crear si no está vacío
                # Parsear el esperable que viene en formato JSON
                try:
                    esperable_obj = json.loads(esperable_data)
                    Esperable.objects.create(
                        criterio=criterio,
                        nivel=esperable_obj.get('nivel', f"Nivel {i+1}"),
                        descripcion=esperable_obj.get('descripcion', ''),
                        puntaje=esperable_obj.get('puntaje', 0),
                        orden=i+1
                    )
                    esperables_creados += 1
                except (json.JSONDecodeError, TypeError) as e:
                    logger.warning(f"Error parsing esperable {i}: {e}")
                    # Fallback para formato antiguo
                    Esperable.objects.create(
                        criterio=criterio,
                        nivel=f"Nivel {i+1}",
                        descripcion=esperable_data.strip(),
                        puntaje=0,
                        orden=i+1
                    )
                    esperables_creados += 1
        
        logger.info(f"Esperables creados: {esperables_creados}")
        
        # Recalcular el puntaje total del criterio basado en los esperables
        puntaje_total = criterio.esperables.aggregate(total=Max('puntaje'))['total'] or 0
        criterio.puntaje = puntaje_total
        criterio.save()
        
        logger.info(f"Criterio {criterio.id} editado exitosamente con puntaje total: {puntaje_total}")
        
        response_data = {
            'success': True,
            'criterio_id': criterio.id,
            'criterio_nombre': str(criterio.nombre),
            'criterio_objetivo': str(criterio.objetivo),
            'criterio_puntaje': float(criterio.puntaje),
            'esperables_count': criterio.esperables.count()
        }
        
        logger.info(f"Respuesta exitosa: {response_data}")
        return JsonResponse(response_data, safe=True)
        
    except Exception as e:
        logger.error(f"Error inesperado en editar_criterio_rubrica: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        return JsonResponse({'error': 'Error interno del servidor. Por favor, intenta nuevamente.'}, status=500)

@login_required
def eliminar_criterio_rubrica(request, curso_id, evaluacion_id, criterio_id):
    """
    Vista AJAX para eliminar un criterio de una rúbrica
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    curso = get_object_or_404(Curso, id=curso_id)
    evaluacion = get_object_or_404(Evaluacion, id=evaluacion_id, curso=curso)
    criterio = get_object_or_404(CriterioRubrica, id=criterio_id, rubrica__evaluacion=evaluacion)
    
    # Verificar permisos
    if not request.user.is_staff and evaluacion.creado_por != request.user:
        return JsonResponse({'error': 'No tienes permisos para realizar esta acción'}, status=403)
    
    # Verificar si la evaluación ya comenzó (no se puede editar después de la fecha de inicio)
    from datetime import date
    hoy = date.today()
    if evaluacion.fecha_inicio and hoy >= evaluacion.fecha_inicio:
        return JsonResponse({
            'error': f'La evaluación comenzó el {evaluacion.fecha_inicio.strftime("%d/%m/%Y")}. No se pueden eliminar criterios después de esta fecha.'
        }, status=403)
    
    criterio.delete()
    
    return JsonResponse({'success': True})

@login_required
def reasignar_ticket(request):
    """
    Vista para reasignar un ticket a otro usuario admin/staff
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        # Verificar que el usuario sea admin/staff
        if not (request.user.is_staff or request.user.is_superuser):
            return JsonResponse({'error': 'No tienes permisos para reasignar tickets'}, status=403)
        
        ticket_id = request.POST.get('ticket_id')
        usuario_asignar_id = request.POST.get('usuario_asignar')
        comentario_interno = request.POST.get('comentario_interno', '')
        
        if not ticket_id or not usuario_asignar_id:
            return JsonResponse({'error': 'Faltan datos requeridos'}, status=400)
        
        # Obtener el ticket
        ticket = TicketSoporte.objects.get(id=ticket_id)
        
        # Obtener el usuario al que se va a asignar
        usuario_asignar = CustomUser.objects.get(id=usuario_asignar_id)
        
        # Verificar que el usuario sea admin/staff
        if not (usuario_asignar.is_staff or usuario_asignar.is_superuser):
            return JsonResponse({'error': 'Solo se pueden asignar tickets a usuarios admin/staff'}, status=400)
        
        # Guardar el usuario anterior para el comentario
        usuario_anterior = ticket.asignado_a
        
        # Determinar si es una asignación inicial o una reasignación
        es_asignacion_inicial = usuario_anterior is None
        
        # Reasignar el ticket
        ticket.asignado_a = usuario_asignar
        
        # Si es una asignación inicial y el ticket está abierto, cambiar a "En proceso"
        if es_asignacion_inicial and ticket.estado == 'abierto':
            ticket.estado = 'en_proceso'
        
        ticket.save()
        
        # Crear comentario interno si se proporcionó
        if comentario_interno.strip():
            if es_asignacion_inicial:
                contenido = f"Ticket asignado a {usuario_asignar.get_full_name()}. {comentario_interno}"
            else:
                contenido = f"Ticket reasignado de {usuario_anterior.get_full_name()} a {usuario_asignar.get_full_name()}. {comentario_interno}"
            ComentarioTicket.objects.create(
                ticket=ticket,
                autor=request.user,
                contenido=contenido,
                es_interno=True
            )
        else:
            # Crear comentario automático
            if es_asignacion_inicial:
                contenido = f"Ticket asignado a {usuario_asignar.get_full_name()}."
            else:
                contenido = f"Ticket reasignado de {usuario_anterior.get_full_name()} a {usuario_asignar.get_full_name()}."
            ComentarioTicket.objects.create(
                ticket=ticket,
                autor=request.user,
                contenido=contenido,
                es_interno=True
            )
        
        # Determinar el mensaje de respuesta
        if es_asignacion_inicial:
            mensaje = f'Ticket asignado exitosamente a {usuario_asignar.get_full_name()} y estado cambiado a "En proceso"'
        else:
            mensaje = f'Ticket reasignado exitosamente a {usuario_asignar.get_full_name()}'
        
        return JsonResponse({
            'success': True,
            'message': mensaje
        })
        
    except TicketSoporte.DoesNotExist:
        return JsonResponse({'error': 'El ticket no existe'}, status=404)
    except CustomUser.DoesNotExist:
        return JsonResponse({'error': 'El usuario no existe'}, status=404)
    except Exception as e:
        logger.debug(f"Error reasignando ticket: {e}")
        return JsonResponse({'error': 'Error interno del servidor'}, status=500)


@login_required
def obtener_usuarios_staff(request):
    """
    Vista AJAX para obtener la lista de usuarios admin/staff para reasignación
    """
    try:
        # Verificar que el usuario sea admin/staff
        if not (request.user.is_staff or request.user.is_superuser):
            return JsonResponse({'error': 'No tienes permisos para ver usuarios staff'}, status=403)
        
        # Obtener usuarios admin/staff (excluyendo al usuario actual)
        usuarios_staff = CustomUser.objects.filter(
            models.Q(is_staff=True) | models.Q(is_superuser=True)
        ).exclude(id=request.user.id).order_by('first_name', 'last_name', 'username')
        
        usuarios_data = []
        for usuario in usuarios_staff:
            usuarios_data.append({
                'id': usuario.id,
                'nombre': usuario.get_full_name() or usuario.username,
                'email': usuario.email,
                'username': usuario.username
            })
        
        return JsonResponse({
            'success': True,
            'usuarios': usuarios_data
        })
        
    except Exception as e:
        logger.debug(f"Error obteniendo usuarios staff: {e}")
        return JsonResponse({'error': 'Error interno del servidor'}, status=500)

@login_required
def cambiar_prioridad_ticket(request):
    """
    Vista para cambiar la prioridad de un ticket
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        # Verificar que el usuario sea admin/staff
        if not (request.user.is_staff or request.user.is_superuser):
            return JsonResponse({'error': 'No tienes permisos para cambiar prioridades'}, status=403)
        
        ticket_id = request.POST.get('ticket_id')
        nueva_prioridad = request.POST.get('nueva_prioridad')
        comentario_prioridad = request.POST.get('comentario_prioridad', '')
        
        if not ticket_id or not nueva_prioridad:
            return JsonResponse({'error': 'Faltan datos requeridos'}, status=400)
        
        # Validar prioridad
        prioridades_validas = ['baja', 'media', 'alta', 'urgente']
        if nueva_prioridad not in prioridades_validas:
            return JsonResponse({'error': 'Prioridad no válida'}, status=400)
        
        # Obtener el ticket
        ticket = TicketSoporte.objects.get(id=ticket_id)
        
        # Guardar la prioridad anterior para el comentario
        prioridad_anterior = ticket.prioridad
        
        # Cambiar la prioridad
        ticket.prioridad = nueva_prioridad
        ticket.save()
        
        # Crear comentario si se proporcionó
        if comentario_prioridad.strip():
            ComentarioTicket.objects.create(
                ticket=ticket,
                autor=request.user,
                contenido=f"Prioridad cambiada de {prioridad_anterior} a {nueva_prioridad}. {comentario_prioridad}",
                es_interno=False
            )
        else:
            # Crear comentario automático
            ComentarioTicket.objects.create(
                ticket=ticket,
                autor=request.user,
                contenido=f"Prioridad cambiada de {prioridad_anterior} a {nueva_prioridad}.",
                es_interno=False
            )
        
        return JsonResponse({
            'success': True,
            'message': f'Prioridad cambiada exitosamente a {nueva_prioridad}'
        })
        
    except TicketSoporte.DoesNotExist:
        return JsonResponse({'error': 'El ticket no existe'}, status=404)
    except Exception as e:
        logger.debug(f"Error cambiando prioridad: {e}")
        return JsonResponse({'error': 'Error interno del servidor'}, status=500)


@login_required
def resolver_ticket(request):
    """
    Vista para resolver un ticket
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        # Verificar que el usuario sea admin/staff
        if not (request.user.is_staff or request.user.is_superuser):
            return JsonResponse({'error': 'No tienes permisos para resolver tickets'}, status=403)
        
        ticket_id = request.POST.get('ticket_id')
        comentario_resolucion = request.POST.get('comentario_resolucion')
        es_interno = request.POST.get('es_interno_resolucion') == 'on'
        
        if not ticket_id or not comentario_resolucion:
            return JsonResponse({'error': 'Faltan datos requeridos'}, status=400)
        
        # Obtener el ticket
        ticket = TicketSoporte.objects.get(id=ticket_id)
        
        # Verificar que el ticket no esté ya resuelto
        if ticket.estado == 'resuelto':
            return JsonResponse({'error': 'El ticket ya está resuelto'}, status=400)
        
        # Cambiar estado a resuelto
        ticket.estado = 'resuelto'
        ticket.fecha_resolucion = timezone.now()
        ticket.save()
        
        # Crear comentario de resolución
        ComentarioTicket.objects.create(
            ticket=ticket,
            autor=request.user,
            contenido=comentario_resolucion,
            es_interno=es_interno
        )
        
        # Agregar un campo temporal para identificar el comentario de resolución
        # Esto se puede hacer agregando un prefijo especial al contenido
        comentario_resolucion_obj = ComentarioTicket.objects.filter(
            ticket=ticket,
            autor=request.user,
            contenido=comentario_resolucion
        ).order_by('-fecha_creacion').first()
        
        if comentario_resolucion_obj:
            # Agregar un marcador especial al contenido para identificar que es de resolución
            comentario_resolucion_obj.contenido = f"[RESUELTO] {comentario_resolucion}"
            comentario_resolucion_obj.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Ticket resuelto exitosamente'
        })
        
    except TicketSoporte.DoesNotExist:
        return JsonResponse({'error': 'El ticket no existe'}, status=404)
    except Exception as e:
        logger.debug(f"Error resolviendo ticket: {e}")
        return JsonResponse({'error': 'Error interno del servidor'}, status=500)

@login_required
def reabrir_ticket(request):
    """
    Vista para reabrir un ticket resuelto
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        ticket_id = request.POST.get('ticket_id')
        comentario_reapertura = request.POST.get('comentario_reapertura')
        es_interno = request.POST.get('es_interno_reapertura') == 'on'
        
        if not ticket_id or not comentario_reapertura:
            return JsonResponse({'error': 'Faltan datos requeridos'}, status=400)
        
        # Obtener el ticket
        ticket = TicketSoporte.objects.get(id=ticket_id)
        
        # Verificar que el usuario sea admin/staff
        if not (request.user.is_staff or request.user.is_superuser):
            return JsonResponse({'error': 'No tienes permisos para reabrir tickets'}, status=403)
        
        # Verificar que el ticket esté resuelto
        if ticket.estado != 'resuelto':
            return JsonResponse({'error': 'Solo se pueden reabrir tickets resueltos'}, status=400)
        
        # Cambiar estado a en_proceso
        ticket.estado = 'en_proceso'
        ticket.fecha_resolucion = None  # Limpiar fecha de resolución
        ticket.save()
        
        # Crear comentario de reapertura con tag
        ComentarioTicket.objects.create(
            ticket=ticket,
            autor=request.user,
            contenido=f"[REAPERTURA] {comentario_reapertura}",
            es_interno=es_interno
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Ticket reabierto exitosamente'
        })
        
    except TicketSoporte.DoesNotExist:
        return JsonResponse({'error': 'El ticket no existe'}, status=404)
    except Exception as e:
        logger.debug(f"Error reabriendo ticket: {e}")
        return JsonResponse({'error': 'Error interno del servidor'}, status=500)

@login_required
def obtener_datos_calificacion(request, calificacion_id):
    """
    Vista para obtener los datos de criterios de una calificación existente
    """
    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    # Verificar permisos
    if not request.user.is_staff:
        return JsonResponse({'error': 'No tienes permisos para acceder a estos datos'}, status=403)
    
    try:
        calificacion = get_object_or_404(Calificacion, id=calificacion_id)
        evaluacion = calificacion.evaluacion
        
        logger.debug(f"Obteniendo datos para calificación {calificacion_id}, evaluación: {evaluacion.nombre}")
        
        # Verificar que la evaluación tenga rúbrica
        if not hasattr(evaluacion, 'rubrica') or not evaluacion.rubrica:
            logger.debug(f"Evaluación {evaluacion.id} no tiene rúbrica asociada")
            return JsonResponse({'error': 'Esta evaluación no tiene rúbrica asociada'}, status=400)
        
        # Obtener el resultado de rúbrica para este estudiante
        try:
            resultado_rubrica = ResultadoRubrica.objects.get(
                rubrica=evaluacion.rubrica,
                estudiante=calificacion.estudiante
            )

            logger.debug(f"ResultadoRubrica encontrado: {resultado_rubrica}")
            logger.info(f"Resultado de rúbrica: {resultado_rubrica}")
            
            # Obtener los puntajes de criterios
            puntajes_criterios = resultado_rubrica.puntajes_criterios.all()

            logger.debug(f"Puntajes de criterios encontrados: {puntajes_criterios.count()}")
            logger.debug(f"Puntajes de criterios: {puntajes_criterios}")
            logger.debug(f"Número de puntajes de criterios: {puntajes_criterios.count()}")
            
            for puntaje in puntajes_criterios:
                logger.debug(f"Puntaje criterio {puntaje.criterio.id}: esperable_id={puntaje.esperable_seleccionado.id if puntaje.esperable_seleccionado else None}, puntaje_obtenido={puntaje.puntaje_obtenido}")
                logger.debug(f"Puntaje criterio {puntaje.criterio.id}: esperable_id={puntaje.esperable_seleccionado.id if puntaje.esperable_seleccionado else None}, puntaje_obtenido={puntaje.puntaje_obtenido}")
            
            criterios_data = []
            for puntaje in puntajes_criterios:
                criterio_data = {
                    'criterio_id': puntaje.criterio.id,
                    'criterio_nombre': puntaje.criterio.nombre,
                    'esperable_id': puntaje.esperable_seleccionado.id if puntaje.esperable_seleccionado else None,
                    'esperable_nivel': puntaje.esperable_seleccionado.nivel if puntaje.esperable_seleccionado else None,
                    'esperable_descripcion': puntaje.esperable_seleccionado.descripcion if puntaje.esperable_seleccionado else None,
                    'esperable_puntaje': float(puntaje.esperable_seleccionado.puntaje) if puntaje.esperable_seleccionado else 0,
                    'puntaje_obtenido': float(puntaje.puntaje_obtenido),
                    'comentarios': puntaje.comentarios or ''
                }
                criterios_data.append(criterio_data)
                logger.debug(f"Criterio {puntaje.criterio.id}: esperable_id={criterio_data['esperable_id']}, puntaje={criterio_data['puntaje_obtenido']}")
            
            logger.debug(f"Criterios data final: {criterios_data}")
            logger.debug(f"Criterios data: {criterios_data}")
            
            return JsonResponse({
                'success': True,
                'criterios': criterios_data,
                'puntaje_total': float(resultado_rubrica.puntaje_total) if resultado_rubrica.puntaje_total else 0,
                'nota_final': float(resultado_rubrica.nota_final) if resultado_rubrica.nota_final else 0
            })
            
        except ResultadoRubrica.DoesNotExist:
            logger.debug(f"No existe ResultadoRubrica para estudiante {calificacion.estudiante.id}")
            # Si no existe ResultadoRubrica, devolver criterios vacíos para permitir edición
            criterios_data = []
            for criterio in evaluacion.rubrica.criterios.all():
                criterios_data.append({
                    'criterio_id': criterio.id,
                    'criterio_nombre': criterio.nombre,
                    'esperable_id': None,
                    'puntaje_obtenido': 0,
                    'comentarios': ''
                })
            
            logger.debug(f"Criterios data vacíos: {criterios_data}")
            
            return JsonResponse({
                'success': True,
                'criterios': criterios_data,
                'puntaje_total': 0,
                'nota_final': 0
            })
            
    except Exception as e:
        logger.debug(f"Error en obtener_datos_calificacion: {str(e)}")
        return JsonResponse({'error': f'Error al obtener datos: {str(e)}'}, status=500)

@login_required
def obtener_esperables_criterio_por_estudiante(request, criterio_id, estudiante_id):
    """
    Vista para obtener los esperables específicos aplicados a un estudiante para un criterio
    """
    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    # Verificar permisos
    if not request.user.is_staff:
        return JsonResponse({'error': 'No tienes permisos para acceder a estos datos'}, status=403)
    
    try:
        criterio = get_object_or_404(CriterioRubrica, id=criterio_id)
        estudiante = get_object_or_404(CustomUser, id=estudiante_id)
        
        logger.debug(f"Buscando esperables para criterio {criterio_id} y estudiante {estudiante_id}")
        
        # Obtener el resultado de rúbrica para este estudiante
        try:
            resultado_rubrica = ResultadoRubrica.objects.get(
                rubrica=criterio.rubrica,
                estudiante=estudiante
            )

            logger.debug(f"ResultadoRubrica encontrado: {resultado_rubrica}")
            
            # Obtener el puntaje específico para este criterio
            puntaje_criterio = PuntajeCriterio.objects.filter(
                resultado_rubrica=resultado_rubrica,
                criterio=criterio
            ).first()
            
            logger.debug(f"PuntajeCriterio encontrado: {puntaje_criterio}")
            
            if puntaje_criterio and puntaje_criterio.esperable_seleccionado:
                # Devolver todos los esperables disponibles, marcando el seleccionado
                esperable_aplicado = puntaje_criterio.esperable_seleccionado
                esperables = criterio.esperables.all()
                esperables_data = []
                
                logger.debug(f"Esperable aplicado: {esperable_aplicado}")
                logger.debug(f"Total esperables disponibles: {esperables.count()}")
                
                for esperable in esperables:
                    esperables_data.append({
                        'id': esperable.id,
                        'nivel': esperable.nivel,
                        'descripcion': esperable.descripcion,
                        'puntaje': float(esperable.puntaje),
                        'es_seleccionado': esperable.id == esperable_aplicado.id
                    })
                
                logger.debug(f"Esperables data: {esperables_data}")
                
                return JsonResponse({
                    'success': True,
                    'esperables': esperables_data,
                    'esperable_seleccionado': esperable_aplicado.id
                })
            else:
                # Si no hay esperable seleccionado, devolver todos los esperables disponibles
                esperables = criterio.esperables.all()
                esperables_data = []
                for esperable in esperables:
                    esperables_data.append({
                        'id': esperable.id,
                        'nivel': esperable.nivel,
                        'descripcion': esperable.descripcion,
                        'puntaje': float(esperable.puntaje),
                        'es_seleccionado': False
                    })
                
                logger.debug(f"No hay esperable seleccionado, devolviendo todos los esperables: {esperables_data}")
                
                return JsonResponse({
                    'success': True,
                    'esperables': esperables_data,
                    'esperable_seleccionado': None
                })
                
        except ResultadoRubrica.DoesNotExist:
            # Si no existe resultado de rúbrica, devolver todos los esperables disponibles
            logger.debug(f"No existe ResultadoRubrica para estudiante {estudiante_id}")
            esperables = criterio.esperables.all()
            esperables_data = []
            for esperable in esperables:
                esperables_data.append({
                    'id': esperable.id,
                    'nivel': esperable.nivel,
                    'descripcion': esperable.descripcion,
                    'puntaje': float(esperable.puntaje),
                    'es_seleccionado': False
                })
            
            return JsonResponse({
                'success': True,
                'esperables': esperables_data,
                'esperable_seleccionado': None
            })
        
    except Exception as e:
        logger.debug(f"Error en obtener_esperables_criterio_por_estudiante: {str(e)}")
        return JsonResponse({'error': f'Error al obtener esperables: {str(e)}'}, status=500)

@login_required
def obtener_esperables_criterio(request, criterio_id):
    """
    Vista para obtener todos los esperables disponibles para un criterio
    """
    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    # Verificar permisos
    if not request.user.is_staff:
        return JsonResponse({'error': 'No tienes permisos para acceder a estos datos'}, status=403)
    
    try:
        criterio = get_object_or_404(CriterioRubrica, id=criterio_id)
        esperables = criterio.esperables.all()
        
        esperables_data = []
        for esperable in esperables:
            esperables_data.append({
                'id': esperable.id,
                'nivel': esperable.nivel,
                'descripcion': esperable.descripcion,
                'puntaje': float(esperable.puntaje),
                'es_seleccionado': False
            })
        
        return JsonResponse({
            'success': True,
            'esperables': esperables_data,
            'esperable_seleccionado': None
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Error al obtener esperables: {str(e)}'}, status=500)

@login_required
def debug_calificaciones(request, curso_id):
    """
    Vista de debug para verificar el estado del sistema de calificaciones
    """
    if not request.user.is_staff:
        return JsonResponse({'error': 'No tienes permisos para acceder a estos datos'}, status=403)
    
    try:
        curso = get_object_or_404(Curso, id=curso_id)
        evaluaciones = curso.evaluaciones.all()
        
        debug_data = {
            'curso': {
                'id': curso.id,
                'nombre': curso.nombre,
                'evaluaciones_count': evaluaciones.count()
            },
            'evaluaciones': []
        }
        
        for evaluacion in evaluaciones:
            eval_data = {
                'id': evaluacion.id,
                'nombre': evaluacion.nombre,
                'nota_maxima': float(evaluacion.nota_maxima),
                'tiene_rubrica': hasattr(evaluacion, 'rubrica') and evaluacion.rubrica is not None
            }
            
            if eval_data['tiene_rubrica']:
                rubrica = evaluacion.rubrica
                eval_data['rubrica'] = {
                    'id': rubrica.id,
                    'nombre': rubrica.nombre,
                    'criterios_count': rubrica.criterios.count(),
                    'puntaje_total': float(rubrica.get_puntaje_total()),
                    'criterios': []
                }
                
                for criterio in rubrica.criterios.all():
                    criterio_data = {
                        'id': criterio.id,
                        'nombre': criterio.nombre,
                        'puntaje': float(criterio.puntaje),
                        'esperables_count': criterio.esperables.count(),
                        'esperables': []
                    }
                    
                    for esperable in criterio.esperables.all():
                        esperable_data = {
                            'id': esperable.id,
                            'nivel': esperable.nivel,
                            'puntaje': float(esperable.puntaje),
                            'descripcion': esperable.descripcion
                        }
                        criterio_data['esperables'].append(esperable_data)
                    
                    eval_data['rubrica']['criterios'].append(criterio_data)
            
            debug_data['evaluaciones'].append(eval_data)
        
        return JsonResponse({
            'success': True,
            'debug_data': debug_data
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Error en debug: {str(e)}'}, status=500)

@login_required
def plataforma_calificaciones_ajax(request, curso_id):
    """
    Vista AJAX para cargar el contenido de calificaciones dinámicamente
    """
    try:
        logger.debug(f"DEBUG: Iniciando plataforma_calificaciones_ajax")
        logger.debug(f"DEBUG: curso_id: {curso_id}")
        logger.debug(f"DEBUG: usuario: {request.user.username}")
        logger.debug(f"DEBUG: es_staff: {request.user.is_staff}")
        
        curso = Curso.objects.get(id=curso_id, activo=True)
        logger.debug(f"DEBUG: Curso encontrado: {curso.nombre}")
        
        # Verificar que el usuario esté inscrito en el curso
        if curso not in request.user.cursos.all():
            return JsonResponse({'error': 'No tienes acceso a este curso'}, status=403)
        
        # Verificar si es una acción específica
        action = request.GET.get('action')
        
        if action == 'ver_rubricas':
            # Cargar vista de rúbricas para todos los usuarios
            evaluaciones_con_rubricas = []
            evaluaciones = Evaluacion.objects.filter(curso=curso, activa=True).order_by('fecha_inicio')
            
            for evaluacion in evaluaciones:
                try:
                    rubrica = evaluacion.rubrica
                    if rubrica and rubrica.activa:
                        evaluaciones_con_rubricas.append({
                            'evaluacion': evaluacion,
                            'rubrica': rubrica
                        })
                except:
                    continue
            
            context = {
                'curso': curso,
                'evaluaciones_con_rubricas': evaluaciones_con_rubricas,
                'mostrar_rubricas': True,
            }
            
            html = render_to_string('pages/plataforma_calificaciones_rubricas_content.html', context, request=request)
            return JsonResponse({'html': html})
        
        elif action == 'ver_calificar':
            # Cargar vista de calificar estudiantes
            from .forms import CalificacionForm
            
            evaluacion_id = request.GET.get('evaluacion_id')
            if not evaluacion_id:
                return JsonResponse({'error': 'ID de evaluación requerido'}, status=400)
            
            try:
                evaluacion = Evaluacion.objects.get(id=evaluacion_id, curso=curso, activa=True)
                
                # Verificar permisos
                if not request.user.is_staff:
                    return JsonResponse({'error': 'No tienes permisos para calificar estudiantes'}, status=403)
                
                # Obtener estudiantes del curso que tienen entregas para esta evaluación
                estudiantes_con_entregas = User.objects.filter(
                    cursos=curso,
                    is_staff=False,
                    is_superuser=False,
                    entregas__evaluacion=evaluacion
                ).distinct().order_by('first_name', 'last_name', 'username')
                
                # Para el formulario: solo estudiantes no calificados
                estudiantes_calificados_ids = evaluacion.calificaciones.values_list('estudiante_id', flat=True)
                estudiantes_disponibles_para_calificar = estudiantes_con_entregas.exclude(id__in=estudiantes_calificados_ids)
                
                # Si no hay estudiantes con entregas, mostrar mensaje
                if not estudiantes_con_entregas.exists():
                    # Obtener todos los estudiantes del curso para mostrar el mensaje
                    todos_estudiantes = User.objects.filter(
                        cursos=curso,
                        is_staff=False,
                        is_superuser=False
                    ).order_by('first_name', 'last_name', 'username')
                    
                    # Crear formulario vacío
                    form = CalificacionForm(curso=curso, evaluacion=evaluacion, estudiantes_con_entregas=[])
                    
                    # Obtener criterios de la rúbrica si existe
                    criterios_rubrica = []
                    if evaluacion.rubrica:
                        criterios_rubrica = evaluacion.rubrica.criterios.all()
                    
                    context = {
                        'curso': curso,
                        'evaluacion': evaluacion,
                        'form': form,
                        'estudiantes': [],
                        'todos_estudiantes': todos_estudiantes,
                        'criterios_rubrica': criterios_rubrica,
                        'mensaje_error': f'No hay estudiantes con entregas para la evaluación "{evaluacion.nombre}". Solo se pueden calificar estudiantes que hayan entregado su trabajo.',
                        'mostrar_calificar': True,
                    }
                    html = render_to_string('pages/calificar_estudiante.html', context, request=request)
                    return JsonResponse({'html': html})
                
                # Obtener calificaciones existentes
                calificaciones_existentes = evaluacion.calificaciones.all().order_by('estudiante__first_name', 'estudiante__last_name')
                
                # Obtener todos los estudiantes del curso
                todos_estudiantes = User.objects.filter(
                    cursos=curso,
                    is_staff=False,
                    is_superuser=False
                ).order_by('first_name', 'last_name', 'username')
                
                # Crear formulario de calificación
                form = CalificacionForm(curso=curso, evaluacion=evaluacion, estudiantes_con_entregas=estudiantes_disponibles_para_calificar)
                
                # Obtener criterios de la rúbrica si existe
                criterios_rubrica = []
                if evaluacion.rubrica:
                    criterios_rubrica = evaluacion.rubrica.criterios.all()
                
                # Calcular estadísticas
                stats = {
                    'total_estudiantes_curso': todos_estudiantes.count(),
                    'total_estudiantes_con_entregas': estudiantes_con_entregas.count(),
                    'estudiantes_ya_calificados': evaluacion.calificaciones.count(),
                    'estudiantes_disponibles_para_calificar': estudiantes_disponibles_para_calificar.count(),
                }
                
                context = {
                    'curso': curso,
                    'evaluacion': evaluacion,
                    'form': form,
                    'estudiantes': estudiantes_disponibles_para_calificar,
                    'todos_estudiantes': todos_estudiantes,
                    'criterios_rubrica': criterios_rubrica,
                    'stats': stats,
                    'mostrar_calificar': True,
                }
                
                html = render_to_string('pages/calificar_estudiante.html', context, request=request)
                return JsonResponse({'html': html})
                
            except Evaluacion.DoesNotExist:
                return JsonResponse({'error': 'Evaluación no encontrada'}, status=404)
            except Exception as e:
                import traceback
                logger.debug(f"Error al cargar calificar: {str(e)}")
                logger.debug(f"Traceback: {traceback.format_exc()}")
                return JsonResponse({'error': f'Error al cargar la vista de calificación: {str(e)}'}, status=500)
        
        elif action == 'eliminar_evaluacion':
            # Cargar vista de eliminación de evaluación
            evaluacion_id = request.GET.get('evaluacion_id')
            if not evaluacion_id:
                return JsonResponse({'error': 'ID de evaluación requerido'}, status=400)
            
            try:
                evaluacion = Evaluacion.objects.get(id=evaluacion_id, curso=curso, activa=True)
                
                # Verificar permisos
                if not request.user.is_staff:
                    return JsonResponse({'error': 'No tienes permisos para eliminar evaluaciones'}, status=403)
                
                # Verificar si la evaluación tiene calificaciones
                tiene_calificaciones = evaluacion.calificaciones.exists()
                
                context = {
                    'curso': curso,
                    'evaluacion': evaluacion,
                    'tiene_calificaciones': tiene_calificaciones,
                    'mostrar_eliminar': True,
                }
                
                html = render_to_string('pages/plataforma_calificaciones_eliminar_content.html', context, request=request)
                return JsonResponse({'html': html})
                
            except Evaluacion.DoesNotExist:
                return JsonResponse({'error': 'Evaluación no encontrada'}, status=404)
            except Exception as e:
                import traceback
                logger.debug(f"Error al cargar eliminación: {str(e)}")
                logger.debug(f"Traceback: {traceback.format_exc()}")
                return JsonResponse({'error': f'Error al cargar la vista de eliminación: {str(e)}'}, status=500)
        
        elif action == 'ver_editar_evaluacion':
            # Cargar vista de editar evaluación
            evaluacion_id = request.GET.get('evaluacion_id')
            if not evaluacion_id:
                return JsonResponse({'error': 'ID de evaluación requerido'}, status=400)
            
            try:
                evaluacion = Evaluacion.objects.get(id=evaluacion_id, curso=curso)
                
                # Verificar permisos
                if not request.user.is_staff:
                    return JsonResponse({'error': 'No tienes permisos para editar evaluaciones'}, status=403)
                
                # Crear formulario
                from .forms import EvaluacionForm
                form = EvaluacionForm(instance=evaluacion, curso=curso)
                
                # Asegurar que las fechas estén en formato ISO para HTML5 date inputs
                # Forzar el formato ISO independientemente de la configuración de Django
                if evaluacion.fecha_inicio:
                    form.fields['fecha_inicio'].widget.attrs['value'] = evaluacion.fecha_inicio.strftime('%Y-%m-%d')
                    # También establecer el formato de visualización para evitar problemas de localización
                    form.fields['fecha_inicio'].widget.format = '%Y-%m-%d'
                if evaluacion.fecha_fin:
                    form.fields['fecha_fin'].widget.attrs['value'] = evaluacion.fecha_fin.strftime('%Y-%m-%d')
                    # También establecer el formato de visualización para evitar problemas de localización
                    form.fields['fecha_fin'].widget.format = '%Y-%m-%d'
                
                context = {
                    'curso': curso,
                    'evaluacion': evaluacion,
                    'form': form,
                    'user': request.user,
                }
                
                # Renderizar solo el contenido del formulario
                html = render_to_string('pages/editar_evaluacion_content.html', context, request=request)
                return JsonResponse({'html': html})
                
            except Evaluacion.DoesNotExist:
                return JsonResponse({'error': 'Evaluación no encontrada'}, status=404)
            except Exception as e:
                import traceback
                logger.debug(f"Error al cargar editar evaluación: {str(e)}")
                logger.debug(f"Traceback: {traceback.format_exc()}")
                return JsonResponse({'error': f'Error al cargar la vista de editar evaluación: {str(e)}'}, status=500)
        
        elif action == 'ver_rubrica':
            # Cargar vista de rúbrica específica
            evaluacion_id = request.GET.get('evaluacion_id')
            if not evaluacion_id:
                return JsonResponse({'error': 'ID de evaluación requerido'}, status=400)
            
            try:
                evaluacion = Evaluacion.objects.get(id=evaluacion_id, curso=curso, activa=True)
                
                # Verificar que el usuario tenga permisos
                if not request.user.is_staff and evaluacion.creado_por != request.user:
                    return JsonResponse({'error': 'No tienes permisos para ver esta rúbrica'}, status=403)
                
                # Obtener la rúbrica
                try:
                    rubrica = evaluacion.rubrica
                except Rubrica.DoesNotExist:
                    return JsonResponse({'error': 'No se encontró una rúbrica para esta evaluación'}, status=404)
                
                # Verificar si la evaluación ya comenzó (no se puede editar después de la fecha de inicio)
                from datetime import date
                hoy = date.today()
                puede_editar = True
                mensaje_restriccion = None
                
                if evaluacion.fecha_inicio and hoy >= evaluacion.fecha_inicio:
                    puede_editar = False
                    mensaje_restriccion = f'La evaluación comenzó el {evaluacion.fecha_inicio.strftime("%d/%m/%Y")}. La rúbrica no se puede editar después de esta fecha.'
                
                context = {
                    'curso': curso,
                    'evaluacion': evaluacion,
                    'rubrica': rubrica,
                    'user': request.user,
                    'puede_editar': puede_editar,
                    'mensaje_restriccion': mensaje_restriccion,
                    'mostrar_rubrica': True,
                }
                
                html = render_to_string('pages/plataforma_calificaciones_rubrica_content.html', context, request=request)
                return JsonResponse({'html': html})
                
            except Evaluacion.DoesNotExist:
                return JsonResponse({'error': 'Evaluación no encontrada'}, status=404)
            except Exception as e:
                logger.debug(f"Error al cargar rúbrica: {str(e)}")
                return JsonResponse({'error': 'Error al cargar la rúbrica'}, status=500)
        
        elif action == 'crear_evaluacion':
            # Cargar formulario de crear evaluación
            from .forms import EvaluacionForm
            
            # Verificar permisos de staff
            if not request.user.is_staff:
                return JsonResponse({'error': 'No tienes permisos para crear evaluaciones'}, status=403)
            
            form = EvaluacionForm(curso=curso)
            context = {
                'curso': curso,
                'form': form,
                'user': request.user,
            }
            html = render_to_string('pages/plataforma_calificaciones_crear_evaluacion_content.html', context, request=request)
            return JsonResponse({'html': html})
        
        elif action == 'ver_estadisticas':
            # Cargar vista de estadísticas
            try:
                logger.debug(f"=== INICIANDO ESTADISTICAS ===")
                logger.debug(f"Usuario: {request.user.username}")
                logger.debug(f"Es staff: {request.user.is_staff}")
                logger.debug(f"Curso ID: {curso.id}")
                
                if request.user.is_staff:
                    # Estadísticas para staff
                    evaluaciones = Evaluacion.objects.filter(curso=curso, activa=True).order_by('-fecha_creacion')
                    calificaciones_curso = Calificacion.objects.filter(evaluacion__curso=curso, nota__isnull=False)
                    
                    # Estadísticas generales
                    if calificaciones_curso.exists():
                        estadisticas = {
                            'promedio_general': calificaciones_curso.aggregate(Avg('nota'))['nota__avg'],
                            'nota_minima': calificaciones_curso.aggregate(Min('nota'))['nota__min'],
                            'nota_maxima': calificaciones_curso.aggregate(Max('nota'))['nota__max'],
                            'total_estudiantes': curso.usuarios.filter(is_staff=False, is_superuser=False).count(),
                        }
                    else:
                        estadisticas = None
                    
                    # Estadísticas por evaluación
                    estadisticas_evaluaciones = []
                    for evaluacion in evaluaciones:
                        calificaciones_eval = calificaciones_curso.filter(evaluacion=evaluacion)
                        stats = {
                            'evaluacion': evaluacion,
                            'total_calificaciones': calificaciones_eval.count(),
                            'promedio': calificaciones_eval.aggregate(Avg('nota'))['nota__avg'] if calificaciones_eval.exists() else None,
                            'nota_minima': calificaciones_eval.aggregate(Min('nota'))['nota__min'] if calificaciones_eval.exists() else None,
                            'nota_maxima': calificaciones_eval.aggregate(Max('nota'))['nota__max'] if calificaciones_eval.exists() else None,
                        }
                        estadisticas_evaluaciones.append(stats)
                    
                    # Estadísticas por estudiante
                    estudiantes = curso.usuarios.filter(is_staff=False, is_superuser=False).order_by('first_name', 'last_name')
                    estadisticas_estudiantes = []
                    
                    for estudiante in estudiantes:
                        calificaciones_est = calificaciones_curso.filter(estudiante=estudiante)
                        stats = {
                            'estudiante': estudiante,
                            'total_calificaciones': calificaciones_est.count(),
                            'promedio': calificaciones_est.aggregate(Avg('nota'))['nota__avg'] if calificaciones_est.exists() else None,
                            'nota_minima': calificaciones_est.aggregate(Min('nota'))['nota__min'] if calificaciones_est.exists() else None,
                            'nota_maxima': calificaciones_est.aggregate(Max('nota'))['nota__max'] if calificaciones_est.exists() else None,
                        }
                        estadisticas_estudiantes.append(stats)
                    
                    context = {
                        'curso': curso,
                        'estadisticas': estadisticas,
                        'estadisticas_evaluaciones': estadisticas_evaluaciones,
                        'estadisticas_estudiantes': estadisticas_estudiantes,
                        'evaluaciones': evaluaciones,
                        'mostrar_estadisticas': True,
                    }
                    
                    html = render_to_string('pages/estadisticas_curso_content.html', context, request=request)
                    
                    logger.debug(f"=== DEBUG ESTADISTICAS ===")
                    logger.debug(f"Estadísticas evaluaciones: {len(estadisticas_evaluaciones)}")
                    logger.debug(f"Estadísticas estudiantes: {len(estadisticas_estudiantes)}")
                    logger.debug(f"HTML length: {len(html)}")
                    
                    return JsonResponse({'html': html})
                else:
                    # Estadísticas para estudiantes
                    calificaciones_usuario = Calificacion.objects.filter(
                        evaluacion__curso=curso,
                        estudiante=request.user
                    ).order_by('-fecha_calificacion')
                    
                    calificaciones_con_nota = calificaciones_usuario.filter(nota__isnull=False)
                    total_evaluaciones = Evaluacion.objects.filter(curso=curso).count()
                    
                    if calificaciones_con_nota.exists():
                        # Calcular promedio ponderado
                        suma_ponderada = 0
                        suma_ponderaciones = 0
                        evaluaciones_calificadas = 0
                        
                        for calificacion in calificaciones_con_nota:
                            nota_ponderada = calificacion.nota * calificacion.evaluacion.ponderacion
                            suma_ponderada += nota_ponderada
                            suma_ponderaciones += calificacion.evaluacion.ponderacion
                            evaluaciones_calificadas += 1
                        
                        if suma_ponderaciones > 0:
                            promedio_ponderado = suma_ponderada / suma_ponderaciones
                            estadisticas_estudiante = {
                                'promedio_ponderado': promedio_ponderado,
                                'evaluaciones_calificadas': evaluaciones_calificadas,
                                'total_evaluaciones': total_evaluaciones,
                                'suma_ponderaciones': suma_ponderaciones
                            }
                        else:
                            estadisticas_estudiante = {
                                'evaluaciones_calificadas': evaluaciones_calificadas,
                                'total_evaluaciones': total_evaluaciones,
                                'suma_ponderaciones': suma_ponderaciones
                            }
                    else:
                        estadisticas_estudiante = {
                            'evaluaciones_calificadas': 0,
                            'total_evaluaciones': total_evaluaciones,
                            'suma_ponderaciones': 0
                        }
                    
                    context = {
                        'curso': curso,
                        'calificaciones_usuario': calificaciones_usuario,
                        'estadisticas_estudiante': estadisticas_estudiante,
                        'mostrar_estadisticas': True,
                    }
                    
                    html = render_to_string('pages/estadisticas_curso_content.html', context, request=request)
                    return JsonResponse({'html': html})
            except Exception as e:
                logger.debug(f"=== ERROR EN ESTADISTICAS ===")
                logger.debug(f"Error: {str(e)}")
                import traceback
                traceback.logger.debug_exc()
                return JsonResponse({'error': 'Error al cargar las estadísticas'}, status=500)
        
        elif action == 'ver_evaluacion':
            # Cargar detalle de una evaluación específica
            evaluacion_id = request.GET.get('evaluacion_id')
            if evaluacion_id:
                try:
                    evaluacion = Evaluacion.objects.get(id=evaluacion_id, curso=curso, activa=True)
                    
                    if request.user.is_staff:
                        # Vista para staff - mostrar entregas y calificaciones
                        entregas = evaluacion.entregas.all().order_by('-fecha_entrega')
                        calificaciones = evaluacion.calificaciones.all().order_by('-fecha_calificacion')
                        
                        context = {
                            'curso': curso,
                            'evaluacion': evaluacion,
                            'entregas': entregas,
                            'calificaciones': calificaciones,
                            'mostrar_detalle_evaluacion': True,
                        }
                    else:
                        # Vista para estudiantes - mostrar su entrega y calificación
                        try:
                            entrega = evaluacion.entregas.get(estudiante=request.user)
                        except Entrega.DoesNotExist:
                            entrega = None
                        
                        try:
                            calificacion = evaluacion.calificaciones.get(estudiante=request.user)
                        except Calificacion.DoesNotExist:
                            calificacion = None
                        
                        context = {
                            'curso': curso,
                            'evaluacion': evaluacion,
                            'entrega': entrega,
                            'calificacion': calificacion,
                            'mostrar_detalle_evaluacion': True,
                        }
                    
                    html = render_to_string('pages/plataforma_calificaciones_evaluacion_content.html', context, request=request)
                    return JsonResponse({'html': html})
                except Evaluacion.DoesNotExist:
                    return JsonResponse({'error': 'La evaluación no existe'}, status=404)
            else:
                return JsonResponse({'error': 'ID de evaluación no proporcionado'}, status=400)
        
        # Vista principal de calificaciones (sin acción específica)
        context = {
            'curso': curso,
            'user': request.user,
        }
        
        if request.user.is_staff:
            # Vista para Staff/Admin
            evaluaciones = Evaluacion.objects.filter(curso=curso, activa=True).order_by('-fecha_creacion')
            context['evaluaciones'] = evaluaciones
            
            # Debug: Imprimir información sobre las evaluaciones
            logger.debug(f"DEBUG: Usuario es staff: {request.user.is_staff}")
            logger.debug(f"DEBUG: Curso ID: {curso.id}")
            logger.debug(f"DEBUG: Evaluaciones encontradas: {evaluaciones.count()}")
            for eval in evaluaciones:
                logger.debug(f"DEBUG: Evaluación: {eval.nombre} (ID: {eval.id})")
            
            # Estadísticas generales del curso
            calificaciones_curso = Calificacion.objects.filter(evaluacion__curso=curso, nota__isnull=False)
            if calificaciones_curso.exists():
                estadisticas = {
                    'promedio_general': calificaciones_curso.aggregate(Avg('nota'))['nota__avg'],
                    'nota_minima': calificaciones_curso.aggregate(Min('nota'))['nota__min'],
                    'nota_maxima': calificaciones_curso.aggregate(Max('nota'))['nota__max'],
                    'total_estudiantes': curso.usuarios.filter(is_staff=False, is_superuser=False).count(),
                }
                context['estadisticas'] = estadisticas
            
            # Estadísticas de entregas por evaluación
            for evaluacion in evaluaciones:
                total_estudiantes = curso.usuarios.filter(is_staff=False, is_superuser=False).count()
                estudiantes_con_entregas = evaluacion.entregas.values('estudiante').distinct().count()
                estudiantes_calificados = evaluacion.calificaciones.count()
                
                evaluacion.stats_entregas = {
                    'total_estudiantes': total_estudiantes,
                    'estudiantes_con_entregas': estudiantes_con_entregas,
                    'estudiantes_sin_entregas': total_estudiantes - estudiantes_con_entregas,
                    'estudiantes_calificados': estudiantes_calificados,
                    'estudiantes_pendientes_calificacion': estudiantes_con_entregas - estudiantes_calificados
                }
            
            # Agregar información sobre la nueva funcionalidad
            context['info_entregas'] = {
                'mensaje': '💡 Solo se pueden calificar estudiantes que hayan entregado su trabajo. Desde la primera entrega se puede acceder a calificar una evaluación.',
                'total_evaluaciones': evaluaciones.count(),
                'evaluaciones_con_entregas': sum(1 for e in evaluaciones if e.stats_entregas['estudiantes_con_entregas'] > 0)
            }
            
        else:
            # Vista para Estudiantes
            calificaciones_usuario = Calificacion.objects.filter(
                evaluacion__curso=curso,
                estudiante=request.user
            ).order_by('-fecha_calificacion')
            context['calificaciones_usuario'] = calificaciones_usuario
            
            # Estadísticas personales del estudiante
            calificaciones_con_nota = calificaciones_usuario.filter(nota__isnull=False)
            total_evaluaciones = Evaluacion.objects.filter(curso=curso).count()
            
            if calificaciones_con_nota.exists():
                # Calcular promedio ponderado
                suma_ponderada = 0
                suma_ponderaciones = 0
                evaluaciones_calificadas = 0
                
                for calificacion in calificaciones_con_nota:
                    # Calcular nota ponderada: nota * ponderacion
                    nota_ponderada = calificacion.nota * calificacion.evaluacion.ponderacion
                    suma_ponderada += nota_ponderada
                    suma_ponderaciones += calificacion.evaluacion.ponderacion
                    evaluaciones_calificadas += 1
                
                # Calcular promedio ponderado
                if suma_ponderaciones > 0:
                    promedio_ponderado = suma_ponderada / suma_ponderaciones
                    context['estadisticas_estudiante'] = {
                        'promedio_ponderado': promedio_ponderado,
                        'evaluaciones_calificadas': evaluaciones_calificadas,
                        'total_evaluaciones': total_evaluaciones,
                        'suma_ponderaciones': suma_ponderaciones
                    }
                else:
                    context['estadisticas_estudiante'] = {
                        'evaluaciones_calificadas': evaluaciones_calificadas,
                        'total_evaluaciones': total_evaluaciones,
                        'suma_ponderaciones': suma_ponderaciones
                    }
            else:
                context['estadisticas_estudiante'] = {
                    'evaluaciones_calificadas': 0,
                    'total_evaluaciones': total_evaluaciones,
                    'suma_ponderaciones': 0
                }
            
            # Calcular promedios por tipo de evaluación
            promedios_por_tipo = {}
            for calificacion in calificaciones_con_nota:
                tipo = calificacion.evaluacion.get_tipo_display()
                if tipo not in promedios_por_tipo:
                    promedios_por_tipo[tipo] = {
                        'notas': [],
                        'ponderaciones': []
                    }
                promedios_por_tipo[tipo]['notas'].append(calificacion.nota)
                promedios_por_tipo[tipo]['ponderaciones'].append(calificacion.evaluacion.ponderacion)
            
            # Calcular promedios
            for tipo, datos in promedios_por_tipo.items():
                if datos['notas']:
                    promedio = sum(datos['notas']) / len(datos['notas'])
                    ponderacion_promedio = sum(datos['ponderaciones']) / len(datos['ponderaciones'])
                    promedios_por_tipo[tipo] = {
                        'promedio': promedio,
                        'ponderacion_promedio': ponderacion_promedio
                    }
            
            context['promedios_por_tipo'] = promedios_por_tipo
        
        # Renderizar solo el contenido de calificaciones
        html = render_to_string('pages/plataforma_calificaciones_content.html', context, request=request)
        
        # Debug: Verificar el contenido del HTML
        logger.debug(f"DEBUG: Longitud del HTML renderizado: {len(html)}")
        
        
        return JsonResponse({'html': html})
        
    except Curso.DoesNotExist:
        return JsonResponse({'error': 'El curso no existe'}, status=404)

@login_required
def plataforma_calificaciones_spa(request, curso_id):
    """
    Vista SPA para calificaciones que extiende de plataforma_aprendizaje.html
    """
    curso = get_object_or_404(Curso, id=curso_id)
    
    # Verificar que el usuario esté inscrito en el curso
    if not request.user.cursos.filter(id=curso_id).exists():
        messages.error(request, 'No tienes acceso a este curso.')
        return redirect('user_space')
    
    context = {
        'curso': curso,
        'user': request.user,
    }
    
    return render(request, 'pages/plataforma_calificaciones_spa.html', context)

def clear_messages(request):
    """
    Vista para limpiar manualmente todos los mensajes de la sesión
    """
    from django.contrib import messages
    from django.http import JsonResponse
    
    if request.method == 'POST':
        storage = messages.get_messages(request)
        storage.used = True  # Marcar todos los mensajes como usados
        return JsonResponse({'status': 'success', 'message': 'Mensajes limpiados'})
    else:
        return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)

@login_required
def editar_evaluacion_ajax(request, curso_id, evaluacion_id):
    """
    Vista AJAX para editar evaluaciones
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        curso = Curso.objects.get(id=curso_id, activo=True)
        evaluacion = Evaluacion.objects.get(id=evaluacion_id, curso=curso)
        
        # Verificar permisos
        if not request.user.is_staff:
            return JsonResponse({'error': 'No tienes permisos para editar evaluaciones'}, status=403)
        
        form = EvaluacionForm(request.POST, instance=evaluacion, curso=curso)
        if form.is_valid():
            evaluacion = form.save(commit=False)
            evaluacion.curso = curso
            evaluacion.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Evaluación "{evaluacion.nombre}" actualizada exitosamente.'
            })
        else:
            return JsonResponse({
                'success': False,
                'error': 'Error en el formulario',
                'form_errors': form.errors
            })
            
    except Curso.DoesNotExist:
        return JsonResponse({'error': 'Curso no encontrado'}, status=404)
    except Evaluacion.DoesNotExist:
        return JsonResponse({'error': 'Evaluación no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Error al actualizar la evaluación: {str(e)}'}, status=500)

@login_required
def eliminar_evaluacion_ajax(request, curso_id):
    """
    Vista AJAX para eliminar evaluaciones
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        curso = Curso.objects.get(id=curso_id, activo=True)
        evaluacion_id = request.GET.get('evaluacion_id')
        
        if not evaluacion_id:
            return JsonResponse({'error': 'ID de evaluación requerido'}, status=400)
        
        evaluacion = Evaluacion.objects.get(id=evaluacion_id, curso=curso)
        
        # Verificar permisos
        if not request.user.is_staff:
            return JsonResponse({'error': 'No tienes permisos para eliminar evaluaciones'}, status=403)
        
        # Verificar si la evaluación tiene calificaciones
        tiene_calificaciones = evaluacion.calificaciones.exists()
        
        # Si no tiene calificaciones, eliminar directamente
        if not tiene_calificaciones:
            nombre_evaluacion = evaluacion.nombre
            evaluacion.delete()
            return JsonResponse({
                'success': True,
                'message': f'Evaluación "{nombre_evaluacion}" eliminada exitosamente.'
            })
        
        # Si tiene calificaciones, verificar confirmación
        confirmacion = request.POST.get('confirmacion', '').strip()
        if confirmacion != 'ELIMINAR':
            return JsonResponse({
                'success': False,
                'error': 'Debes escribir "ELIMINAR" para confirmar la eliminación.'
            })
        
        # Eliminar la evaluación y todas sus calificaciones asociadas
        nombre_evaluacion = evaluacion.nombre
        evaluacion.delete()
        return JsonResponse({
            'success': True,
            'message': f'Evaluación "{nombre_evaluacion}" eliminada exitosamente.'
        })
        
    except Curso.DoesNotExist:
        return JsonResponse({'error': 'Curso no encontrado'}, status=404)
    except Evaluacion.DoesNotExist:
        return JsonResponse({'error': 'Evaluación no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Error al eliminar la evaluación: {str(e)}'}, status=500)

@login_required
def crear_evaluacion_ajax(request, curso_id):
    """
    Vista AJAX para cargar el formulario de crear evaluación dinámicamente
    """
    try:
        logger.debug(f"DEBUG: Iniciando crear_evaluacion_ajax")
        logger.debug(f"DEBUG: curso_id: {curso_id}")
        logger.debug(f"DEBUG: usuario: {request.user.username}")
        logger.debug(f"DEBUG: es_staff: {request.user.is_staff}")
        
        curso = Curso.objects.get(id=curso_id, activo=True)
        logger.debug(f"DEBUG: Curso encontrado: {curso.nombre}")
        
        # Verificar que el usuario esté inscrito en el curso
        if curso not in request.user.cursos.all():
            return JsonResponse({'error': 'No tienes acceso a este curso'}, status=403)
        
        # Verificar permisos de staff
        if not request.user.is_staff:
            return JsonResponse({'error': 'No tienes permisos para crear evaluaciones'}, status=403)
        
        # Verificar si es una acción específica
        action = request.GET.get('action')
        
        if action == 'crear_evaluacion':
            # Cargar formulario de crear evaluación
            from .forms import EvaluacionForm
            
            if request.method == 'POST':
                form = EvaluacionForm(request.POST, curso=curso)
                if form.is_valid():
                    evaluacion = form.save(commit=False)
                    evaluacion.curso = curso
                    evaluacion.creado_por = request.user
                    evaluacion.save()
                    
                    # Retornar respuesta de éxito
                    return JsonResponse({
                        'success': True,
                        'message': f'Evaluación "{evaluacion.nombre}" creada exitosamente.',
                        'redirect_url': reverse('plataforma_calificaciones', kwargs={'curso_id': curso_id})
                    })
                else:
                    # Formulario con errores
                    context = {
                        'curso': curso,
                        'form': form,
                        'user': request.user,
                    }
                    html = render_to_string('pages/plataforma_calificaciones_crear_evaluacion_content.html', context, request=request)
                    return JsonResponse({'html': html})
            else:
                # Formulario inicial
                form = EvaluacionForm(curso=curso)
                context = {
                    'curso': curso,
                    'form': form,
                    'user': request.user,
                }
                html = render_to_string('pages/plataforma_calificaciones_crear_evaluacion_content.html', context, request=request)
                return JsonResponse({'html': html})
        
        else:
            # Acción por defecto: mostrar formulario
            from .forms import EvaluacionForm
            form = EvaluacionForm(curso=curso)
            context = {
                'curso': curso,
                'form': form,
                'user': request.user,
            }
            html = render_to_string('pages/plataforma_calificaciones_crear_evaluacion_content.html', context, request=request)
            return JsonResponse({'html': html})
            
    except Curso.DoesNotExist:
        return JsonResponse({'error': 'Curso no encontrado'}, status=404)
    except Exception as e:
        logger.error(f"Error en crear_evaluacion_ajax: {str(e)}")
        return JsonResponse({'error': f'Error interno del servidor: {str(e)}'}, status=500)
